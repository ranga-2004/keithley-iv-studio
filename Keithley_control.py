"""
Keithley 2450 — Lab View
Professional measurement application for I-V characterisation and FET analysis.
Author : Rangaraajan Muralidaran
Version: 4.0  —  Cycling, bias stress, parameter extraction,
                  measurement quality (averaging / settling / pre-check),
                  ghost overlay, mean±σ band, cycle CSV export.
"""

import pyvisa
import time
import csv
import os
import copy
import json
import numpy as np
import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.ticker as mticker
import matplotlib.cm as _cm          # cycle gradient colours

# ═══════════════════════════════════════════════════════════════════════════════
#  DESIGN TOKENS  —  Light (default) and Dark themes
# ═══════════════════════════════════════════════════════════════════════════════

LIGHT = dict(
    BG      = "#ECEEF2",
    PANEL   = "#FFFFFF",
    PANEL2  = "#F4F6F9",
    PANEL3  = "#EBF0F7",
    BORDER  = "#CBD2DC",
    BORDER2 = "#B0BACC",
    TEXT    = "#111827",
    TEXT2   = "#4B5563",
    TEXT3   = "#9CA3AF",
    ACCENT  = "#1E3A5F",
    ACCENT2 = "#0369A1",
    ACCENT3 = "#E8F0FB",
    SUCCESS = "#166534",
    SUCCESS2= "#DCFCE7",
    DANGER  = "#991B1B",
    DANGER2 = "#FEE2E2",
    WARN    = "#92400E",
    WARN2   = "#FEF3C7",
    SEP     = "#E5E9EF",
    HOVER   = "#F0F4FF",
    PLOT_BG = "#FAFBFD",
    PLOT_GRID = "#E5E9EF",
)

DARK = dict(
    BG      = "#0D1117",
    PANEL   = "#161B22",
    PANEL2  = "#1C2431",
    PANEL3  = "#1A2840",
    BORDER  = "#30363D",
    BORDER2 = "#484F58",
    TEXT    = "#E6EDF3",
    TEXT2   = "#8B949E",
    TEXT3   = "#484F58",
    ACCENT  = "#1F6FEB",
    ACCENT2 = "#58A6FF",
    ACCENT3 = "#0D2044",
    SUCCESS = "#3FB950",
    SUCCESS2= "#0A2E12",
    DANGER  = "#FF7B72",
    DANGER2 = "#3D0C0C",
    WARN    = "#E3B341",
    WARN2   = "#2E1E00",
    SEP     = "#21262D",
    HOVER   = "#1F2D3D",
    PLOT_BG = "#0D1117",
    PLOT_GRID = "#21262D",
)

_THEME = LIGHT.copy()

def _t(key):
    return _THEME[key]

def _sync_globals():
    import sys
    m = sys.modules[__name__]
    for k, v in _THEME.items():
        setattr(m, k, v)

BG      = _THEME["BG"]
PANEL   = _THEME["PANEL"]
PANEL2  = _THEME["PANEL2"]
PANEL3  = _THEME["PANEL3"]
BORDER  = _THEME["BORDER"]
BORDER2 = _THEME["BORDER2"]
TEXT    = _THEME["TEXT"]
TEXT2   = _THEME["TEXT2"]
TEXT3   = _THEME["TEXT3"]
ACCENT  = _THEME["ACCENT"]
ACCENT2 = _THEME["ACCENT2"]
ACCENT3 = _THEME["ACCENT3"]
SUCCESS = _THEME["SUCCESS"]
SUCCESS2= _THEME["SUCCESS2"]
DANGER  = _THEME["DANGER"]
DANGER2 = _THEME["DANGER2"]
WARN    = _THEME["WARN"]
WARN2   = _THEME["WARN2"]
SEP     = _THEME["SEP"]
HOVER   = _THEME["HOVER"]

Y1_PAL  = ["#1D4ED8","#0E7490","#15803D","#7C3AED","#B45309"]
Y2_PAL  = ["#BE185D","#9333EA","#C2410C","#0F766E","#374151"]
FET_PAL = ["#1D4ED8","#DC2626","#16A34A","#9333EA","#D97706",
           "#0E7490","#65A30D","#BE185D","#0369A1","#374151"]

FONT_H1   = ("Segoe UI", 14, "bold")
FONT_H2   = ("Segoe UI", 11, "bold")
FONT_H3   = ("Segoe UI", 10, "bold")
FONT_BODY = ("Segoe UI",  9)
FONT_SM   = ("Segoe UI",  8)
FONT_MONO = ("Consolas",  9, "bold")
FONT_MONO2= ("Consolas", 10, "bold")

TSP_SAFE_ABORT = (
    "pcall(function() "
    "  local s = trigger.model.state() "
    "  if s == trigger.STATE_RUNNING or "
    "     s == trigger.STATE_WAITING  or "
    "     s == trigger.STATE_PAUSED   then "
    "    trigger.model.abort() "
    "  end "
    "end)"
)

# ═══════════════════════════════════════════════════════════════════════════════
#  REUSABLE UI COMPONENTS
# ═══════════════════════════════════════════════════════════════════════════════

class Card(tk.Frame):
    def __init__(self, parent, title="", accent=ACCENT, **kw):
        kw.setdefault("bg", _THEME["PANEL"])
        kw.setdefault("highlightbackground", _THEME["BORDER"])
        kw.setdefault("highlightthickness", 1)
        super().__init__(parent, **kw)
        if title:
            hdr = tk.Frame(self, bg=accent)
            hdr.pack(fill="x")
            tk.Label(hdr, text=f"  {title}", bg=accent, fg="white",
                     font=FONT_H3, anchor="w").pack(fill="x", padx=2, pady=5)
        self.body = tk.Frame(self, bg=_THEME["PANEL"])
        self.body.pack(fill="x", padx=8, pady=(6, 8))


class Section(tk.Frame):
    def __init__(self, parent, title="", open_=True, **kw):
        kw.setdefault("bg", BG)
        super().__init__(parent, **kw)
        self._open = open_

        hdr = tk.Frame(self, bg=_THEME["PANEL2"], cursor="hand2",
                       highlightbackground=_THEME["BORDER"], highlightthickness=1)
        hdr.pack(fill="x")
        tk.Frame(hdr, bg=_THEME["ACCENT2"], width=3).pack(side="left", fill="y")
        self._arr = tk.Label(hdr, text="▾" if open_ else "▸",
                             bg=_THEME["PANEL2"], fg=_THEME["ACCENT2"], font=("Segoe UI", 10))
        self._arr.pack(side="left", padx=(6, 2), pady=3)
        tk.Label(hdr, text=title, bg=_THEME["PANEL2"], fg=_THEME["ACCENT2"],
                 font=FONT_H3).pack(side="left", pady=3)
        hdr.bind("<Button-1>", self._toggle)
        for w in hdr.winfo_children():
            w.bind("<Button-1>", self._toggle)

        self.body = tk.Frame(self, bg=_THEME["PANEL"],
                             highlightbackground=_THEME["BORDER"], highlightthickness=1)
        if open_:
            self.body.pack(fill="x", pady=(0, 2))

    def _toggle(self, _=None):
        if self._open:
            self.body.pack_forget()
            self._arr.config(text="▸")
        else:
            self.body.pack(fill="x", pady=(0, 2))
            self._arr.config(text="▾")
        self._open = not self._open


class StatusBar(tk.Frame):
    STATES = {
        "ready":    ("#22C55E", ""),
        "running":  ("#F59E0B", "●  Running…"),
        "stopping": ("#EF4444", "●  Stopping…"),
        "done":     ("#22C55E", "●  Complete"),
        "error":    ("#EF4444", "●  Error"),
        "stopped":  ("#EF4444", "●  Stopped"),
    }

    def __init__(self, parent, **kw):
        kw.setdefault("bg", _THEME["PANEL"])
        kw.setdefault("highlightbackground", _THEME["BORDER"])
        kw.setdefault("highlightthickness", 1)
        super().__init__(parent, height=28, **kw)
        self.pack_propagate(False)
        self._lbl = tk.Label(self, text="", bg=_THEME["PANEL"],
                             fg=_THEME["TEXT2"], font=FONT_MONO, anchor="w")
        self._lbl.pack(side="left", padx=12, pady=3)
        self._extra = tk.Label(self, text="", bg=_THEME["PANEL"], fg=_THEME["TEXT2"],
                               font=FONT_BODY, anchor="w")
        self._extra.pack(side="left", padx=4)
        tk.Label(self, text="Keithley 2450 - Lab View  ·  Rangaraajan Muralidaran",
                 bg=_THEME["PANEL"], fg=_THEME["TEXT3"], font=FONT_SM).pack(side="right", padx=12)

    def set(self, state, extra=""):
        col, txt = self.STATES.get(state, ("#94A3B8", state))
        self._lbl.config(text=txt, fg=col)
        self._extra.config(text=extra)

    def msg(self, text, col=TEXT2):
        self._lbl.config(text=f"●  {text}", fg=col)
        self._extra.config(text="")


class ProgressRow(tk.Frame):
    def __init__(self, parent, **kw):
        kw.setdefault("bg", _THEME["PANEL"])
        super().__init__(parent, **kw)
        self._bar = ttk.Progressbar(self, orient="horizontal",
                                    mode="determinate", maximum=100,
                                    style="App.Horizontal.TProgressbar")
        self._bar.pack(side="left", fill="x", expand=True, ipady=2)
        self._pct = tk.Label(self, text="", bg=_THEME["PANEL"], fg=_THEME["ACCENT2"],
                             font=FONT_MONO, width=6)
        self._pct.pack(side="left", padx=(8, 0))
        self._timer = tk.Label(self, text="", bg=_THEME["PANEL"], fg=_THEME["TEXT3"],
                               font=FONT_SM, width=8)
        self._timer.pack(side="left", padx=(4, 0))
        self._t0 = None
        self._after_id = None

    def start(self):
        self._t0 = time.time()
        self._bar["value"] = 0
        self._pct.config(text="0 %")
        self._tick()

    def _tick(self):
        if self._t0 is None:
            return
        elapsed = int(time.time() - self._t0)
        self._timer.config(text=f"{elapsed//60:02d}:{elapsed%60:02d}")
        self._after_id = self.after(1000, self._tick)

    def update(self, pct):
        self._bar["value"] = pct
        self._pct.config(text=f"{pct} %")

    def stop(self, success=True):
        if self._after_id:
            self.after_cancel(self._after_id)
            self._after_id = None
        if success:
            self._bar["value"] = 100
            self._pct.config(text="100 %")
        else:
            self._pct.config(text="—")
        self._t0 = None

    def reset(self):
        self.stop(False)
        self._bar["value"] = 0
        self._timer.config(text="")


def _chk(parent, text, var, **kw):
    kw.setdefault("bg", _THEME["PANEL"])
    return tk.Checkbutton(parent, text=text, variable=var,
                          bg=kw.pop("bg"), activebackground=_THEME["PANEL"],
                          selectcolor=_THEME["PANEL"], relief="flat",
                          highlightthickness=0, font=FONT_BODY,
                          fg=_THEME["TEXT"], **kw)


def _row(parent, label, width=18):
    f = tk.Frame(parent, bg=_THEME["PANEL"])
    f.pack(fill="x", pady=2)
    lbl = tk.Label(f, text=label, bg=_THEME["PANEL"], fg=_THEME["TEXT2"], font=FONT_BODY,
                   width=width, anchor="w")
    lbl.pack(side="left")
    return f


def _entry(parent, default="", width=12, key=None, store=None):
    e = ttk.Entry(parent, width=width, font=FONT_BODY)
    e.insert(0, default)
    e.pack(side="right", fill="x", expand=True)
    if key and store is not None:
        store[key] = e
    return e


def _combo(parent, values, default, width=12, key=None, store=None, state="readonly"):
    c = ttk.Combobox(parent, values=values, state=state, width=width,
                     font=FONT_BODY)
    c.set(default)
    c.pack(side="right", fill="x", expand=True)
    if key and store is not None:
        store[key] = c
    return c


def _param(parent, label, default=None, key=None, store=None,
           combo_vals=None, combo_key=None, lbl_w=18):
    f = _row(parent, label, lbl_w)
    if combo_vals:
        w = _combo(f, combo_vals, combo_vals[0],
                   key=combo_key, store=store)
    else:
        w = _entry(f, default or "", key=key, store=store)
    return w


# ═══════════════════════════════════════════════════════════════════════════════
#  STYLE SETUP
# ═══════════════════════════════════════════════════════════════════════════════

def setup_styles():
    p  = _THEME["PANEL"]
    p2 = _THEME["PANEL2"]
    p3 = _THEME["PANEL3"]
    t  = _THEME["TEXT"]
    t2 = _THEME["TEXT2"]
    bd = _THEME["BORDER"]
    bd2= _THEME["BORDER2"]
    a2 = _THEME["ACCENT2"]
    a3 = _THEME["ACCENT3"]
    sep= _THEME["SEP"]

    s = ttk.Style()
    s.theme_use("clam")

    s.configure("TFrame",        background=p)
    s.configure("TLabel",        background=p, foreground=t,  font=FONT_BODY)
    s.configure("TCheckbutton",  background=p, foreground=t,  font=FONT_BODY)
    s.configure("TRadiobutton",  background=p, foreground=t,  font=FONT_BODY)

    s.configure("TEntry",
                fieldbackground=p, foreground=t,
                insertcolor=a2, bordercolor=bd,
                lightcolor=bd, darkcolor=bd, font=FONT_BODY,
                padding=(4, 3))
    s.map("TEntry",
          bordercolor=[("focus", a2)],
          fieldbackground=[("disabled", p2)])

    s.configure("TCombobox",
                fieldbackground=p, background=p2,
                foreground=t, arrowcolor=a2,
                bordercolor=bd, font=FONT_BODY, padding=(4, 3))
    s.map("TCombobox",
          fieldbackground=[("readonly", p), ("disabled", p2)],
          bordercolor=[("focus", a2)])

    s.configure("TButton",
                background=p2, foreground=t,
                bordercolor=bd2, font=FONT_BODY, padding=(8, 4),
                relief="flat")
    s.map("TButton",
          background=[("active", a3), ("pressed", a3)],
          bordercolor=[("active", a2)])

    for name, bg, abg, fg in [
        ("Run.TButton",    "#16A34A", "#15803D", "white"),
        ("Stop.TButton",   "#DC2626", "#B91C1C", "white"),
        ("Accent.TButton", a2,        "#0284C7", "white"),
    ]:
        s.configure(name, background=bg, foreground=fg,
                    font=FONT_H3, padding=(10, 6), relief="flat")
        s.map(name, background=[("active", abg), ("pressed", abg)])

    s.configure("App.Horizontal.TProgressbar",
                troughcolor=p2, background=a2,
                bordercolor=bd, thickness=8, lightcolor=a2,
                darkcolor=a2)

    s.configure("Treeview",
                background=p, foreground=t,
                fieldbackground=p, rowheight=24, font=FONT_BODY,
                bordercolor=bd)
    s.configure("Treeview.Heading",
                background=p2, foreground=_THEME["ACCENT"], font=FONT_H3,
                bordercolor=bd, relief="flat")
    s.map("Treeview",
          background=[("selected", a3)],
          foreground=[("selected", _THEME["ACCENT2"])])

    s.configure("TSeparator", background=sep)
    s.configure("TScrollbar",
                background=p2, troughcolor=p,
                bordercolor=bd, arrowcolor=t2)


# ═══════════════════════════════════════════════════════════════════════════════
#  LIST VALUE EDITOR
# ═══════════════════════════════════════════════════════════════════════════════

class ListEditor(tk.Toplevel):
    def __init__(self, parent, title, initial=None, accent=ACCENT, on_save=None):
        super().__init__(parent)
        self.title(title)
        self.geometry("360x460")
        self.configure(bg=_THEME["PANEL"])
        self.resizable(False, True)
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.grab_set()
        self._on_save = on_save
        self._accent  = accent

        hdr = tk.Frame(self, bg=accent)
        hdr.pack(fill="x")
        tk.Label(hdr, text=f"  {title}", bg=accent, fg="white",
                 font=FONT_H3).pack(side="left", padx=8, pady=7)

        body = tk.Frame(self, bg=_THEME["PANEL"]); body.pack(fill="both", expand=True, padx=8, pady=8)

        col_hdr = tk.Frame(body, bg=accent); col_hdr.pack(fill="x")
        for txt, w, anch in [("Index", 60, "center"), ("Value", 240, "center")]:
            tk.Label(col_hdr, text=txt, width=w//8, bg=accent, fg="white",
                     font=FONT_SM, anchor=anch).pack(side="left", ipadx=6, ipady=4,
                                                     fill="x" if txt=="Value" else None,
                                                     expand=(txt=="Value"))

        tv_f = tk.Frame(body, bg=_THEME["PANEL"]); tv_f.pack(fill="both", expand=True)
        vsb  = ttk.Scrollbar(tv_f, orient="vertical")
        self.tv = ttk.Treeview(tv_f, columns=("idx","val"), show="",
                               height=12, yscrollcommand=vsb.set,
                               selectmode="browse")
        vsb.config(command=self.tv.yview)
        self.tv.column("idx", width=60,  anchor="center", stretch=False)
        self.tv.column("val", width=240, anchor="center", stretch=True)
        self.tv.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self.tv.tag_configure("odd",  background=_THEME["PANEL"])
        self.tv.tag_configure("even", background=_THEME["PANEL2"])
        self.tv.bind("<Double-1>", self._inline_edit)

        hint = tk.Label(body, text="Double-click a value to edit",
                        bg=_THEME["PANEL"], fg=_THEME["TEXT3"], font=FONT_SM)
        hint.pack(anchor="w", pady=(2, 0))

        btn_f = tk.Frame(self, bg=_THEME["PANEL2"],
                         highlightbackground=_THEME["BORDER"], highlightthickness=1)
        btn_f.pack(fill="x")
        btn_left  = tk.Frame(btn_f, bg=_THEME["PANEL2"]); btn_left.pack(side="left",  padx=8, pady=6)
        btn_right = tk.Frame(btn_f, bg=_THEME["PANEL2"]); btn_right.pack(side="right", padx=8, pady=6)

        ttk.Button(btn_left,  text="+ Add Row",     command=self._add).pack(side="left", padx=(0,4))
        ttk.Button(btn_left,  text="✕ Delete",      command=self._delete).pack(side="left", padx=(0,4))
        ttk.Button(btn_left,  text="⬆ Import CSV",  command=self._import).pack(side="left")
        ttk.Button(btn_right, text="Cancel",         command=self.destroy).pack(side="right", padx=(4,0))
        ttk.Button(btn_right, text="Save & Close",   style="Accent.TButton",
                   command=self._save).pack(side="right")

        data = initial or []
        if data:
            for i, v in enumerate(data, 1):
                self.tv.insert("","end", values=(i, f"{v:g}"),
                               tags=("odd" if i%2 else "even",))
        else:
            for i in range(1, 6):
                self.tv.insert("","end", values=(i,""),
                               tags=("odd" if i%2 else "even",))

    def _inline_edit(self, event):
        if self.tv.identify("region", event.x, event.y) != "cell": return
        if self.tv.identify_column(event.x) != "#2": return
        iid = self.tv.identify_row(event.y)
        if not iid: return
        x, y, w, h = self.tv.bbox(iid, "#2")
        cur = self.tv.item(iid, "values")[1]
        ed = tk.Entry(self.tv, font=FONT_BODY, bg=_THEME["PANEL"], fg=_THEME["TEXT"],
                      insertbackground=_THEME["ACCENT2"], relief="solid",
                      bd=1, highlightthickness=1,
                      highlightcolor=ACCENT2, highlightbackground=_THEME["BORDER"])
        ed.insert(0, cur); ed.select_range(0, tk.END)
        ed.place(x=x, y=y, width=w, height=h); ed.focus_set()
        def commit(_=None):
            idx = self.tv.item(iid, "values")[0]
            tag = "odd" if int(idx)%2 else "even"
            self.tv.item(iid, values=(idx, ed.get().strip()), tags=(tag,))
            ed.destroy()
        ed.bind("<Return>", commit); ed.bind("<Tab>", commit)
        ed.bind("<Escape>", lambda _: ed.destroy())
        ed.bind("<FocusOut>", commit)

    def _add(self):
        n = len(self.tv.get_children()) + 1
        self.tv.insert("","end", values=(n,""),
                       tags=("odd" if n%2 else "even",))

    def _delete(self):
        sel = self.tv.selection()
        if not sel: return
        self.tv.delete(sel[0])
        for i, iid in enumerate(self.tv.get_children(), 1):
            cur = self.tv.item(iid,"values")[1]
            self.tv.item(iid, values=(i, cur),
                         tags=("odd" if i%2 else "even",))

    def _import(self):
        path = filedialog.askopenfilename(
            parent=self, title="Import Values",
            filetypes=[("CSV","*.csv"),("Excel","*.xlsx *.xls"),("All","*.*")])
        if not path: return
        try:
            vals = []; ext = path.rsplit(".", 1)[-1].lower()
            if ext == "csv":
                with open(path, newline="", encoding="utf-8-sig") as f:
                    for row in csv.reader(f):
                        for cell in row:
                            try: vals.append(float(cell.strip())); break
                            except ValueError: continue
            elif ext in ("xlsx","xls"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    for row in wb.active.iter_rows(values_only=True):
                        for cell in row:
                            try: vals.append(float(cell)); break
                            except: continue
                except ImportError:
                    messagebox.showerror("Missing Package",
                                         "openpyxl not installed.\npip install openpyxl",
                                         parent=self); return
            if not vals:
                messagebox.showwarning("No Data","No numeric values found.", parent=self)
                return
            for iid in self.tv.get_children(): self.tv.delete(iid)
            for i, v in enumerate(vals, 1):
                self.tv.insert("","end", values=(i, f"{v:g}"),
                               tags=("odd" if i%2 else "even",))
        except Exception as e:
            messagebox.showerror("Import Error", str(e), parent=self)

    def _save(self):
        vals = []
        for iid in self.tv.get_children():
            try: vals.append(float(self.tv.item(iid,"values")[1]))
            except: pass
        if self._on_save:
            self._on_save(vals)
        self.destroy()

    def get_values(self):
        vals = []
        for iid in self.tv.get_children():
            try: vals.append(float(self.tv.item(iid,"values")[1]))
            except: pass
        return vals


# ═══════════════════════════════════════════════════════════════════════════════
#  PLOT HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _style_plot(fig, ax, ax2=None, title=""):
    bg      = _THEME["PANEL"]
    plot_bg = _THEME["PLOT_BG"]
    border  = _THEME["BORDER"]
    text    = _THEME["TEXT"]
    text2   = _THEME["TEXT2"]
    sep     = _THEME["PLOT_GRID"]
    is_dark = (bg == DARK["PANEL"] or plot_bg == DARK["PLOT_BG"])
    fig.patch.set_facecolor(bg)
    ax.set_facecolor(plot_bg)
    spine_col = "#30363D" if is_dark else border
    spine_lw  = 0.5 if is_dark else 0.8
    for sp in ax.spines.values():
        sp.set_color(spine_col); sp.set_linewidth(spine_lw)
    ax.tick_params(colors=text2, labelsize=8, width=0.6)
    ax.grid(True, linestyle="--", color=sep, linewidth=0.7,
            alpha=0.6 if is_dark else 0.9)
    ax.set_title(title, color=text, fontsize=10, fontweight="bold", pad=8)
    if ax2:
        ax2.set_facecolor("none")
        for sp in ax2.spines.values():
            sp.set_color(spine_col); sp.set_linewidth(spine_lw)
        ax2.tick_params(colors=text2, labelsize=8, width=0.6)


def _eng_formatter(ax, axis="y"):
    fmt = mticker.EngFormatter(sep="")
    if axis == "y":
        ax.yaxis.set_major_formatter(fmt)
    else:
        ax.xaxis.set_major_formatter(fmt)


def _eng_str(val):
    try:
        return mticker.EngFormatter(sep="")(val)
    except Exception:
        return str(val)


# ═══════════════════════════════════════════════════════════════════════════════
#  PERSISTENT CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

class _Config:
    _DEFAULTS = {
        "visa_address": "USB0::0x05E6::0x2450::04465297::INSTR",
        "save_dir":     r"D:\KEITHLEY 2450\Data",
        "sample_id":    "DUT_001",
        "dark_mode":    True,
    }

    def __init__(self):
        script_dir  = os.path.dirname(os.path.abspath(__file__))
        self._path  = os.path.join(script_dir, "keithley_config.json")
        self._data  = dict(self._DEFAULTS)
        self._load()

    def _load(self):
        try:
            with open(self._path, "r", encoding="utf-8") as f:
                saved = json.load(f)
            for k in self._DEFAULTS:
                if k in saved:
                    self._data[k] = saved[k]
        except (FileNotFoundError, json.JSONDecodeError, OSError):
            pass

    def save(self):
        try:
            with open(self._path, "w", encoding="utf-8") as f:
                json.dump(self._data, f, indent=2)
        except OSError:
            pass

    def get(self, key):
        return self._data.get(key, self._DEFAULTS.get(key))

    def set(self, key, value):
        self._data[key] = value
        self.save()


_cfg = _Config()


# ═══════════════════════════════════════════════════════════════════════════════
#  COMPARE MANAGER
# ═══════════════════════════════════════════════════════════════════════════════

class CompareManager:
    REF_PALETTE = [
        "#F97316","#A855F7","#EC4899","#14B8A6","#EAB308",
        "#EF4444","#8B5CF6","#06B6D4","#84CC16","#F43F5E",
    ]

    def __init__(self, owner, replot_cb, save_dir_cb, mode="IV"):
        self._owner     = owner
        self._replot    = replot_cb
        self._save_dir  = save_dir_cb
        self._mode      = mode
        self.refs       = []

    def _next_colour(self):
        used = {r["colour"] for r in self.refs}
        for c in self.REF_PALETTE:
            if c not in used: return c
        return self.REF_PALETTE[len(self.refs) % len(self.REF_PALETTE)]

    def _parse_csv(self, path):
        curves = []
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers, rows = [], []
            for row in reader:
                if not row: continue
                if row[0].startswith("#"): continue
                if not headers: headers = [h.strip() for h in row]
                else: rows.append(row)
        if not headers or not rows:
            raise ValueError("No data found in CSV.")

        cols = {h: [] for h in headers}
        for row in rows:
            for i, h in enumerate(headers):
                try: cols[h].append(float(row[i]) if i < len(row) else float("nan"))
                except ValueError: cols[h].append(float("nan"))

        if self._mode == "IV":
            v_col = next((h for h in headers if "Voltage" in h), None)
            if v_col is None: raise ValueError("No Voltage column found.")
            x = cols[v_col]
            i_cols = [h for h in headers if ("Current" in h or
                      h.endswith("(A)") or " I " in h)]
            if not i_cols: raise ValueError("No Current column found.")
            for ic in i_cols:
                curves.append({"x": x, "y": cols[ic], "col_name": ic})
        else:
            if len(headers) < 2: raise ValueError("Need at least 2 columns.")
            x_col = headers[0]
            x = cols[x_col]
            i_cols = [h for h in headers[1:] if h in cols]
            if not i_cols: raise ValueError("No data columns found.")
            for ic in i_cols:
                curves.append({"x": x, "y": cols[ic], "col_name": ic})
        return curves

    def add_files(self, paths=None):
        if paths is None:
            paths = filedialog.askopenfilenames(
                title="Add Reference CSV(s)",
                initialdir=self._save_dir(),
                filetypes=[("CSV files","*.csv"),("All files","*.*")],
                parent=self._owner)
            if not paths: return
        added = 0
        errors = []
        for path in paths:
            try:
                curves = self._parse_csv(path)
                name   = os.path.splitext(os.path.basename(path))[0]
                self.refs.append({
                    "path":    path,
                    "label":   name,
                    "visible": True,
                    "colour":  self._next_colour(),
                    "curves":  curves,
                })
                added += 1
            except Exception as e:
                errors.append(f"{os.path.basename(path)}: {e}")
        if errors:
            messagebox.showwarning("Compare — Load Errors",
                                   "\n".join(errors), parent=self._owner)
        if added:
            self._replot()

    def remove(self, idx):
        if 0 <= idx < len(self.refs):
            del self.refs[idx]
            self._replot()

    def clear(self):
        self.refs.clear()
        self._replot()

    def toggle_visible(self, idx):
        if 0 <= idx < len(self.refs):
            self.refs[idx]["visible"] = not self.refs[idx]["visible"]
            self._replot()

    def rename(self, idx, new_label):
        if 0 <= idx < len(self.refs):
            self.refs[idx]["label"] = new_label
            self._replot()

    def set_colour(self, idx, colour):
        if 0 <= idx < len(self.refs):
            self.refs[idx]["colour"] = colour
            self._replot()

    def draw(self, ax, x_ch=None):
        for ri, ref in enumerate(self.refs):
            if not ref["visible"]: continue
            col   = ref["colour"]
            label = ref["label"]
            curves = ref["curves"]
            for ci, c in enumerate(curves):
                xd = c["x"]; yd = c["y"]
                n  = min(len(xd), len(yd))
                if n == 0: continue
                lbl = (f"{label}  {c['col_name']}"
                       if len(curves) > 1 else label)
                ax.plot(xd[:n], yd[:n], "--", lw=1.8,
                        color=col, alpha=0.75, zorder=2,
                        label=lbl)

    def open_panel(self):
        ComparePanelWindow(self._owner, self)


class ComparePanelWindow(tk.Toplevel):
    def __init__(self, parent, mgr):
        super().__init__(parent)
        self.mgr = mgr
        self.title("Compare CSV Manager")
        self.configure(bg=_THEME["BG"])
        self.geometry("700x480")
        self.resizable(True, True)
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self._build()

    def _build(self):
        for w in self.winfo_children(): w.destroy()
        BG =_THEME["BG"]; PNL=_THEME["PANEL"]; P2=_THEME["PANEL2"]
        BD =_THEME["BORDER"]; TXT=_THEME["TEXT"]; T2=_THEME["TEXT2"]
        T3 =_THEME["TEXT3"]; A2=_THEME["ACCENT2"]; DNG=_THEME["DANGER"]

        tk.Frame(self, bg=A2, height=3).pack(fill="x")

        hdr = tk.Frame(self, bg=P2); hdr.pack(fill="x")
        tk.Label(hdr, text="  📊  Compare CSV Manager",
                 bg=P2, fg=TXT, font=("Segoe UI",12,"bold")).pack(
                 side="left", padx=14, pady=8)
        tk.Label(hdr, text="Load 1 or more older CSVs to overlay as dashed reference lines",
                 bg=P2, fg=T2, font=("Segoe UI",8)).pack(side="left", padx=4)

        brow = tk.Frame(self, bg=BG); brow.pack(fill="x", padx=10, pady=(8,4))
        tk.Button(brow, text="+ Add CSV(s)",
                  bg=_THEME["SUCCESS"], fg="white",
                  font=("Segoe UI",9,"bold"), relief="flat", bd=0,
                  padx=14, pady=5, cursor="hand2",
                  activebackground=_THEME["SUCCESS"],
                  command=self._add).pack(side="left", padx=(0,6))
        tk.Button(brow, text="✕ Clear All",
                  bg=P2, fg=DNG,
                  font=("Segoe UI",9), relief="flat", bd=0,
                  padx=10, pady=5, cursor="hand2",
                  activebackground=BG,
                  command=self._clear_all).pack(side="left", padx=(0,6))
        tk.Label(brow, text=f"Mode: {self.mgr._mode}  ·  "
                             "Eye = toggle visible  ·  Click label to rename  ·  ● = colour",
                 bg=BG, fg=T3, font=("Segoe UI",7)).pack(side="left", padx=12)
        tk.Button(brow, text="Close",
                  bg=P2, fg=TXT, font=("Segoe UI",9),
                  relief="flat", bd=0, padx=10, pady=5, cursor="hand2",
                  activebackground=BG, command=self.destroy).pack(side="right")

        tk.Frame(self, bg=_THEME["SEP"], height=1).pack(fill="x", padx=10)

        container = tk.Frame(self, bg=BG); container.pack(fill="both", expand=True, padx=10, pady=6)
        cvs = tk.Canvas(container, bg=BG, highlightthickness=0)
        vsb = ttk.Scrollbar(container, orient="vertical", command=cvs.yview)
        self._inner = tk.Frame(cvs, bg=BG)
        self._inner.bind("<Configure>",
                         lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.create_window((0,0), window=self._inner, anchor="nw")
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        cvs.bind("<MouseWheel>",
                 lambda e: cvs.yview_scroll(int(-e.delta/120), "units"))
        self._render_list()

        foot = tk.Frame(self, bg=P2); foot.pack(fill="x", padx=10, pady=(0,6))
        tk.Label(foot, text="Reference colours:", bg=P2, fg=T2,
                 font=("Segoe UI",8)).pack(side="left", padx=8, pady=4)
        for c in CompareManager.REF_PALETTE[:10]:
            tk.Frame(foot, bg=c, width=18, height=14).pack(
                side="left", padx=2, pady=6)

    def _render_list(self):
        for w in self._inner.winfo_children(): w.destroy()
        BG  = _THEME["BG"]; PNL=_THEME["PANEL"]; BD=_THEME["BORDER"]
        TXT = _THEME["TEXT"]; T2=_THEME["TEXT2"]; T3=_THEME["TEXT3"]

        if not self.mgr.refs:
            tk.Label(self._inner,
                     text="No reference CSVs loaded yet.\nClick  + Add CSV(s)  to get started.",
                     bg=BG, fg=T3, font=("Segoe UI",10),
                     justify="center").pack(expand=True, pady=40)
            return

        hrow = tk.Frame(self._inner, bg=PNL)
        hrow.pack(fill="x", pady=(0,2))
        for txt, w in [("", 5), ("Vis", 4), ("Label (click to rename)", 28),
                       ("File", 24), ("Curves", 7), ("Colour", 8), ("", 8)]:
            tk.Label(hrow, text=txt, bg=PNL, fg=T2,
                     font=("Segoe UI",8,"bold"),
                     width=w, anchor="w").pack(side="left", padx=2, pady=3)

        for idx, ref in enumerate(self.mgr.refs):
            row_bg = PNL if idx % 2 == 0 else BG
            row = tk.Frame(self._inner, bg=row_bg,
                           highlightbackground=BD, highlightthickness=1)
            row.pack(fill="x", pady=1)

            tk.Label(row, text=str(idx+1), bg=row_bg, fg=T3,
                     font=("Segoe UI",8), width=3, anchor="e").pack(
                     side="left", padx=(6,2), pady=4)

            eye = "👁" if ref["visible"] else "—"
            eye_col = _THEME["ACCENT2"] if ref["visible"] else T3
            tk.Button(row, text=eye, bg=row_bg, fg=eye_col,
                      font=("Segoe UI",9), relief="flat", bd=0,
                      padx=4, cursor="hand2",
                      command=lambda i=idx: self._toggle(i)).pack(
                      side="left", padx=2)

            lbl_var = tk.StringVar(value=ref["label"])
            lbl_entry = tk.Entry(row, textvariable=lbl_var, width=26,
                                  bg=row_bg, fg=TXT,
                                  insertbackground=_THEME["ACCENT2"],
                                  relief="flat", font=("Segoe UI",9),
                                  highlightthickness=1,
                                  highlightbackground=BD)
            lbl_entry.pack(side="left", padx=4, ipady=2)

            def _commit(event, i=idx, v=lbl_var):
                self.mgr.rename(i, v.get().strip() or f"Ref {i+1}")
                self._render_list()

            lbl_entry.bind("<Return>",   _commit)
            lbl_entry.bind("<FocusOut>", _commit)

            short = os.path.basename(ref["path"])
            if len(short) > 22: short = short[:19] + "…"
            tk.Label(row, text=short, bg=row_bg, fg=T2,
                     font=("Segoe UI",8), width=22, anchor="w").pack(
                     side="left", padx=4)

            nc = len(ref["curves"])
            tk.Label(row, text=f"{nc} col{'s' if nc!=1 else ''}",
                     bg=row_bg, fg=T3,
                     font=("Segoe UI",8), width=6).pack(side="left", padx=4)

            swatch = tk.Frame(row, bg=ref["colour"],
                               width=28, height=16, cursor="hand2")
            swatch.pack(side="left", padx=6)
            swatch.bind("<Button-1>", lambda e, i=idx: self._cycle_colour(i))

            tk.Button(row, text="✕",
                      bg=row_bg, fg=_THEME["DANGER"],
                      font=("Segoe UI",9), relief="flat", bd=0,
                      padx=6, cursor="hand2",
                      command=lambda i=idx: self._remove(i)).pack(
                      side="right", padx=6)

    def _add(self):
        self.mgr.add_files()
        self._render_list()

    def _toggle(self, idx):
        self.mgr.toggle_visible(idx)
        self._render_list()

    def _remove(self, idx):
        self.mgr.remove(idx)
        self._render_list()

    def _clear_all(self):
        if not self.mgr.refs: return
        if messagebox.askyesno("Clear All", "Remove all reference CSV overlays?",
                                parent=self):
            self.mgr.clear()
            self._render_list()

    def _cycle_colour(self, idx):
        ref = self.mgr.refs[idx]
        pal = CompareManager.REF_PALETTE
        try: next_idx = (pal.index(ref["colour"]) + 1) % len(pal)
        except ValueError: next_idx = 0
        self.mgr.set_colour(idx, pal[next_idx])
        self._render_list()


class KeithleyApp:
    CHANNELS = ["Voltage (V)", "Current (A)", "Time (s)",
                "Resistance (Ω)", "Power (W)"]

    def __init__(self, root):
        self.root = root
        self.root.title("Keithley 2450 - Lab View")
        self.root.configure(bg=_THEME["BG"])
        self.root.minsize(1280, 720)

        self.save_dir    = _cfg.get("save_dir")
        self.master_addr = _cfg.get("visa_address")
        self.rm          = pyvisa.ResourceManager()
        self.inst        = None
        self.is_running  = False
        self._stop_flag  = False
        self.inputs      = {}
        self.sweep_data  = []
        self._compare    = CompareManager(
            owner=None,
            replot_cb=lambda: self._replot() if hasattr(self, 'ax') else None,
            save_dir_cb=lambda: self.save_dir,
            mode="IV")
        self.ax_x        = None
        self.ax_y1       = {}
        self.ax_y2       = {}

        self.mq_avg      = tk.StringVar(value="1")
        self.mq_settle   = tk.BooleanVar(value=False)
        self.mq_tol      = tk.StringVar(value="1.0")
        self.mq_precheck = tk.BooleanVar(value=True)

        self._dark_mode  = bool(_cfg.get("dark_mode"))
        if self._dark_mode:
            _THEME.update(DARK)
        _sync_globals()

        self._cycle_runs    = []
        self._cycle_names   = []
        self._current_cycle = 0
        self._overlay_data  = None
        self.cyc_enabled    = tk.BooleanVar(value=False)
        self.cyc_n          = tk.StringVar(value="5")
        self.cyc_rest       = tk.StringVar(value="0")
        self.cyc_gradient   = tk.BooleanVar(value=True)
        self.cyc_show_all   = tk.BooleanVar(value=True)
        self.cyc_stats      = tk.BooleanVar(value=True)
        self.cyc_overlay    = tk.BooleanVar(value=True)
        self.cyc_trigger    = tk.StringVar(value="Auto")
        self._cyc_manual_go = threading.Event()
        self.cyc_custom_names = tk.BooleanVar(value=False)

        self.an_vth    = tk.BooleanVar(value=True)
        self.an_ss     = tk.BooleanVar(value=True)
        self.an_ionoff = tk.BooleanVar(value=True)
        self.an_hyst   = tk.BooleanVar(value=False)

        if not os.path.exists(self.save_dir):
            try: os.makedirs(self.save_dir)
            except OSError: pass

        setup_styles()
        self._build_ui()
        self._compare._owner = self.root
        self._start_conn_checker()

    # ── Instrument helpers ────────────────────────────────────────────────────

    def _safe_abort(self):
        try:
            self.inst.write(TSP_SAFE_ABORT)
            time.sleep(0.08)
        except Exception:
            pass

    def _force_stop(self):
        if not self.is_running:
            return
        self._stop_flag = True
        self.status.set("stopping")
        try:
            if self.inst:
                self.inst.write(TSP_SAFE_ABORT)
                self.inst.write("smu.source.output = smu.OFF")
        except Exception:
            pass

    def _set_conn_status(self, connected: bool, addr: str = ""):
        def _do():
            if not hasattr(self, "_conn_lbl"): return
            if connected:
                short = addr.split("::")[-2] if "::" in addr else addr[:18]
                self._conn_lbl.config(text=f"● SMU: {short}", fg="#4ADE80")
            else:
                self._conn_lbl.config(text="● SMU: Disconnected", fg="#EF4444")
        self.root.after(0, _do)

    def _start_conn_checker(self):
        def _ping():
            if not self.is_running:
                addr = self.master_addr.strip()
                try:
                    rm   = pyvisa.ResourceManager()
                    inst = rm.open_resource(addr)
                    inst.timeout = 1500
                    idn = inst.query("*IDN?").strip()
                    inst.close()
                    rm.close()
                    connected = bool(idn)
                except Exception:
                    connected = False
                self._set_conn_status(connected, addr if connected else "")
            if hasattr(self, "_conn_check_id"):
                self._conn_check_id = self.root.after(3000, _ping)
        self._conn_check_id = self.root.after(1000, _ping)

    def _stop_conn_checker(self):
        if hasattr(self, "_conn_check_id"):
            try: self.root.after_cancel(self._conn_check_id)
            except Exception: pass

    # ── Dark / Light theme ────────────────────────────────────────────────────

    def _toggle_dark(self):
        self._dark_mode = not self._dark_mode
        palette = DARK if self._dark_mode else LIGHT
        _THEME.update(palette)
        _sync_globals()
        _cfg.set("dark_mode", self._dark_mode)
        if self._dark_mode:
            self._dark_btn.config(text="🌙  Dark", bg="#1A2535",
                                  fg="#FCD34D", activebackground="#243040")
        else:
            self._dark_btn.config(text="☀  Light", bg="#243F6A",
                                  fg="#93C5FD", activebackground="#2A4A7F")
        self.root.configure(bg=_THEME["BG"])
        self._apply_theme(self.root)
        setup_styles()
        self.root.option_add("*TCombobox*Listbox.background",       _THEME["PANEL"])
        self.root.option_add("*TCombobox*Listbox.foreground",       _THEME["TEXT"])
        self.root.option_add("*TCombobox*Listbox.selectBackground", _THEME["ACCENT3"])
        self.root.option_add("*TCombobox*Listbox.selectForeground", _THEME["TEXT"])
        panel = _THEME["PANEL"]
        if hasattr(self, "fig"):
            self.fig.patch.set_facecolor(panel)
        if hasattr(self, "canvas"):
            self.canvas.get_tk_widget().config(bg=panel)
        for w in self.root.winfo_children():
            if isinstance(w, tk.Toplevel):
                w.configure(bg=_THEME["BG"])
                self._apply_theme(w)
                if hasattr(w, "fig"):
                    w.fig.patch.set_facecolor(panel)
                if hasattr(w, "canvas"):
                    w.canvas.get_tk_widget().config(bg=panel)
                    if hasattr(w, "_replot"):
                        if w.curves or w._cycle_data or w._overlay_data:
                            w._replot()
                        else:
                            w.fig.clear()
                            w.fig.patch.set_facecolor(panel)
                            w.ax = w.fig.add_subplot(111)
                            w.fig.subplots_adjust(left=0.11, right=0.95,
                                                  top=0.93, bottom=0.10)
                            _style_plot(w.fig, w.ax,
                                        title="Configure terminals and press RUN")
                            w.canvas.draw()
        if self.sweep_data or self._cycle_runs or self._overlay_data:
            self._replot()
        else:
            if hasattr(self, "fig") and hasattr(self, "ax"):
                self.fig.clear()
                self.fig.patch.set_facecolor(panel)
                self.ax  = self.fig.add_subplot(111)
                self.ax2 = self.ax.twinx()
                self.ax2.set_visible(False)
                self.fig.subplots_adjust(left=0.11, right=0.95,
                                         top=0.93, bottom=0.10)
                _style_plot(self.fig, self.ax, title="I-V Characteristic")
                if hasattr(self, "canvas"):
                    self.canvas.draw()

    def _save_fig_light(self, fig, path):
        was_dark = self._dark_mode
        if not was_dark:
            fig.savefig(path, dpi=300, facecolor=LIGHT["PANEL"],
                        bbox_inches="tight")
            return
        _THEME.update(LIGHT)
        _sync_globals()
        try:
            fig.patch.set_facecolor(LIGHT["PANEL"])
            for ax in fig.get_axes():
                ax.set_facecolor(LIGHT["PLOT_BG"])
                for sp in ax.spines.values():
                    sp.set_color(LIGHT["BORDER"])
                ax.tick_params(colors=LIGHT["TEXT2"])
                ax.xaxis.label.set_color(LIGHT["TEXT"])
                ax.yaxis.label.set_color(ax.yaxis.label.get_color())
                ax.title.set_color(LIGHT["TEXT"])
                ax.grid(True, linestyle="--", color=LIGHT["PLOT_GRID"],
                        linewidth=0.7, alpha=0.9)
                leg = ax.get_legend()
                if leg:
                    leg.get_frame().set_facecolor(LIGHT["PANEL"])
                    leg.get_frame().set_edgecolor(LIGHT["BORDER"])
                    for txt in leg.get_texts():
                        txt.set_color(LIGHT["TEXT"])
            fig.canvas.draw()
            fig.savefig(path, dpi=300, facecolor=LIGHT["PANEL"],
                        bbox_inches="tight")
        finally:
            _THEME.update(DARK)
            _sync_globals()
            fig.patch.set_facecolor(DARK["PANEL"])
            for ax in fig.get_axes():
                ax.set_facecolor(DARK["PLOT_BG"])
                for sp in ax.spines.values():
                    sp.set_color(DARK["BORDER"])
                ax.tick_params(colors=DARK["TEXT2"])
                ax.xaxis.label.set_color(DARK["TEXT"])
                ax.title.set_color(DARK["TEXT"])
                ax.grid(True, linestyle="--", color=DARK["PLOT_GRID"],
                        linewidth=0.7, alpha=0.6)
                leg = ax.get_legend()
                if leg:
                    leg.get_frame().set_facecolor(DARK["PANEL"])
                    leg.get_frame().set_edgecolor(DARK["BORDER"])
                    for txt in leg.get_texts():
                        txt.set_color(DARK["TEXT"])
            try:
                fig.canvas.draw_idle()
            except Exception:
                pass

    def _apply_theme(self, widget):
        p  = _THEME["PANEL"]
        p2 = _THEME["PANEL2"]
        p3 = _THEME["PANEL3"]
        bg = _THEME["BG"]
        t  = _THEME["TEXT"]
        t2 = _THEME["TEXT2"]
        t3 = _THEME["TEXT3"]
        a2 = _THEME["ACCENT2"]
        a3 = _THEME["ACCENT3"]
        bd = _THEME["BORDER"]
        sep= _THEME["SEP"]

        w_class = widget.winfo_class()

        try:
            cur_bg = widget.cget("bg")
        except tk.TclError:
            cur_bg = None

        bg_map = {
            LIGHT["BG"]:     bg,  DARK["BG"]:     bg,
            LIGHT["PANEL"]:  p,   DARK["PANEL"]:  p,
            LIGHT["PANEL2"]: p2,  DARK["PANEL2"]: p2,
            LIGHT["PANEL3"]: p3,  DARK["PANEL3"]: p3,
            LIGHT["SEP"]:    sep, DARK["SEP"]:    sep,
            LIGHT["HOVER"]:  _THEME["HOVER"], DARK["HOVER"]: _THEME["HOVER"],
            LIGHT["ACCENT3"]:a3,  DARK["ACCENT3"]:a3,
            "#F0FDF4": _THEME["PANEL"] if _THEME["BG"]==DARK["BG"] else "#F0FDF4",
            "#EFF6FF": _THEME["PANEL"] if _THEME["BG"]==DARK["BG"] else "#EFF6FF",
            "#FDF2F8": _THEME["PANEL"] if _THEME["BG"]==DARK["BG"] else "#FDF2F8",
            "#0A1F0F": _THEME["PANEL"] if _THEME["BG"]==LIGHT["BG"] else "#0A1F0F",
            "#0A1220": _THEME["PANEL"] if _THEME["BG"]==LIGHT["BG"] else "#0A1220",
            "#1A0A14": _THEME["PANEL"] if _THEME["BG"]==LIGHT["BG"] else "#1A0A14",
        }

        if cur_bg in bg_map and w_class not in ("Button",):
            try:
                widget.config(bg=bg_map[cur_bg])
            except tk.TclError:
                pass

        for attr, old_vals, new_val in [
            ("fg", [LIGHT["TEXT"],    DARK["TEXT"]],    t),
            ("fg", [LIGHT["TEXT2"],   DARK["TEXT2"]],   t2),
            ("fg", [LIGHT["TEXT3"],   DARK["TEXT3"]],   t3),
            ("fg", [LIGHT["ACCENT2"], DARK["ACCENT2"]], a2),
            ("highlightbackground",
             [LIGHT["BORDER"], DARK["BORDER"],
              LIGHT["BORDER2"],DARK["BORDER2"]], bd),
            ("highlightcolor",
             [LIGHT["ACCENT2"], DARK["ACCENT2"],
              LIGHT["BORDER"],  DARK["BORDER"]], a2),
        ]:
            try:
                cur = widget.cget(attr)
                if cur in old_vals:
                    widget.config(**{attr: new_val})
            except tk.TclError:
                pass

        if w_class in ("Frame", "Canvas", "Labelframe"):
            try:
                widget.config(highlightbackground=bd, highlightcolor=bd)
            except tk.TclError:
                pass

        if w_class in ("Checkbutton", "Radiobutton"):
            try:
                widget.config(selectcolor=p, activebackground=p,
                              activeforeground=t)
            except tk.TclError:
                pass

        if w_class == "Entry":
            try:
                widget.config(bg=p, fg=t, insertbackground=a2)
            except tk.TclError:
                pass

        if w_class == "Text":
            try:
                widget.config(bg=p, fg=t, insertbackground=a2,
                              highlightbackground=bd, highlightcolor=a2)
            except tk.TclError:
                pass

        if w_class == "Canvas":
            try:
                widget.config(bg=_THEME["PANEL"])
            except tk.TclError:
                pass

        for child in widget.winfo_children():
            self._apply_theme(child)

    # ── UI construction ───────────────────────────────────────────────────────

    def _build_ui(self):
        hdr = tk.Frame(self.root, bg=ACCENT, height=56)
        hdr.pack(fill="x"); hdr.pack_propagate(False)

        logo_f = tk.Frame(hdr, bg=ACCENT); logo_f.pack(side="left", padx=14, pady=8)
        tk.Label(logo_f, text="KEITHLEY 2450", bg=ACCENT, fg="white",
                 font=FONT_H1).pack(anchor="w")
        tk.Label(logo_f, text="Source Measure Unit  ·  Lab View",
                 bg=ACCENT, fg="#7EB6E0", font=FONT_SM).pack(anchor="w")

        self._conn_lbl = tk.Label(
            hdr, text="● SMU: Disconnected",
            bg=ACCENT, fg="#EF4444", font=FONT_MONO)
        self._conn_lbl.pack(side="right", padx=(12, 14))

        self._dark_btn = tk.Button(
            hdr, text="🌙  Dark", font=FONT_SM,
            bg="#1A2535", fg="#FCD34D", relief="flat", bd=0,
            activebackground="#243040", cursor="hand2",
            padx=10, pady=3, command=self._toggle_dark)
        self._dark_btn.pack(side="right", padx=(0, 8))

        right_f = tk.Frame(hdr, bg=ACCENT); right_f.pack(side="right", padx=8)

        addr_f = tk.Frame(right_f, bg=ACCENT); addr_f.pack(side="right", padx=(0,8))
        tk.Label(addr_f, text="VISA Address", bg=ACCENT, fg="#7EB6E0",
                 font=FONT_SM).pack(anchor="e")
        self._addr_entry = tk.Entry(addr_f, width=36,
                                    bg="#1A3255", fg="white",
                                    insertbackground="white",
                                    relief="flat", font=FONT_MONO, bd=0,
                                    highlightthickness=1,
                                    highlightbackground="#3A5F8A",
                                    highlightcolor="#5A8FBF")
        self._addr_entry.insert(0, self.master_addr)
        self._addr_entry.pack(pady=(2,0), ipady=3)

        def _commit_addr(_=None):
            addr = self._addr_entry.get().strip()
            self.master_addr = addr
            _cfg.set("visa_address", addr)

        self._addr_entry.bind("<FocusOut>", _commit_addr)
        self._addr_entry.bind("<Return>",   _commit_addr)

        dir_f = tk.Frame(right_f, bg=ACCENT); dir_f.pack(side="right", padx=(0,8))
        dir_lbl_row = tk.Frame(dir_f, bg=ACCENT); dir_lbl_row.pack(fill="x")
        tk.Label(dir_lbl_row, text="Save Folder", bg=ACCENT, fg="#7EB6E0",
                 font=FONT_SM).pack(side="left")

        def _browse_dir():
            d = filedialog.askdirectory(initialdir=self.save_dir,
                                         title="Select save folder")
            if d:
                self.save_dir = d
                self._dir_entry.delete(0, tk.END)
                self._dir_entry.insert(0, d)
                _cfg.set("save_dir", d)

        tk.Button(dir_lbl_row, text="…", font=FONT_SM,
                  bg="#243F6A", fg="white", relief="flat", bd=0,
                  activebackground="#2A4A7F", cursor="hand2",
                  padx=4, pady=0, command=_browse_dir).pack(side="right", padx=(4,0))

        self._dir_entry = tk.Entry(dir_f, width=28,
                                   bg="#1A3255", fg="white",
                                   insertbackground="white",
                                   relief="flat", font=FONT_MONO, bd=0,
                                   highlightthickness=1,
                                   highlightbackground="#3A5F8A",
                                   highlightcolor="#5A8FBF")
        self._dir_entry.insert(0, self.save_dir)
        self._dir_entry.pack(pady=(2,0), ipady=3)

        def _commit_dir(_=None):
            d = self._dir_entry.get().strip()
            if d:
                self.save_dir = d
                _cfg.set("save_dir", d)

        self._dir_entry.bind("<FocusOut>", _commit_dir)
        self._dir_entry.bind("<Return>",   _commit_dir)

        sid_f = tk.Frame(right_f, bg=ACCENT); sid_f.pack(side="right", padx=(0,8))
        tk.Label(sid_f, text="Sample ID", bg=ACCENT, fg="#7EB6E0",
                 font=FONT_SM).pack(anchor="e")
        self.sample_id = tk.Entry(sid_f, width=14,
                                  bg="#1A3255", fg="white",
                                  insertbackground="white",
                                  relief="flat", font=FONT_MONO, bd=0,
                                  highlightthickness=1,
                                  highlightbackground="#3A5F8A",
                                  highlightcolor="#5A8FBF")
        self.sample_id.insert(0, _cfg.get("sample_id"))
        self.sample_id.pack(pady=(2,0), ipady=3)
        self.sample_id.bind("<FocusOut>",
            lambda _: _cfg.set("sample_id", self.sample_id.get().strip()))
        self.sample_id.bind("<Return>",
            lambda _: _cfg.set("sample_id", self.sample_id.get().strip()))

        self.status = StatusBar(self.root)
        self.status.pack(fill="x", side="bottom")

        paned = tk.PanedWindow(self.root, orient=tk.HORIZONTAL,
                               bg=_THEME["BG"], sashwidth=6,
                               sashrelief="flat", sashpad=2)
        paned.pack(fill="both", expand=True, padx=0, pady=0)

        lf = tk.Frame(paned, bg=_THEME["BG"], width=560)
        paned.add(lf, minsize=500)
        self._build_left_panel(lf)

        rf = tk.Frame(paned, bg=_THEME["BG"])
        paned.add(rf, stretch="always")
        self._build_right_panel(rf)

    def _build_left_panel(self, parent):
        cvs = tk.Canvas(parent, bg=_THEME["BG"], highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=cvs.yview)
        inner = tk.Frame(cvs, bg=_THEME["BG"])
        inner.bind("<Configure>",
                   lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        self._ctrl_win = cvs.create_window((0,0), window=inner, anchor="nw")
        cvs.bind("<Configure>",
                 lambda e: cvs.itemconfig(self._ctrl_win, width=e.width))
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        cvs.bind_all("<MouseWheel>",
                     lambda e: cvs.yview_scroll(int(-1*(e.delta/120)), "units"))
        self._build_controls(inner)

    def _build_controls(self, p):
        pad = {"padx": 10, "pady": (6,0)}

        src = Card(p, title="Source & Sweep"); src.pack(fill="x", **pad)
        b = src.body

        mf = _row(b, "Source Mode", 16)
        modes = ["Voltage Sweep","Current Sweep","Voltage List Sweep",
                 "Current List Sweep","Voltage Bias","Current Bias"]
        self.inputs["Mode"] = _combo(mf, modes, "Voltage Sweep",
                                     key="Mode", store=self.inputs)
        self.inputs["Mode"].bind("<<ComboboxSelected>>", self._update_ui_logic)

        chk_f = tk.Frame(b, bg=_THEME["PANEL"]); chk_f.pack(fill="x", pady=(4,2))
        self.inputs["Stepper"] = tk.BooleanVar()
        self.inputs["Dual"]    = tk.BooleanVar()
        self.inputs["PulseEn"] = tk.BooleanVar()
        _chk(chk_f, "Stepper (Node 2)",  self.inputs["Stepper"],
             command=self._update_ui_logic).pack(side="left", padx=(0,12))
        _chk(chk_f, "Dual Sweep",        self.inputs["Dual"]).pack(side="left", padx=(0,12))
        _chk(chk_f, "Pulse Mode",        self.inputs["PulseEn"],
             command=self._update_ui_logic).pack(side="left")

        tk.Frame(b, bg=_THEME["SEP"], height=1).pack(fill="x", pady=(6,4))

        self.f_sweep = tk.Frame(b, bg=_THEME["PANEL"])
        self.f_sweep.pack(fill="x")
        _param(self.f_sweep, "Start",  "0",   key="Start",  store=self.inputs, lbl_w=16)
        _param(self.f_sweep, "Stop",   "5",   key="Stop",   store=self.inputs, lbl_w=16)
        _param(self.f_sweep, "Step",   "0.1", key="Step",   store=self.inputs, lbl_w=16)
        _param(self.f_sweep, "Points", "51",  key="Points", store=self.inputs, lbl_w=16)
        self.inputs["Start"].bind("<FocusOut>",  self._calc_points)
        self.inputs["Stop"].bind("<FocusOut>",   self._calc_points)
        self.inputs["Step"].bind("<FocusOut>",   self._calc_points)
        self.inputs["Points"].bind("<FocusOut>", self._calc_step)

        _param(self.f_sweep, "Sweep Type", combo_vals=["Linear","Logarithmic"],
               combo_key="SweepType", store=self.inputs, lbl_w=16)

        self.f_list = tk.Frame(b, bg=_THEME["PANEL"])
        self._build_list_table(self.f_list)

        self.f_pulse = tk.Frame(b, bg=_THEME["PANEL2"],
                                highlightbackground=_THEME["BORDER"], highlightthickness=1)
        pulse_hdr = tk.Frame(self.f_pulse, bg=_THEME["ACCENT2"])
        pulse_hdr.pack(fill="x")
        tk.Label(pulse_hdr, text="  Pulse Parameters",
                 bg=_THEME["ACCENT2"], fg="white", font=FONT_SM).pack(side="left", pady=3, padx=4)
        _param(self.f_pulse, "Bias Level",   "0.00",  key="PulseBias", store=self.inputs, lbl_w=16)
        _param(self.f_pulse, "On Time (s)",  "0.001", key="OnTime",    store=self.inputs, lbl_w=16)
        _param(self.f_pulse, "Off Time (s)", "0.01",  key="OffTime",   store=self.inputs, lbl_w=16)

        tk.Frame(b, bg=_THEME["SEP"], height=1).pack(fill="x", pady=(6,4))
        _param(b, "Source Range",  combo_vals=["Best Fixed","Auto","200mV","2V","20V","200V"],
               combo_key="RangeV", store=self.inputs, lbl_w=16)
        _param(b, "Compliance",    "0.01", key="Limit",  store=self.inputs, lbl_w=16)
        _param(b, "Source Delay (s)", "0.1", key="Delay", store=self.inputs, lbl_w=16)

        self.f_stepper = tk.Frame(b, bg=_THEME["PANEL3"],
                                  highlightbackground=_THEME["ACCENT2"], highlightthickness=1)
        step_hdr = tk.Frame(self.f_stepper, bg=ACCENT)
        step_hdr.pack(fill="x")
        tk.Label(step_hdr, text="  Stepper  —  Node 2 Configuration",
                 bg=ACCENT, fg="white", font=FONT_SM).pack(side="left", pady=3, padx=4)
        _param(self.f_stepper, "Start (V)", "2.0", key="StepStart",  store=self.inputs, lbl_w=16)
        _param(self.f_stepper, "Stop (V)",  "3.0", key="StepStop",   store=self.inputs, lbl_w=16)
        _param(self.f_stepper, "Points",    "3",   key="StepPoints", store=self.inputs, lbl_w=16)

        meas = Card(p, title="Measure"); meas.pack(fill="x", **pad)
        mb = meas.body

        def meas_group(label, en_key, en_val, rk, rv, xlbl, xk, xv, xd):
            grp = tk.Frame(mb, bg=_THEME["PANEL"],
                           highlightbackground=_THEME["SEP"], highlightthickness=1)
            grp.pack(fill="x", pady=3)
            ghdr = tk.Frame(grp, bg=_THEME["PANEL2"]); ghdr.pack(fill="x")
            self.inputs[en_key] = tk.BooleanVar(value=en_val)
            _chk(ghdr, f"  {label}", self.inputs[en_key],
                 bg=_THEME["PANEL2"]).pack(side="left", padx=4, pady=3)
            gbody = tk.Frame(grp, bg=_THEME["PANEL"]); gbody.pack(fill="x", padx=8, pady=(2,4))
            r1 = tk.Frame(gbody, bg=_THEME["PANEL"]); r1.pack(fill="x", pady=1)
            tk.Label(r1, text="Range", bg=_THEME["PANEL"], fg=_THEME["TEXT2"], font=FONT_SM,
                     width=14, anchor="w").pack(side="left")
            self.inputs[rk] = ttk.Combobox(r1, values=rv, state="readonly",
                                            width=14, font=FONT_SM)
            self.inputs[rk].set("Auto")
            self.inputs[rk].pack(side="left", padx=2)
            r2 = tk.Frame(gbody, bg=_THEME["PANEL"]); r2.pack(fill="x", pady=1)
            tk.Label(r2, text=xlbl, bg=_THEME["PANEL"], fg=_THEME["TEXT2"], font=FONT_SM,
                     width=14, anchor="w").pack(side="left")
            self.inputs[xk] = ttk.Combobox(r2, values=xv, state="readonly",
                                            width=14, font=FONT_SM)
            self.inputs[xk].set(xd)
            self.inputs[xk].pack(side="left", padx=2)

        meas_group("Current", "MeasI", True,
                   "RangeI",
                   ["Auto","10pA","100pA","1nA","10nA","100nA",
                    "1uA","10uA","100uA","1mA","10mA","100mA","1A"],
                   "Min Auto Range","MinRangeI",
                   ["10pA","100pA","1nA","10nA","100nA",
                    "1uA","10uA","100uA","1mA"], "10nA")

        vg = tk.Frame(mb, bg=_THEME["PANEL"],
                      highlightbackground=_THEME["SEP"], highlightthickness=1)
        vg.pack(fill="x", pady=3)
        vghdr = tk.Frame(vg, bg=_THEME["PANEL2"]); vghdr.pack(fill="x")
        self.inputs["MeasV"] = tk.BooleanVar(value=False)
        _chk(vghdr, "  Voltage", self.inputs["MeasV"],
             bg=_THEME["PANEL2"]).pack(side="left", padx=4, pady=3)
        vgb = tk.Frame(vg, bg=_THEME["PANEL"]); vgb.pack(fill="x", padx=8, pady=(2,4))
        vt  = tk.Frame(vgb, bg=_THEME["PANEL"]); vt.pack(fill="x", pady=1)
        tk.Label(vt, text="Report", bg=_THEME["PANEL"], fg=_THEME["TEXT2"], font=FONT_SM,
                 width=14, anchor="w").pack(side="left")
        self.inputs["MeasVType"] = ttk.Combobox(vt, values=["Programmed","Actual"],
                                                 state="readonly", width=14, font=FONT_SM)
        self.inputs["MeasVType"].set("Programmed")
        self.inputs["MeasVType"].pack(side="left", padx=2)

        meas_group("Resistance", "MeasR", False,
                   "RangeR",
                   ["Auto","20Ω","200Ω","2kΩ","20kΩ","200kΩ","2MΩ","20MΩ","200MΩ"],
                   "Min Auto Range","MinRangeR",
                   ["20Ω","200Ω","2kΩ","20kΩ"], "20Ω")

        extras = tk.Frame(mb, bg=_THEME["PANEL"]); extras.pack(fill="x", pady=(4,0))
        self.inputs["Timestamp"] = tk.BooleanVar(value=True)
        self.inputs["MeasP"]     = tk.BooleanVar(value=False)
        _chk(extras, "Include Timestamp", self.inputs["Timestamp"]).pack(side="left")
        _chk(extras, "Power (W)",         self.inputs["MeasP"]).pack(side="left", padx=12)

        spd = Card(p, title="Instrument Settings"); spd.pack(fill="x", **pad)
        sb = spd.body
        _param(sb, "NPLC",           "1",    key="NPLC",      store=self.inputs, lbl_w=20)
        _param(sb, "Auto Zero",      None,   combo_vals=["Once","Off","On"],
               combo_key="AutoZero", store=self.inputs, lbl_w=20)
        _param(sb, "Input Jacks",    None,   combo_vals=["Rear","Front"],
               combo_key="Terminals", store=self.inputs, lbl_w=20)
        _param(sb, "Sense Mode",     None,   combo_vals=["2-Wire","4-Wire"],
               combo_key="Sense",    store=self.inputs, lbl_w=20)
        _param(sb, "Output OFF State", None, combo_vals=["Normal","High-Z","Zero"],
               combo_key="OffState", store=self.inputs, lbl_w=20)
        _param(sb, "High Capacitance", None, combo_vals=["Off","On"],
               combo_key="HighC",    store=self.inputs, lbl_w=20)

        cyc = Card(p, title="Cycling & Stability"); cyc.pack(fill="x", **pad)
        cb = cyc.body

        cyc_r1 = tk.Frame(cb, bg=_THEME["PANEL"]); cyc_r1.pack(fill="x", pady=2)
        _chk(cyc_r1, "Enable Cycling", self.cyc_enabled,
             command=self._update_ui_logic).pack(side="left")
        tk.Label(cyc_r1, text="Cycles:", bg=_THEME["PANEL"], fg=_THEME["TEXT2"],
                 font=FONT_BODY).pack(side="left", padx=(12,3))
        ttk.Entry(cyc_r1, textvariable=self.cyc_n,
                  width=5, font=FONT_BODY).pack(side="left")
        tk.Label(cyc_r1, text="Rest (s):", bg=_THEME["PANEL"], fg=_THEME["TEXT2"],
                 font=FONT_BODY).pack(side="left", padx=(10,3))
        ttk.Entry(cyc_r1, textvariable=self.cyc_rest,
                  width=6, font=FONT_BODY).pack(side="left")

        cyc_tr = tk.Frame(cb, bg=_THEME["PANEL"]); cyc_tr.pack(fill="x", pady=(2,0))
        tk.Label(cyc_tr, text="Trigger:", bg=_THEME["PANEL"], fg=_THEME["TEXT2"],
                 font=FONT_BODY).pack(side="left", padx=(0,6))
        for val, lbl in [("Auto",   "⟳  Auto  (after rest)"),
                         ("Manual", "☞  Manual  (wait for button)")]:
            tk.Radiobutton(cyc_tr, text=lbl, variable=self.cyc_trigger, value=val,
                           bg=_THEME["PANEL"], fg=_THEME["TEXT"], selectcolor=_THEME["PANEL"],
                           activebackground=_THEME["PANEL"], font=FONT_BODY,
                           command=self._update_ui_logic).pack(side="left", padx=(0,12))

        self._next_cycle_btn = tk.Button(
            cb, text="▶  Next Cycle", font=FONT_H3,
            bg="#16A34A", fg="white", relief="flat", bd=0,
            activebackground="#15803D", cursor="hand2",
            padx=14, pady=5,
            command=self._manual_next_cycle)

        self._cyc_name_frame = tk.Frame(cb, bg=_THEME["PANEL"],
                                         highlightbackground=_THEME["BORDER"],
                                         highlightthickness=1)
        self._cyc_name_frame.pack(fill="x", pady=(4,2))

        cyc_name_hdr = tk.Frame(self._cyc_name_frame, bg=_THEME["PANEL2"])
        cyc_name_hdr.pack(fill="x")
        tk.Label(cyc_name_hdr, text="  Cycle Names  — one per line (auto-used if filled)",
                 bg=_THEME["PANEL2"], fg=_THEME["TEXT2"], font=FONT_SM).pack(
                 side="left", pady=3, fill="x", expand=True)

        self._cyc_names_text = tk.Text(self._cyc_name_frame,
                                        height=4, width=28,
                                        font=FONT_BODY,
                                        bg=_THEME["PANEL"], fg=_THEME["TEXT"],
                                        insertbackground=_THEME["ACCENT2"],
                                        relief="flat", bd=0,
                                        highlightthickness=0)
        self._cyc_names_text.pack(fill="x", padx=4, pady=(2,0))
        tk.Label(self._cyc_name_frame,
                 text="  e.g.  Fresh,  +20°C,  Stressed,  Recovery…",
                 bg=_THEME["PANEL"], fg=_THEME["TEXT3"], font=FONT_SM, anchor="w").pack(
                 fill="x", padx=4, pady=(0,2))

        self._cyc_tags_lbl = tk.Label(self._cyc_name_frame,
                                       text="", bg=_THEME["PANEL"],
                                       fg=_THEME["ACCENT2"], font=FONT_SM,
                                       anchor="w", justify="left",
                                       wraplength=320)
        self._cyc_tags_lbl.pack(fill="x", padx=6, pady=(0,4))

        cyc_r2 = tk.Frame(cb, bg=_THEME["PANEL"]); cyc_r2.pack(fill="x", pady=(4,0))
        for txt, var in [("Gradient", self.cyc_gradient),
                         ("Show all", self.cyc_show_all),
                         ("Mean±σ",   self.cyc_stats)]:
            _chk(cyc_r2, txt, var, command=self._replot).pack(side="left", padx=(0,10))

        self.cyc_status_lbl = tk.Label(cb, text="", bg=_THEME["PANEL"],
                                        fg=_THEME["ACCENT2"], font=FONT_MONO)
        self.cyc_status_lbl.pack(anchor="w", pady=(2,0))

        mq = Card(p, title="Measurement Quality"); mq.pack(fill="x", **pad)
        mqb = mq.body

        mq_r1 = tk.Frame(mqb, bg=_THEME["PANEL"]); mq_r1.pack(fill="x", pady=2)
        tk.Label(mq_r1, text="Avg / point", bg=_THEME["PANEL"], fg=_THEME["TEXT2"],
                 font=FONT_BODY, width=20, anchor="w").pack(side="left")
        ttk.Combobox(mq_r1, textvariable=self.mq_avg,
                     values=["1","2","4","8","16"], state="readonly",
                     width=6, font=FONT_BODY).pack(side="left")
        tk.Label(mq_r1, text="readings", bg=_THEME["PANEL"], fg=_THEME["TEXT3"],
                 font=FONT_SM).pack(side="left", padx=4)

        mq_r2 = tk.Frame(mqb, bg=_THEME["PANEL"]); mq_r2.pack(fill="x", pady=2)
        _chk(mq_r2, "Settling check", self.mq_settle).pack(side="left")
        tk.Label(mq_r2, text="Tol", bg=_THEME["PANEL"], fg=_THEME["TEXT2"],
                 font=FONT_SM).pack(side="left", padx=(12,2))
        ttk.Entry(mq_r2, textvariable=self.mq_tol,
                  width=5, font=FONT_BODY).pack(side="left")
        tk.Label(mq_r2, text="%  (repeats until consecutive reads agree)",
                 bg=_THEME["PANEL"], fg=_THEME["TEXT3"], font=FONT_SM).pack(side="left", padx=4)

        mq_r3 = tk.Frame(mqb, bg=_THEME["PANEL"]); mq_r3.pack(fill="x", pady=(2,0))
        _chk(mq_r3, "Compliance pre-check before sweep",
             self.mq_precheck).pack(side="left")

        run_card = tk.Frame(p, bg=_THEME["BG"]); run_card.pack(fill="x", padx=10, pady=(8,12))
        ttk.Button(run_card, text="▶   RUN SWEEP",
                   style="Run.TButton",
                   command=self.start_thread).pack(side="left", fill="x",
                                                   expand=True, ipady=5, padx=(0,6))
        ttk.Button(run_card, text="■  STOP",
                   style="Stop.TButton",
                   command=self._force_stop).pack(side="left", ipady=5, ipadx=10)

    def _build_list_table(self, parent):
        thdr = tk.Frame(parent, bg=ACCENT); thdr.pack(fill="x")
        tk.Label(thdr, text="Index", width=8, bg=ACCENT, fg="white",
                 font=FONT_SM, anchor="center").pack(side="left", ipadx=4, ipady=3)
        self._list_col_lbl = tk.Label(thdr, text="Value",
                                      bg=ACCENT, fg="white", font=FONT_SM,
                                      anchor="center")
        self._list_col_lbl.pack(side="left", fill="x", expand=True, ipadx=4, ipady=3)
        tf = tk.Frame(parent, bg=_THEME["PANEL"]); tf.pack(fill="x")
        vsb = ttk.Scrollbar(tf, orient="vertical")
        self.list_tv = ttk.Treeview(tf, columns=("idx","val"), show="",
                                    height=7, yscrollcommand=vsb.set,
                                    selectmode="browse")
        vsb.config(command=self.list_tv.yview)
        self.list_tv.column("idx", width=60,  anchor="center", stretch=False)
        self.list_tv.column("val", width=150, anchor="center", stretch=True)
        self.list_tv.pack(side="left", fill="x", expand=True)
        vsb.pack(side="right", fill="y")
        self.list_tv.tag_configure("odd",  background=_THEME["PANEL"])
        self.list_tv.tag_configure("even", background=_THEME["PANEL2"])
        for i in range(1, 6):
            self.list_tv.insert("","end", values=(i,""),
                                tags=("odd" if i%2 else "even",))
        self.list_tv.bind("<Double-1>", self._list_inline_edit)
        br = tk.Frame(parent, bg=_THEME["PANEL"]); br.pack(fill="x", pady=(3,0))
        ttk.Button(br, text="+ Row",        command=self._list_add).pack(side="left", padx=(0,2))
        ttk.Button(br, text="✕ Delete",     command=self._list_del).pack(side="left", padx=(0,6))
        ttk.Button(br, text="⬆ Import",     command=self._list_import).pack(side="left")
        self._list_import_lbl = tk.Label(br, text="", bg=_THEME["PANEL"], fg=_THEME["TEXT3"], font=FONT_SM)
        self._list_import_lbl.pack(side="left", padx=6)

    def _list_add(self):
        n = len(self.list_tv.get_children()) + 1
        self.list_tv.insert("","end", values=(n,""),
                            tags=("odd" if n%2 else "even",))

    def _list_del(self):
        sel = self.list_tv.selection()
        if not sel: return
        self.list_tv.delete(sel[0])
        for i, iid in enumerate(self.list_tv.get_children(), 1):
            v = self.list_tv.item(iid,"values")[1]
            self.list_tv.item(iid, values=(i,v),
                              tags=("odd" if i%2 else "even",))

    def _list_inline_edit(self, event):
        if self.list_tv.identify("region", event.x, event.y) != "cell": return
        if self.list_tv.identify_column(event.x) != "#2": return
        iid = self.list_tv.identify_row(event.y)
        if not iid: return
        x, y, w, h = self.list_tv.bbox(iid, "#2")
        cur = self.list_tv.item(iid,"values")[1]
        ed = tk.Entry(self.list_tv, font=FONT_BODY, bg=_THEME["PANEL"], fg=_THEME["TEXT"],
                      insertbackground=_THEME["ACCENT2"], relief="solid", bd=1,
                      highlightthickness=1, highlightcolor=_THEME["ACCENT2"],
                      highlightbackground=_THEME["BORDER"])
        ed.insert(0, cur); ed.select_range(0, tk.END)
        ed.place(x=x, y=y, width=w, height=h); ed.focus_set()
        def commit(_=None):
            idx = self.list_tv.item(iid,"values")[0]
            tag = "odd" if int(idx)%2 else "even"
            self.list_tv.item(iid, values=(idx, ed.get().strip()), tags=(tag,))
            ed.destroy()
        ed.bind("<Return>", commit); ed.bind("<Tab>", commit)
        ed.bind("<Escape>", lambda _: ed.destroy())
        ed.bind("<FocusOut>", commit)

    def _list_import(self):
        path = filedialog.askopenfilename(
            title="Import List Values",
            filetypes=[("CSV","*.csv"),("Excel","*.xlsx *.xls"),("All","*.*")])
        if not path: return
        try:
            vals=[]; ext=path.rsplit(".",1)[-1].lower()
            if ext=="csv":
                with open(path, newline="", encoding="utf-8-sig") as f:
                    for row in csv.reader(f):
                        for cell in row:
                            try: vals.append(float(cell.strip())); break
                            except ValueError: continue
            elif ext in ("xlsx","xls"):
                try:
                    import openpyxl
                    wb=openpyxl.load_workbook(path, read_only=True, data_only=True)
                    for row in wb.active.iter_rows(values_only=True):
                        for cell in row:
                            try: vals.append(float(cell)); break
                            except: continue
                except ImportError:
                    messagebox.showerror("Missing Package","pip install openpyxl"); return
            if not vals:
                messagebox.showwarning("No Data","No numeric values found."); return
            for iid in self.list_tv.get_children(): self.list_tv.delete(iid)
            for i,v in enumerate(vals,1):
                self.list_tv.insert("","end", values=(i,f"{v:g}"),
                                    tags=("odd" if i%2 else "even",))
            self._list_import_lbl.config(text=os.path.basename(path))
        except Exception as e:
            messagebox.showerror("Import Error", str(e))

    def _get_list_values(self):
        vals=[]
        for iid in self.list_tv.get_children():
            try: vals.append(float(self.list_tv.item(iid,"values")[1]))
            except: pass
        return vals

    def _build_right_panel(self, parent):
        tb = tk.Frame(parent, bg=_THEME["PANEL"],
                      highlightbackground=_THEME["BORDER"], highlightthickness=1)
        tb.pack(fill="x", padx=8, pady=(8,4))

        prow = tk.Frame(tb, bg=_THEME["PANEL"]); prow.pack(fill="x", padx=10, pady=(7,0))
        self.progress = ProgressRow(prow)
        self.progress.pack(fill="x")

        brow = tk.Frame(tb, bg=_THEME["PANEL"]); brow.pack(fill="x", padx=10, pady=(4,7))
        ttk.Button(brow, text="Clear Plot",   command=self._clear_plot).pack(side="left", padx=(0,4))
        ttk.Button(brow, text="Export CSV",   command=self._save_csv).pack(side="left", padx=(0,4))
        ttk.Button(brow, text="Cycle CSV",    command=self._save_cycle_csv).pack(side="left", padx=(0,4))
        ttk.Button(brow, text="Save PNG",     command=self._save_png).pack(side="left", padx=(0,4))
        ttk.Button(brow, text="Auto Scale",   command=self._auto_scale).pack(side="left", padx=(0,4))

        self._ghost_btn = tk.Button(
            brow, text="Ghost: ON", font=FONT_SM,
            bg=_THEME["ACCENT3"], fg=_THEME["ACCENT2"], relief="flat", bd=0,
            activebackground=_THEME["PANEL2"], cursor="hand2",
            padx=8, pady=2, command=self._toggle_ghost)
        self._ghost_btn.pack(side="left", padx=(8,4))

        ttk.Button(brow, text="⚗ Analysis",
                   command=self._run_analysis_iv).pack(side="left", padx=(4,0))

        self._cmp_btn = tk.Button(
            brow, text="📊 Compare CSV",
            font=FONT_SM,
            bg=_THEME["PANEL2"], fg=_THEME["TEXT2"],
            relief="flat", bd=0,
            activebackground=_THEME["HOVER"], cursor="hand2",
            padx=8, pady=2,
            command=lambda: self._compare.open_panel())
        self._cmp_btn.pack(side="left", padx=(8,0))

        an_f = tk.Frame(parent, bg=_THEME["PANEL"],
                        highlightbackground=_THEME["BORDER2"], highlightthickness=1)
        an_f.pack(fill="x", padx=8, pady=(0,4))

        an_hdr = tk.Frame(an_f, bg="#166534"); an_hdr.pack(fill="x")
        tk.Label(an_hdr, text="  ⚗  Parameter Extraction",
                 bg="#166534", fg="white", font=FONT_H3).pack(side="left", pady=4)

        self._an_open = False
        self._an_body = tk.Frame(an_f, bg=_THEME["PANEL"])

        def _toggle_an():
            if self._an_open:
                self._an_body.pack_forget()
                self._an_open = False
                an_arr.config(text="▸")
            else:
                self._an_body.pack(fill="x")
                self._an_open = True
                an_arr.config(text="▾")

        an_arr = tk.Label(an_hdr, text="▸", bg="#166534", fg="white",
                          font=("Segoe UI", 10), cursor="hand2")
        an_arr.pack(side="right", padx=10)
        an_arr.bind("<Button-1>", lambda _: _toggle_an())
        an_hdr.bind("<Button-1>", lambda _: _toggle_an())

        an_bdy = self._an_body
        an_r1  = tk.Frame(an_bdy, bg=_THEME["PANEL"]); an_r1.pack(fill="x", padx=10, pady=(6,2))
        for txt, var in [("Vth",      self.an_vth),
                         ("SS mV/dec",self.an_ss),
                         ("Ion/Ioff", self.an_ionoff),
                         ("Hysteresis",self.an_hyst)]:
            _chk(an_r1, txt, var).pack(side="left", padx=(0,10))

        an_r2 = tk.Frame(an_bdy, bg=_THEME["PANEL"]); an_r2.pack(fill="x", padx=10, pady=(0,6))
        ttk.Button(an_r2, text="Run Analysis", style="Accent.TButton",
                   command=self._run_analysis_iv).pack(side="left", padx=(0,12))
        self.an_result_lbl = tk.Label(an_r2, text="—", bg=_THEME["PANEL"],
                                       fg=_THEME["ACCENT2"], font=FONT_MONO)
        self.an_result_lbl.pack(side="left", fill="x")
        self.an_trend_lbl  = tk.Label(an_bdy, text="", bg=_THEME["PANEL"],
                                       fg=_THEME["TEXT2"], font=FONT_SM)
        self.an_trend_lbl.pack(anchor="w", padx=10, pady=(0,4))

        body = tk.Frame(parent, bg=_THEME["BG"])
        body.pack(fill="both", expand=True, padx=8, pady=(0,8))

        chart_f = tk.Frame(body, bg=_THEME["PANEL"],
                           highlightbackground=_THEME["BORDER"],
                           highlightthickness=1)
        chart_f.pack(side="left", fill="both", expand=True)

        self.fig = Figure(figsize=(7,5), dpi=100, facecolor=_THEME["PANEL"])
        self.fig.subplots_adjust(left=0.11, right=0.88, top=0.93, bottom=0.10)
        self.ax  = self.fig.add_subplot(111)
        self.ax2 = self.ax.twinx()
        self.ax2.set_visible(False)
        _style_plot(self.fig, self.ax, title="I-V Characteristic")

        self.canvas = FigureCanvasTkAgg(self.fig, master=chart_f)
        self.canvas.draw()
        self.canvas.get_tk_widget().config(bg=_THEME["PANEL"])
        self.canvas.get_tk_widget().pack(fill="both", expand=True, padx=2, pady=2)

        self._build_axis_panel(body)

    def _build_axis_panel(self, parent):
        sb = tk.Frame(parent, bg=_THEME["PANEL"], width=172,
                      highlightbackground=_THEME["BORDER"], highlightthickness=1)
        sb.pack(side="right", fill="y", padx=(6,0))
        sb.pack_propagate(False)

        self.ax_x  = tk.StringVar(value="Voltage (V)")
        self.ax_y1 = {ch: tk.BooleanVar(value=(ch=="Current (A)"))
                      for ch in self.CHANNELS}
        self.ax_y2 = {ch: tk.BooleanVar(value=False) for ch in self.CHANNELS}

        def section(title, col):
            tk.Frame(sb, bg=col, height=2).pack(fill="x")
            tk.Label(sb, text=title, bg=_THEME["PANEL2"], fg=col,
                     font=FONT_H3, anchor="w").pack(fill="x", padx=8, pady=(5,2))
            body = tk.Frame(sb, bg=_THEME["PANEL"])
            body.pack(fill="x", padx=4, pady=(0,4))
            tk.Frame(sb, bg=_THEME["SEP"], height=1).pack(fill="x")
            return body

        xb = section("X-Axis", ACCENT)
        for ch in self.CHANNELS:
            r = tk.Frame(xb, bg=_THEME["PANEL"]); r.pack(fill="x", pady=1)
            tk.Label(r, text="◆", bg=_THEME["PANEL"], fg=_THEME["TEXT3"],
                     font=("Segoe UI",7)).pack(side="left", padx=(4,2))
            tk.Radiobutton(r, text=ch, variable=self.ax_x, value=ch,
                           bg=_THEME["PANEL"], fg=_THEME["TEXT"], selectcolor=_THEME["PANEL"],
                           activebackground=_THEME["HOVER"], activeforeground=ACCENT,
                           font=FONT_SM, command=self._replot).pack(side="left")

        y1b = section("Y1-Axis (left)", Y1_PAL[0])
        for i, ch in enumerate(self.CHANNELS):
            col = Y1_PAL[i % len(Y1_PAL)]
            r = tk.Frame(y1b, bg=_THEME["PANEL"]); r.pack(fill="x", pady=1)
            tk.Label(r, text="■", bg=_THEME["PANEL"], fg=col,
                     font=("Segoe UI",9)).pack(side="left", padx=(4,2))
            tk.Checkbutton(r, text=ch, variable=self.ax_y1[ch],
                           bg=_THEME["PANEL"], fg=_THEME["TEXT"], selectcolor=_THEME["PANEL"],
                           activebackground=_THEME["HOVER"], activeforeground=col,
                           font=FONT_SM, relief="flat",
                           highlightthickness=0,
                           command=self._replot).pack(side="left")

        y2b = section("Y2-Axis (right)", Y2_PAL[0])
        for i, ch in enumerate(self.CHANNELS):
            col = Y2_PAL[i % len(Y2_PAL)]
            r = tk.Frame(y2b, bg=_THEME["PANEL"]); r.pack(fill="x", pady=1)
            tk.Label(r, text="■", bg=_THEME["PANEL"], fg=col,
                     font=("Segoe UI",9)).pack(side="left", padx=(4,2))
            tk.Checkbutton(r, text=ch, variable=self.ax_y2[ch],
                           bg=_THEME["PANEL"], fg=_THEME["TEXT"], selectcolor=_THEME["PANEL"],
                           activebackground=_THEME["HOVER"], activeforeground=col,
                           font=FONT_SM, relief="flat",
                           highlightthickness=0,
                           command=self._replot).pack(side="left")

    # ── UI logic ──────────────────────────────────────────────────────────────

    def _update_ui_logic(self, _=None):
        mode = self.inputs["Mode"].get()
        is_list = "List" in mode
        if is_list:
            self.f_sweep.pack_forget()
            self.f_list.pack(fill="x", pady=(2,0))
            self._list_col_lbl.config(
                text="Voltage Value" if "Voltage" in mode else "Current Value")
        else:
            self.f_list.pack_forget()
            self.f_sweep.pack(fill="x", pady=(2,0))

        if self.inputs["PulseEn"].get():
            self.f_pulse.pack(fill="x", pady=(4,0))
        else:
            self.f_pulse.pack_forget()

        if self.inputs["Stepper"].get():
            self.f_stepper.pack(fill="x", pady=(4,0))
        else:
            self.f_stepper.pack_forget()

    def _calc_points(self, _=None):
        try:
            s = float(self.inputs["Start"].get())
            e = float(self.inputs["Stop"].get())
            st= float(self.inputs["Step"].get())
            if st == 0: return
            n = int(round(abs(e-s)/abs(st))) + 1
            self.inputs["Points"].delete(0, tk.END)
            self.inputs["Points"].insert(0, str(n))
        except: pass

    def _calc_step(self, _=None):
        try:
            s  = float(self.inputs["Start"].get())
            e  = float(self.inputs["Stop"].get())
            n  = int(self.inputs["Points"].get())
            if n <= 1: return
            st = (e-s)/(n-1)
            self.inputs["Step"].delete(0, tk.END)
            self.inputs["Step"].insert(0, f"{st:.6g}")
        except: pass

    # ── Plot ──────────────────────────────────────────────────────────────────

    def _manual_next_cycle(self):
        self._cyc_manual_go.set()

    def _get_cycle_name(self, cycle_idx: int) -> str:
        try:
            raw   = self._cyc_names_text.get("1.0", tk.END)
            lines = [l.strip() for l in raw.splitlines() if l.strip()]
            if cycle_idx < len(lines):
                return lines[cycle_idx]
        except Exception:
            pass
        return f"Cyc {cycle_idx + 1}"

    def _update_next_btn(self, waiting: bool):
        def _do():
            if not hasattr(self, "_next_cycle_btn"): return
            if waiting:
                self._next_cycle_btn.config(
                    text="▶  Next Cycle  →", state="normal",
                    bg="#16A34A")
                self._next_cycle_btn.pack(fill="x", padx=0, pady=(4,2))
            else:
                self._next_cycle_btn.config(state="disabled",
                                             bg=_THEME["TEXT3"])
                self._next_cycle_btn.pack_forget()
        self.root.after(0, _do)

    def _toggle_ghost(self):
        self.cyc_overlay.set(not self.cyc_overlay.get())
        if self.cyc_overlay.get():
            self._ghost_btn.config(text="Ghost: ON",
                                   bg=_THEME["ACCENT3"], fg=_THEME["ACCENT2"])
        else:
            self._ghost_btn.config(text="Ghost: OFF",
                                   bg=_THEME["PANEL2"], fg=_THEME["TEXT3"])
        self._replot()

    def _replot(self):
        has_data    = bool(self.sweep_data) or bool(self._cycle_runs)
        has_ghost   = bool(self.cyc_overlay.get() and self._overlay_data)
        has_compare = bool(self._compare.refs)
        if not has_data and not has_ghost and not has_compare:
            return

        x_ch   = self.ax_x.get()
        y1_chs = [c for c in self.CHANNELS if self.ax_y1[c].get()]
        y2_chs = [c for c in self.CHANNELS if self.ax_y2[c].get()]

        if not has_data and not has_ghost and has_compare:
            if not y1_chs:
                y1_chs = ["Current (A)"]
            if not x_ch:
                x_ch = "Voltage (V)"

        self.fig.clear()
        self.fig.patch.set_facecolor(_THEME["PANEL"])
        self.ax  = self.fig.add_subplot(111)
        self.ax2 = self.ax.twinx()
        self.fig.subplots_adjust(left=0.11, right=0.88 if y2_chs else 0.95,
                                 top=0.93, bottom=0.10)
        _style_plot(self.fig, self.ax, self.ax2 if y2_chs else None)

        self.ax.set_xlabel(x_ch, color=_THEME["TEXT"], fontsize=9, labelpad=5)
        if y1_chs:
            self.ax.set_ylabel(", ".join(y1_chs), color=Y1_PAL[0],
                               fontsize=9, labelpad=5)
            self.ax.tick_params(axis="y", colors=Y1_PAL[0])
            _eng_formatter(self.ax, "y")
        if y2_chs:
            self.ax2.set_ylabel(", ".join(y2_chs), color=Y2_PAL[0],
                                fontsize=9, labelpad=5)
            self.ax2.tick_params(axis="y", colors=Y2_PAL[0])
            _eng_formatter(self.ax2, "y")
        else:
            self.ax2.set_visible(False)

        if has_ghost:
            ghost_col  = "#888888" if not self._dark_mode else "#4A5568"
            ghost_col2 = "#AA7777" if not self._dark_mode else "#7B4E4E"
            for curve in self._overlay_data:
                xd = curve.get(x_ch, [])
                if not xd: continue
                for ch in y1_chs:
                    yd = curve.get(ch, [])
                    n  = min(len(xd), len(yd))
                    if n == 0: continue
                    self.ax.plot(xd[:n], yd[:n], "-", lw=1.4,
                                 color=ghost_col, alpha=0.50, zorder=1)
                for ch in y2_chs:
                    yd = curve.get(ch, [])
                    n  = min(len(xd), len(yd))
                    if n == 0: continue
                    self.ax2.plot(xd[:n], yd[:n], "-", lw=1.4,
                                  color=ghost_col2, alpha=0.40, zorder=1)

        if self._compare.refs:
            self._compare.draw(self.ax)

        cyc_enabled = self.cyc_enabled.get()
        all_cycles  = list(self._cycle_runs)
        if self.sweep_data:
            all_cycles.append(list(self.sweep_data))

        use_cycles = cyc_enabled and len(all_cycles) >= 1

        if use_cycles:
            n_cyc = len(all_cycles)
            cyc_cols = [Y1_PAL[i % len(Y1_PAL)] for i in range(n_cyc)]

            for ci, (cyc_curves, c_col) in enumerate(zip(all_cycles, cyc_cols)):
                is_current = (ci == len(all_cycles) - 1 and self.is_running)
                alpha = 1.0 if (is_current or n_cyc == 1) \
                        else max(0.45, 1.0 - 0.07 * (n_cyc - ci - 1))
                lw    = 1.8 if (is_current or n_cyc == 1) \
                        else (1.0 if n_cyc > 4 else 1.4)
                ms    = 2.5 if (is_current or n_cyc == 1) \
                        else (1.5 if n_cyc > 4 else 2.0)
                sfx_live = " [live]" if is_current else ""

                for curve in cyc_curves:
                    xd = curve.get(x_ch, [])
                    if not xd: continue
                    sv   = curve.get("step_val", 0)
                    sfx  = f"  [V₂={sv:.3g}V]" if sv else ""
                    for j, ch in enumerate(y1_chs):
                        yd = curve.get(ch, [])
                        n  = min(len(xd), len(yd))
                        if n == 0: continue
                        col = c_col
                        self.ax.plot(xd[:n], yd[:n], "o-",
                                     markersize=ms, linewidth=lw,
                                     color=col, alpha=alpha,
                                     label=f"{ch}  {self._get_cycle_name(ci)}{sfx}{sfx_live}" if n_cyc > 1
                                           else f"{ch}{sfx}",
                                     zorder=3 + ci)
                    for j, ch in enumerate(y2_chs):
                        yd = curve.get(ch, [])
                        n  = min(len(xd), len(yd))
                        if n == 0: continue
                        col = Y2_PAL[j % len(Y2_PAL)]
                        self.ax2.plot(xd[:n], yd[:n], "s--",
                                      markersize=ms, linewidth=lw,
                                      color=col, alpha=alpha,
                                      label=f"{ch}  {self._get_cycle_name(ci)}{sfx}{sfx_live} [Y2]" if n_cyc > 1
                                            else f"{ch}{sfx} [Y2]",
                                      zorder=3 + ci)

            completed_cycs = all_cycles[:-1] if self.is_running else all_cycles
            if (self.cyc_show_all.get() and self.cyc_stats.get()
                    and len(completed_cycs) >= 2 and y1_chs):
                try:
                    ch    = y1_chs[0]
                    ref_x = np.array(completed_cycs[0][0].get(x_ch, []))
                    if len(ref_x) >= 3:
                        all_y = []
                        for cyc_c in completed_cycs:
                            for cv in cyc_c:
                                xd = np.array(cv.get(x_ch, []))
                                yd = np.array(cv.get(ch, []))
                                if len(xd) < 2: continue
                                order = np.argsort(xd)
                                yi = np.interp(ref_x, xd[order], yd[order])
                                all_y.append(yi)
                        if len(all_y) >= 2:
                            ay  = np.array(all_y)
                            mu  = np.nanmean(ay, axis=0)
                            sg  = np.nanstd(ay, axis=0)
                            v   = ~np.isnan(mu)
                            band_col = "#22D3EE" if self._dark_mode else "#3B82F6"
                            self.ax.fill_between(
                                ref_x[v], (mu - sg)[v], (mu + sg)[v],
                                alpha=0.20 if self._dark_mode else 0.15,
                                color=band_col, zorder=2,
                                label=f"μ±σ  ({len(completed_cycs)} cycles)")
                            self.ax.plot(ref_x[v], mu[v], "--",
                                         lw=1.8, color=band_col,
                                         alpha=0.90, zorder=4, label="Mean")
                except Exception:
                    pass

        else:
            src = (self.sweep_data if self.sweep_data
                   else (self._cycle_runs[-1] if self._cycle_runs else []))
            for curve in src:
                xd = curve.get(x_ch, [])
                if not xd: continue
                sv  = curve.get("step_val", 0)
                sfx = f"  [V₂={sv:.3g}V]" if sv else ""
                for j, ch in enumerate(y1_chs):
                    yd = curve.get(ch, [])
                    n  = min(len(xd), len(yd))
                    if n == 0: continue
                    col = Y1_PAL[j % len(Y1_PAL)]
                    self.ax.plot(xd[:n], yd[:n], "o-",
                                 markersize=2.5, linewidth=1.6,
                                 color=col, label=f"{ch}{sfx}")
                for j, ch in enumerate(y2_chs):
                    yd = curve.get(ch, [])
                    n  = min(len(xd), len(yd))
                    if n == 0: continue
                    col = Y2_PAL[j % len(Y2_PAL)]
                    self.ax2.plot(xd[:n], yd[:n], "s--",
                                  markersize=2.5, linewidth=1.6,
                                  color=col, label=f"{ch}{sfx} [Y2]")

        handles, auto_labels = self.ax.get_legend_handles_labels()
        if handles:
            self.ax.legend(handles, auto_labels, fontsize=7.5, loc="best",
                           framealpha=0.85,
                           edgecolor=_THEME["BORDER"],
                           facecolor=_THEME["PANEL"],
                           labelcolor=_THEME["TEXT"],
                           labelspacing=0.3)
        self.canvas.draw_idle()

    def _clear_plot(self):
        if self.is_running: return
        if not messagebox.askyesno("Clear", "Clear all sweep data and plot?"):
            return
        self.sweep_data     = []
        self._cycle_runs    = []
        self._cycle_names   = []
        self._overlay_data  = None
        self._current_cycle = 0
        self.fig.clear()
        self.fig.patch.set_facecolor(_THEME["PANEL"])
        self.ax  = self.fig.add_subplot(111)
        self.ax2 = self.ax.twinx()
        self.ax2.set_visible(False)
        self.fig.subplots_adjust(left=0.11, right=0.95, top=0.93, bottom=0.10)
        _style_plot(self.fig, self.ax, title="I-V Characteristic")
        self.canvas.draw()
        self.progress.reset()
        self.status.set("ready")
        if hasattr(self, "an_result_lbl"):
            self.an_result_lbl.config(text="—", fg=_THEME["ACCENT2"])
        if hasattr(self, "an_trend_lbl"):
            self.an_trend_lbl.config(text="")
        if hasattr(self, "cyc_status_lbl"):
            self.cyc_status_lbl.config(text="")
        if hasattr(self, "_cyc_tags_lbl"):
            self._cyc_tags_lbl.config(text="")

    def _auto_scale(self):
        try:
            self.ax.relim(); self.ax.autoscale_view()
            if self.ax2.get_visible():
                self.ax2.relim(); self.ax2.autoscale_view()
            self.canvas.draw_idle()
        except: pass

    # ── Measurement thread ────────────────────────────────────────────────────

    def start_thread(self):
        if self.is_running: return
        self.is_running  = True
        self._stop_flag  = False
        if self.cyc_overlay.get():
            src = []
            if self._cycle_runs:
                for cyc in self._cycle_runs:
                    src.extend(cyc)
            elif self.sweep_data:
                src = self.sweep_data
            if src:
                self._overlay_data = copy.deepcopy(src)
        self.sweep_data     = []
        self._cycle_runs    = []
        self._cycle_names   = []
        self._current_cycle = 0
        self.fig.clear()
        self.fig.patch.set_facecolor(_THEME["PANEL"])
        self.ax  = self.fig.add_subplot(111)
        self.ax2 = self.ax.twinx()
        self.ax2.set_visible(False)
        self.fig.subplots_adjust(left=0.11, right=0.95, top=0.93, bottom=0.10)
        _style_plot(self.fig, self.ax, title="Acquiring…")
        self.canvas.draw()
        self.status.set("running")
        self.progress.start()
        threading.Thread(target=self._run_process, daemon=True).start()

    def _run_process(self):
        p = self.inputs
        try:
            self.inst = self.rm.open_resource(self.master_addr)
            self.inst.timeout = 30000
            self._set_conn_status(True, self.master_addr)

            self.inst.write("*CLS")
            self.inst.write("reset()")
            time.sleep(0.5)
            self.inst.write("defbuffer1.clear()")
            time.sleep(0.1)

            def tsp_check(label=""):
                resp = self.inst.query(
                    "print(errorqueue.count, errorqueue.next())").strip()
                parts = resp.split(",", 1)
                try:
                    n = int(float(parts[0]))
                except (ValueError, IndexError):
                    return
                if n > 0 and len(parts) > 1:
                    raise RuntimeError(f"TSP [{label}]: {parts[1].strip()}")

            self.inst.write(
                f"smu.terminals = smu.TERMINALS_{p['Terminals'].get().upper()}")
            sense = "SENSE_4WIRE" if "4" in p["Sense"].get() else "SENSE_2WIRE"
            self.inst.write(f"smu.measure.sense = smu.{sense}")
            self.inst.write(f"smu.measure.nplc = {p['NPLC'].get()}")
            tsp_check("nplc")

            az = p["AutoZero"].get()
            if az == "Once":
                self.inst.write("smu.measure.autozero.enable = smu.OFF")
                self.inst.write("smu.measure.autozero.once()")
            elif az == "On":
                self.inst.write("smu.measure.autozero.enable = smu.ON")
            else:
                self.inst.write("smu.measure.autozero.enable = smu.OFF")

            hc = "smu.ON" if p["HighC"].get() == "On" else "smu.OFF"
            self.inst.write(f"smu.measure.highcapacitance = {hc}")
            tsp_check("highcapacitance")

            def fmt_range(v):
                return (v.replace("pA","e-12").replace("nA","e-9")
                         .replace("uA","e-6").replace("mA","e-3")
                         .replace("Ω","").replace("kΩ","e3").replace("MΩ","e6")
                         .replace("A",""))

            ri = p["RangeI"].get()
            if ri == "Auto":
                self.inst.write("smu.measure.autorange = smu.ON")
                self.inst.write(
                    f"smu.measure.autorangelow = {fmt_range(p['MinRangeI'].get())}")
            else:
                self.inst.write(f"smu.measure.range = {fmt_range(ri)}")

            mode_str   = p["Mode"].get()
            is_voltage = "Voltage" in mode_str
            if is_voltage:
                self.inst.write("smu.source.func = smu.FUNC_DC_VOLTAGE")
                self.inst.write(f"smu.source.ilimit.level = {p['Limit'].get()}")
            else:
                self.inst.write("smu.source.func = smu.FUNC_DC_CURRENT")
                self.inst.write(f"smu.source.vlimit.level = {p['Limit'].get()}")

            vr = p["RangeV"].get()
            if vr in ("Auto","Best Fixed"):
                self.inst.write("smu.source.autorange = smu.ON")
            else:
                vmap = {"200mV":"0.2","2V":"2","20V":"20","200V":"200"}
                self.inst.write(f"smu.source.range = {vmap.get(vr,'20')}")

            steps = [0.0]
            if p["Stepper"].get():
                steps = np.linspace(
                    float(p["StepStart"].get()),
                    float(p["StepStop"].get()),
                    max(int(p["StepPoints"].get()), 1)
                ).tolist()

            try:   mq_avg = max(1, int(self.mq_avg.get()))
            except: mq_avg = 1
            mq_settle  = self.mq_settle.get()
            try:   mq_tol = max(0.001, float(self.mq_tol.get())) / 100.0
            except: mq_tol = 0.01
            mq_precheck = self.mq_precheck.get()

            dual_on = p["Dual"].get()
            delay   = p["Delay"].get()

            if mq_precheck:
                self.root.after(0, lambda: self.status.msg("Pre-check…", WARN))
                try:
                    sv1_f = float(p["Start"].get())
                    self.inst.write("smu.source.output = smu.ON")
                    self.inst.write(f"smu.source.level = {sv1_f:.6g}")
                    time.sleep(max(float(delay), 0.1))
                    i_test = float(self.inst.query("print(smu.measure.read())").strip())
                    comp_f = float(p["Limit"].get())
                    if abs(i_test) >= comp_f * 0.95:
                        self.inst.write("smu.source.level = 0")
                        raise RuntimeError(
                            f"Compliance pre-check failed:\n"
                            f"I = {i_test:.4g} A at start point.\n"
                            f"Device may be shorted or compliance too low.\n\n"
                            f"Raise Compliance or reduce Start voltage.")
                    self.inst.write("smu.source.level = 0")
                    time.sleep(0.05)
                except RuntimeError:
                    raise
                except Exception:
                    pass

            if mq_avg > 1:
                try:
                    base_nplc = float(p["NPLC"].get())
                    self.inst.write(f"smu.measure.nplc = {base_nplc * mq_avg:.3g}")
                except Exception:
                    pass

            cyc_enabled  = self.cyc_enabled.get()
            cyc_manual   = cyc_enabled and (self.cyc_trigger.get() == "Manual")
            try:    n_cycles = max(1, int(self.cyc_n.get())) if cyc_enabled else 1
            except: n_cycles = 1
            try:    cyc_rest = max(0.0, float(self.cyc_rest.get()))
            except: cyc_rest = 0.0

            total_pts_all = n_cycles * len(steps)

            for cycle_idx in range(n_cycles):
                if self._stop_flag: break

                self._current_cycle = cycle_idx

                if cycle_idx > 0 and not self._stop_flag:
                    if cyc_manual:
                        self._cyc_manual_go.clear()
                        self._update_next_btn(waiting=True)
                        self.root.after(0, lambda ci=cycle_idx, nc=n_cycles:
                            self.status.msg(
                                f"⏸  Cycle {ci}/{nc} done  —  "
                                f"press  ▶ Next Cycle  to continue", "#60A5FA"))
                        while not self._cyc_manual_go.wait(timeout=0.2):
                            if self._stop_flag: break
                        self._update_next_btn(waiting=False)
                        if self._stop_flag: break
                    else:
                        if cyc_rest > 0:
                            self.root.after(0, lambda ci=cycle_idx, nc=n_cycles:
                                self.status.msg(
                                    f"Cycle {ci}/{nc}  → Rest {cyc_rest:.1f}s…",
                                    "#60A5FA"))
                            t_rest = time.time()
                            while time.time() - t_rest < cyc_rest:
                                if self._stop_flag: break
                                rem = cyc_rest - (time.time() - t_rest)
                                self.root.after(0, lambda r=rem, ci=cycle_idx,
                                                nc=n_cycles:
                                    self.status.msg(
                                        f"Cycle {ci}/{nc}  → Rest  {r:.1f}s remaining…",
                                        "#60A5FA"))
                                time.sleep(0.5)

                if n_cycles > 1:
                    self.root.after(0, lambda ci=cycle_idx, nc=n_cycles:
                        self.status.msg(
                            f"{self._get_cycle_name(ci)}  ({ci+1}/{nc})  — Sweeping…",
                            WARN))

                cycle_curves = []

                for step_idx, step_val in enumerate(steps):
                    if self._stop_flag: break

                    if p["Stepper"].get():
                        self.inst.write(f"node[2].smu.source.level = {step_val}")
                        self.inst.write("node[2].smu.source.output = node[2].smu.ON")
                        self.root.after(0, lambda v=step_val:
                            self.status.msg(f"Stepping  V₂ = {v:.3f} V", WARN))

                    if "List" in mode_str:
                        lv = self._get_list_values()
                        if not lv:
                            raise ValueError("List sweep: no values defined.")
                        fwd = ",".join(f"{v:g}" for v in lv)
                        rev = ",".join(f"{v:g}" for v in reversed(lv))
                        passes = [
                            (f"smu.source.sweeplist('sw',{{{fwd}}},{delay},"
                             f"1,smu.RANGE_BEST,smu.OFF)", len(lv))
                        ]
                        if dual_on:
                            passes.append((
                                f"smu.source.sweeplist('sw',{{{rev}}},{delay},"
                                f"1,smu.RANGE_BEST,smu.OFF)", len(lv)))
                    else:
                        sv1 = p["Start"].get()
                        sv2 = p["Stop"].get()
                        pts = int(p["Points"].get())
                        if p["PulseEn"].get():
                            on_t  = p["OnTime"].get()
                            off_t = p["OffTime"].get()
                            passes = [
                                (f"smu.source.pulse.sweeplinear('sw',{sv1},{sv2},"
                                 f"{pts},{on_t},{off_t},1,smu.RANGE_BEST,smu.OFF)", pts)
                            ]
                            if dual_on:
                                passes.append((
                                    f"smu.source.pulse.sweeplinear('sw',{sv2},{sv1},"
                                    f"{pts},{on_t},{off_t},1,smu.RANGE_BEST,smu.OFF)", pts))
                        else:
                            passes = [
                                (f"smu.source.sweeplinear('sw',{sv1},{sv2},"
                                 f"{pts},{delay},1,smu.RANGE_BEST,smu.OFF)", pts)
                            ]
                            if dual_on:
                                passes.append((
                                    f"smu.source.sweeplinear('sw',{sv2},{sv1},"
                                    f"{pts},{delay},1,smu.RANGE_BEST,smu.OFF)", pts))

                    pts_total = sum(n for _, n in passes)
                    live = {
                        "step_val":       step_val,
                        "cycle":          cycle_idx,
                        "Voltage (V)":    [],
                        "Current (A)":    [],
                        "Time (s)":       [],
                        "Resistance (Ω)": [],
                        "Power (W)":      [],
                    }
                    cycle_curves.append(live)
                    self.sweep_data = cycle_curves

                    for pass_idx, (pass_cmd, pass_pts) in enumerate(passes):
                        if self._stop_flag:
                            self._safe_abort(); break

                        self.inst.write("defbuffer1.clear()")
                        time.sleep(0.05)
                        self.inst.write(pass_cmd)
                        tsp_check(f"sweep_cmd[{pass_idx}]")
                        self.inst.write("trigger.model.initiate()")
                        tsp_check(f"initiate[{pass_idx}]")

                        last    = 0
                        t0      = time.time()
                        timeout = pass_pts * (float(p["NPLC"].get())/50.0 +
                                              float(delay)) * 5 + 30

                        while True:
                            if self._stop_flag:
                                self._safe_abort(); break
                            if time.time() - t0 > timeout:
                                raise RuntimeError(
                                    f"Pass {pass_idx+1} timed out "
                                    f"(collected {last}/{pass_pts} pts)")

                            try:
                                cnt = int(float(
                                    self.inst.query("print(defbuffer1.n)").strip()))
                            except ValueError:
                                self.inst.write("*CLS"); time.sleep(0.05); continue

                            if cnt > last:
                                s1, e1 = last+1, cnt
                                nsrc = [float(x) for x in self.inst.query(
                                    f"printbuffer({s1},{e1},defbuffer1.sourcevalues)"
                                ).split(",")]
                                nmeas= [float(x) for x in self.inst.query(
                                    f"printbuffer({s1},{e1},defbuffer1.readings)"
                                ).split(",")]
                                nts  = [float(x) for x in self.inst.query(
                                    f"printbuffer({s1},{e1},defbuffer1.relativetimestamps)"
                                ).split(",")]
                                nv = nsrc  if is_voltage else nmeas
                                ni = nmeas if is_voltage else nsrc
                                live["Voltage (V)"].extend(nv)
                                live["Current (A)"].extend(ni)
                                live["Time (s)"].extend(nts)
                                live["Resistance (Ω)"].extend(
                                    [abs(v/i) if i else float("inf")
                                     for v, i in zip(nv, ni)])
                                live["Power (W)"].extend(
                                    [abs(v*i) for v, i in zip(nv, ni)])
                                last = cnt
                                self.root.after(0, self._replot)

                            done = (cycle_idx * len(steps) + step_idx) * pts_total + last
                            total_all = n_cycles * len(steps) * pts_total
                            pct = min(int(100 * done / max(total_all, 1)), 99)
                            self.root.after(0, lambda pv=pct: self.progress.update(pv))

                            if cnt >= pass_pts:
                                break
                            time.sleep(0.05)

                    self._safe_abort()

                self._cycle_runs.append(list(cycle_curves))
                cycle_name = self._get_cycle_name(cycle_idx)
                self._cycle_names.append(cycle_name)
                self.sweep_data = []
                lbl_txt = (f"{cycle_name} complete  "
                           f"·  {len(cycle_curves)} curve{'s' if len(cycle_curves)!=1 else ''}")
                tags_txt = "  ".join(f"✓ {n}" for n in self._cycle_names)
                self.root.after(0, lambda t=lbl_txt, tg=tags_txt: (
                    self.cyc_status_lbl.config(text=t)
                    if hasattr(self, "cyc_status_lbl") else None,
                    self._cyc_tags_lbl.config(text=tg)
                    if hasattr(self, "_cyc_tags_lbl") else None,
                ))
                self.root.after(0, self._replot)

            self.inst.write("smu.source.output = smu.OFF")
            if p["Stepper"].get():
                self.inst.write("node[2].smu.source.output = node[2].smu.OFF")
            self.inst.close()
            self._set_conn_status(False)

            self._update_next_btn(waiting=False)

            if self._stop_flag:
                self.root.after(0, lambda: (
                    self.status.set("stopped"),
                    self.progress.stop(False)
                ))
            else:
                n_c = len(self._cycle_runs)
                self.root.after(0, lambda nc=n_c: (
                    self.status.set("done"),
                    self.progress.stop(True)
                ))
                self.root.after(0, self._auto_save)
                if any([self.an_vth.get(), self.an_ss.get(),
                        self.an_ionoff.get(), self.an_hyst.get()]):
                    self.root.after(300, self._run_analysis_iv)

        except Exception as e:
            err = str(e)
            self.root.after(0, lambda: (
                messagebox.showerror("Instrument Error", err),
                self.status.set("error"),
                self.progress.stop(False)
            ))
        finally:
            self.is_running = False
            self._stop_flag = False
            self._set_conn_status(False)

    def _extract_vth_iv(self, x, y):
        """Vth via linear extrapolation at max gm (√Id vs V)."""
        try:
            x = np.array(x, dtype=float)
            y = np.abs(np.array(y, dtype=float))
            order = np.argsort(x); x, y = x[order], y[order]
            if np.sum(y > 0) < 4: return float("nan")
            sq = np.sqrt(np.clip(y, 0, None))
            gm = np.gradient(sq, x)
            idx = int(np.argmax(gm))
            return float("nan") if gm[idx] == 0 else float(x[idx] - sq[idx]/gm[idx])
        except Exception:
            return float("nan")

    def _calc_ss_iv(self, x, y):
        """Subthreshold swing in mV/decade."""
        try:
            x = np.array(x, dtype=float)
            y = np.abs(np.array(y, dtype=float))
            order = np.argsort(x); x, y = x[order], y[order]
            valid = y > 1e-15
            if np.sum(valid) < 4: return float("nan")
            log_y = np.log10(y[valid])
            slope = np.gradient(log_y, x[valid])
            best  = float(np.max(slope))
            return float("nan") if best <= 0 else float(1000.0 / best)
        except Exception:
            return float("nan")

    def _calc_ionoff_iv(self, y):
        """Ion/Ioff ratio."""
        try:
            y = np.abs(np.array(y, dtype=float))
            v = y > 0
            if np.sum(v) < 2: return float("nan")
            return float(np.max(y[v]) / np.min(y[v]))
        except Exception:
            return float("nan")

    def _calc_hyst_iv(self, curve):
        """Hysteresis area between forward and reverse sweep halves."""
        try:
            x = np.array(curve.get("Voltage (V)", []), dtype=float)
            y = np.abs(np.array(curve.get("Current (A)", []), dtype=float))
            n = len(x) // 2
            if n < 3: return float("nan")
            xf, yf = x[:n], y[:n]
            xr_r = x[n:][::-1]; yr_r = y[n:][::-1]
            yr_i = np.interp(xf, xr_r, yr_r)
            return float(np.trapz(np.abs(yf - yr_i), xf))
        except Exception:
            return float("nan")

    def _run_analysis_iv(self):
        src = (self._cycle_runs[-1] if self._cycle_runs else self.sweep_data)
        if not src:
            if hasattr(self, "an_result_lbl"):
                self.an_result_lbl.config(
                    text="No data — run a sweep first.", fg=_THEME["DANGER"])
            return

        curve = src[0]
        x = curve.get("Voltage (V)", [])
        y = curve.get("Current (A)", [])
        eng = mticker.EngFormatter(sep="")
        parts = []

        if self.an_vth.get():
            v = self._extract_vth_iv(x, y)
            parts.append(f"Vth = {v:.3g} V" if not np.isnan(v) else "Vth = N/A")

        if self.an_ss.get():
            v = self._calc_ss_iv(x, y)
            parts.append(f"SS = {v:.0f} mV/dec" if not np.isnan(v) else "SS = N/A")

        if self.an_ionoff.get():
            v = self._calc_ionoff_iv(y)
            parts.append(f"Ion/Ioff = {eng(v)}" if not np.isnan(v) else "Ion/Ioff = N/A")

        if self.an_hyst.get():
            v = self._calc_hyst_iv(curve)
            parts.append(f"Hyst = {eng(v)} A·V" if not np.isnan(v) else "Hyst = N/A (need dual sweep)")

        result = "    ".join(parts) if parts else "No extractors enabled."
        if hasattr(self, "an_result_lbl"):
            self.an_result_lbl.config(text=result, fg=_THEME["ACCENT2"])

        if len(self._cycle_runs) > 1 and hasattr(self, "an_trend_lbl"):
            trend_parts = []
            vth_list, ion_list = [], []
            for cyc in self._cycle_runs:
                if not cyc: continue
                c = cyc[0]
                vth_list.append(self._extract_vth_iv(
                    c.get("Voltage (V)",[]), c.get("Current (A)",[])))
                yc = np.abs(np.array(c.get("Current (A)", []), dtype=float))
                ion_list.append(float(np.max(yc)) if len(yc) else float("nan"))
            vth_ok = [v for v in vth_list if not np.isnan(v)]
            if len(vth_ok) >= 2 and self.an_vth.get():
                dv = vth_ok[-1] - vth_ok[0]
                trend_parts.append(f"ΔVth = {dv:+.3g} V over {len(vth_ok)} cycles")
            ion_ok = [v for v in ion_list if not np.isnan(v)]
            if len(ion_ok) >= 2 and ion_ok[0] != 0:
                drift = (ion_ok[-1]/ion_ok[0] - 1) * 100
                trend_parts.append(f"Ion drift = {drift:+.1f}%")
            self.an_trend_lbl.config(
                text="Cycle trends:  " + "    ".join(trend_parts)
                if trend_parts else "")

    # ── CSV / PNG ─────────────────────────────────────────────────────────────

    def _make_filename(self):
        sid = self.sample_id.get().strip() or "DUT"
        return f"{sid}_{time.strftime('%Y%m%d_%H%M%S')}"

    def _build_csv_rows(self, source=None):
        p  = self.inputs
        ts = p["Timestamp"].get()
        hs = p["Stepper"].get()
        mr = p["MeasR"].get()
        mp = p["MeasP"].get()
        use_cycles = (self.cyc_enabled.get() and len(self._cycle_runs) > 1)

        if use_cycles:
            header = ["Voltage (V)"]
            for ci, cyc in enumerate(self._cycle_runs):
                cyc_name = (self._cycle_names[ci]
                            if ci < len(self._cycle_names) else f"Cyc {ci+1}")
                for curve in cyc:
                    sv = curve.get("step_val", 0)
                    sfx = f" [V2={sv:.3g}V]" if sv else ""
                    header.append(f"{cyc_name}{sfx}  I (A)")
                    if mr: header.append(f"{cyc_name}{sfx}  R (Ω)")
                    if mp: header.append(f"{cyc_name}{sfx}  P (W)")

            ref_x = self._cycle_runs[0][0].get("Voltage (V)", [])
            max_pts = max((max(len(cv.get("Voltage (V)", []))
                              for cv in cyc) for cyc in self._cycle_runs if cyc),
                         default=0)
            rows = []
            for idx in range(max_pts):
                row = [f"{ref_x[idx]:.9g}" if idx < len(ref_x) else ""]
                for cyc in self._cycle_runs:
                    for curve in cyc:
                        iv = curve.get("Current (A)", [])
                        rv = curve.get("Resistance (Ω)", [])
                        pw = curve.get("Power (W)", [])
                        row.append(f"{iv[idx]:.9g}" if idx < len(iv) else "")
                        if mr: row.append(f"{rv[idx]:.6g}" if idx < len(rv) else "")
                        if mp: row.append(f"{pw[idx]:.9g}" if idx < len(pw) else "")
                rows.append(row)
            return header, rows
        else:
            header = []
            if ts: header.append("Time (s)")
            if hs: header.append("Step_Value (V)")
            header += ["Voltage (V)", "Current (A)"]
            if mr: header.append("Resistance (Ω)")
            if mp: header.append("Power (W)")
            rows = []
            data = source if source is not None else self.sweep_data
            for curve in data:
                v = curve.get("Voltage (V)", [])
                i = curve.get("Current (A)", [])
                t = curve.get("Time (s)", [])
                sv = curve.get("step_val", 0)
                for idx in range(min(len(v), len(i))):
                    row = []
                    if ts: row.append(f"{t[idx]:.6f}" if idx < len(t) else "")
                    if hs: row.append(f"{sv:.6f}")
                    row += [f"{v[idx]:.9g}", f"{i[idx]:.9g}"]
                    if mr:
                        try:    row.append(f"{abs(v[idx]/i[idx]):.6g}")
                        except: row.append("inf")
                    if mp: row.append(f"{abs(v[idx]*i[idx]):.9g}")
                    rows.append(row)
            return header, rows

    def _write_csv(self, path, source=None):
        header, rows = self._build_csv_rows(source)
        p = self.inputs
        n_cyc = len(self._cycle_runs)
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["# Keithley 2450  Lab View"])
            w.writerow(["# Sample ID",  self.sample_id.get()])
            w.writerow(["# Date/Time",  time.strftime("%d/%m/%Y  %H:%M:%S")])
            w.writerow(["# Mode",       p["Mode"].get()])
            w.writerow(["# NPLC",       p["NPLC"].get()])
            w.writerow(["# Compliance", p["Limit"].get()])
            w.writerow(["# Terminals",  p["Terminals"].get()])
            w.writerow(["# Sense",      p["Sense"].get()])
            if n_cyc > 1:
                w.writerow(["# Cycles", n_cyc])
            if hasattr(self, "an_result_lbl"):
                result = self.an_result_lbl.cget("text")
                if result and result != "—":
                    w.writerow(["# Extracted", result])
            w.writerow([])
            w.writerow(header)
            w.writerows(rows)

    def _auto_save(self):
        try:
            base = self._make_filename()
            self._write_csv(os.path.join(self.save_dir, f"{base}.csv"))
            self._save_fig_light(self.fig,
                                 os.path.join(self.save_dir, f"{base}.png"))
            n_cyc = len(self._cycle_runs)
            extra = f"  ({n_cyc} cycles)" if n_cyc > 1 else ""
            self.root.after(0, lambda:
                self.status.msg(
                    f"Auto-saved{extra}  →  {self.save_dir}", SUCCESS))
        except Exception as e:
            self.root.after(0, lambda err=str(e):
                messagebox.showwarning("Auto-save Failed",
                                       f"Could not auto-save:\n{err}"))

    def _save_csv(self):
        if not self.sweep_data and not self._cycle_runs:
            messagebox.showwarning("No Data","Run a sweep first."); return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv", initialdir=self.save_dir,
            initialfile=f"{self._make_filename()}.csv",
            filetypes=[("CSV","*.csv"),("All","*.*")])
        if not path: return
        try:
            self._write_csv(path)
            messagebox.showinfo("Saved", f"CSV saved:\n{path}")
        except Exception as e:
            messagebox.showerror("Save Error", str(e))

    def _save_cycle_csv(self):
        if not self._cycle_runs:
            messagebox.showwarning("No Cycle Data",
                "No cycle data available.\n"
                "Enable cycling and run a measurement first.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv", initialdir=self.save_dir,
            initialfile=f"{self._make_filename()}_cycles.csv",
            filetypes=[("CSV","*.csv"),("All","*.*")])
        if not path: return
        try:
            p = self.inputs
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["# Keithley 2450  Lab View  —  Cycle Export"])
                w.writerow(["# Sample ID",  self.sample_id.get()])
                w.writerow(["# Mode",       p["Mode"].get()])
                w.writerow(["# Cycles",     len(self._cycle_runs)])
                w.writerow(["# Date",       time.strftime("%d/%m/%Y  %H:%M:%S")])
                if hasattr(self, "an_result_lbl"):
                    res = self.an_result_lbl.cget("text")
                    if res and res != "—":
                        w.writerow(["# Extracted", res])
                if hasattr(self, "an_trend_lbl"):
                    trend = self.an_trend_lbl.cget("text")
                    if trend:
                        w.writerow(["# Trend", trend])
                w.writerow([])
                x_col  = self.ax_x.get()
                header = [x_col]
                for ci, cyc in enumerate(self._cycle_runs):
                    cyc_name = (self._cycle_names[ci]
                                if ci < len(self._cycle_names)
                                else f"Cyc{ci+1}")
                    for curve in cyc:
                        sv = curve.get("step_val", 0)
                        sfx = f"  [V₂={sv:.3g}V]" if sv else ""
                        header.append(f"{cyc_name}  I{sfx} (A)")
                        if p["MeasR"].get():
                            header.append(f"{cyc_name}  R{sfx} (Ω)")
                w.writerow(header)
                ref_x = self._cycle_runs[0][0].get("Voltage (V)", []) if self._cycle_runs else []
                max_pts = max(
                    (max(len(cv.get("Voltage (V)",[]))
                         for cv in cyc) for cyc in self._cycle_runs if cyc),
                    default=0)
                for i in range(max_pts):
                    row = [f"{ref_x[i]:.6g}" if i < len(ref_x) else ""]
                    for cyc in self._cycle_runs:
                        for curve in cyc:
                            iv = curve.get("Current (A)", [])
                            rv = curve.get("Resistance (Ω)", [])
                            row.append(f"{iv[i]:.9g}" if i < len(iv) else "")
                            if p["MeasR"].get():
                                row.append(f"{rv[i]:.6g}" if i < len(rv) else "")
                    w.writerow(row)
            messagebox.showinfo("Saved", f"Cycle CSV saved:\n{path}")
        except Exception as e:
            messagebox.showerror("Save Error", str(e))

    def _save_png(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".png", initialdir=self.save_dir,
            initialfile=f"{self._make_filename()}.png",
            filetypes=[("PNG","*.png"),("All","*.*")])
        if path:
            try:
                self._save_fig_light(self.fig, path)
                self.status.msg(f"PNG saved: {os.path.basename(path)}", SUCCESS)
            except Exception as e:
                messagebox.showerror("Save Error", str(e))


# ═══════════════════════════════════════════════════════════════════════════════
#  FET CHARACTERISATION WINDOW
# ═══════════════════════════════════════════════════════════════════════════════

class FETWindow(tk.Toplevel):
    MODES     = ["Transfer Curve  (Id vs Vg)", "Output Curve  (Id vs Vds)"]
    TERMINALS = ["Gate", "Drain", "Source"]
    SMU_OPTS  = ["NONE","GNDU",
                 "SMU1 (Node 1 — Master)","SMU2 (Node 2 — Slave)"]
    OP_MODES  = ["Open","Voltage Bias","Voltage Linear Sweep",
                 "Voltage List Sweep","Voltage Log Sweep","Voltage Step",
                 "Current Bias","Current Linear Sweep","Current List Sweep",
                 "Current Log Sweep","Current Step","Common"]
    RANGE_OPTS= ["Best Fixed","Auto","200mV","2V","20V","200V"]
    OVP_OPTS  = ["OFF","20V","40V","60V","100V","200V"]

    _TSP    = {"SMU1 (Node 1 — Master)":"smu",
               "SMU2 (Node 2 — Slave)": "node[2].smu"}
    _NO_OP  = {"NONE","GNDU"}
    _SWEEP  = {"Voltage Linear Sweep","Current Linear Sweep",
               "Voltage Step","Current Step"}
    _LOG    = {"Voltage Log Sweep","Current Log Sweep"}
    _BIAS   = {"Voltage Bias","Current Bias"}
    _LIST   = {"Voltage List Sweep","Current List Sweep"}

    TERM_COL = {"Gate":"#16A34A","Drain":"#1D4ED8","Source":"#BE185D"}
    TERM_BG_LIGHT = {"Gate":"#F0FDF4","Drain":"#EFF6FF","Source":"#FDF2F8"}
    TERM_BG_DARK  = {"Gate":"#0A1F0F","Drain":"#0A1220","Source":"#1A0A14"}

    @property
    def TERM_BG(self):
        return self.TERM_BG_DARK if self.app._dark_mode else self.TERM_BG_LIGHT

    def __init__(self, parent_app):
        super().__init__(parent_app.root)
        self.app        = parent_app
        self.title("Keithley 2450  —  FET Characterisation")
        self.configure(bg=_THEME["BG"])
        self.resizable(True, True)
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.state("zoomed")
        try: self.attributes("-zoomed", True)
        except Exception: pass
        setup_styles()
        self.option_add("*TCombobox*Listbox.background",       _THEME["PANEL"])
        self.option_add("*TCombobox*Listbox.foreground",       _THEME["TEXT"])
        self.option_add("*TCombobox*Listbox.selectBackground", _THEME["ACCENT3"])
        self.option_add("*TCombobox*Listbox.selectForeground", _THEME["TEXT"])
        self.is_running = False
        self._stop      = False
        self.inst       = None
        self.curves     = []
        self._compare   = CompareManager(
            owner=self,
            replot_cb=lambda: self._replot() if hasattr(self, 'ax') else None,
            save_dir_cb=lambda: self.app.save_dir,
            mode="FET")
        self.v           = {t: {} for t in self.TERMINALS}
        self._pf         = {t: {} for t in self.TERMINALS}
        self._list_vals  = {t: [] for t in self.TERMINALS}

        self._cycle_data       = []
        self._cycle_names_fet  = []
        self._current_cycle    = 0
        self._overlay_data     = None
        self.cyc_enabled       = tk.BooleanVar(value=False)
        self.cyc_n             = tk.StringVar(value="5")
        self.cyc_rest          = tk.StringVar(value="0")
        self.cyc_show_all      = tk.BooleanVar(value=True)
        self.cyc_gradient      = tk.BooleanVar(value=True)
        self.cyc_stats         = tk.BooleanVar(value=True)
        self.cyc_overlay       = tk.BooleanVar(value=True)
        self.cyc_custom_names  = tk.BooleanVar(value=False)

        self.str_enabled    = tk.BooleanVar(value=False)
        self.str_v          = tk.StringVar(value="20")
        self.str_dur        = tk.StringVar(value="60")
        self.str_int        = tk.StringVar(value="10")
        self.str_term       = tk.StringVar(value="Gate")

        self.an_vth         = tk.BooleanVar(value=True)
        self.an_ss          = tk.BooleanVar(value=True)
        self.an_ionoff      = tk.BooleanVar(value=True)
        self.an_mob         = tk.BooleanVar(value=False)
        self.an_hyst        = tk.BooleanVar(value=False)
        self.an_vds_lin     = tk.StringVar(value="0.1")
        self.an_cox         = tk.StringVar(value="0")
        self.an_W           = tk.StringVar(value="1000")
        self.an_L           = tk.StringVar(value="50")

        self.mq_avg         = tk.StringVar(value="1")
        self.mq_settle      = tk.BooleanVar(value=False)
        self.mq_tol         = tk.StringVar(value="1.0")
        self.mq_maxatt      = tk.StringVar(value="5")
        self.mq_precheck    = tk.BooleanVar(value=True)

        self._build_styles()
        self._build_ui()
        self._start_fet_conn_checker()
        if self.app._dark_mode:
            self.configure(bg=_THEME["BG"])
            self.app._apply_theme(self)
            panel = _THEME["PANEL"]
            if hasattr(self, "fig"):
                self.fig.patch.set_facecolor(panel)
            if hasattr(self, "canvas"):
                self.canvas.get_tk_widget().config(bg=panel)
                self.fig.clear()
                self.fig.patch.set_facecolor(panel)
                self.ax = self.fig.add_subplot(111)
                self.fig.subplots_adjust(left=0.11, right=0.95, top=0.93, bottom=0.10)
                _style_plot(self.fig, self.ax,
                            title="Configure terminals and press RUN")
                self.canvas.draw()

    def _build_styles(self):
        s = ttk.Style()
        for name, bg, abg in [("FRun.TButton", "#16A34A","#15803D"),
                               ("FStop.TButton","#DC2626","#B91C1C")]:
            s.configure(name, background=bg, foreground="white",
                        font=FONT_H3, padding=(10,6), relief="flat")
            s.map(name, background=[("active",abg)])

    def _start_fet_conn_checker(self):
        def _ping():
            if self.is_running:
                pass
            else:
                addr = self.addr_var.get().strip() if hasattr(self, "addr_var") else ""
                try:
                    rm   = pyvisa.ResourceManager()
                    inst = rm.open_resource(addr)
                    inst.timeout = 1500
                    idn = inst.query("*IDN?").strip()
                    inst.close()
                    rm.close()
                    connected = bool(idn)
                except Exception:
                    connected = False
                def _upd(c=connected, a=addr):
                    if not hasattr(self, "_conn_lbl"): return
                    if c:
                        short = a.split("::")[-2] if "::" in a else a[:18]
                        self._conn_lbl.config(text=f"● SMU: {short}", fg="#4ADE80")
                    else:
                        self._conn_lbl.config(text="● SMU: Disconnected", fg="#EF4444")
                try: self.after(0, _upd)
                except Exception: pass
            try:
                self._fet_conn_id = self.after(3000, _ping)
            except Exception:
                pass
        self._fet_conn_id = self.after(1000, _ping)

    def _rebuild_cell_bgs(self):
        try:
            for widget in self._inner.winfo_descendants():
                wc = widget.winfo_class()
                if wc not in ("Frame",): continue
                try:
                    cb = widget.cget("bg")
                    swap = {
                        "#F0FDF4": self.TERM_BG["Gate"],
                        "#EFF6FF": self.TERM_BG["Drain"],
                        "#FDF2F8": self.TERM_BG["Source"],
                        "#0A1F0F": self.TERM_BG["Gate"],
                        "#0A1220": self.TERM_BG["Drain"],
                        "#1A0A14": self.TERM_BG["Source"],
                    }
                    if cb in swap:
                        widget.config(bg=swap[cb])
                except Exception:
                    pass
        except Exception:
            pass

    def _toggle_dark_fet(self):
        self.app._dark_mode = not self.app._dark_mode
        palette = DARK if self.app._dark_mode else LIGHT
        _THEME.update(palette)
        _sync_globals()
        _cfg.set("dark_mode", self.app._dark_mode)
        setup_styles()
        self.option_add("*TCombobox*Listbox.background",       _THEME["PANEL"])
        self.option_add("*TCombobox*Listbox.foreground",       _THEME["TEXT"])
        self.option_add("*TCombobox*Listbox.selectBackground", _THEME["ACCENT3"])
        self.option_add("*TCombobox*Listbox.selectForeground", _THEME["TEXT"])
        self.configure(bg=_THEME["BG"])
        self.app._apply_theme(self)
        if hasattr(self, "_inner"):
            self._rebuild_cell_bgs()
        panel = _THEME["PANEL"]
        if hasattr(self, "fig"):
            self.fig.patch.set_facecolor(panel)
        if hasattr(self, "canvas"):
            self.canvas.get_tk_widget().config(bg=panel)
            if self.curves or self._cycle_data or self._overlay_data:
                self._replot()
            else:
                self.fig.clear()
                self.fig.patch.set_facecolor(panel)
                self.ax = self.fig.add_subplot(111)
                self.fig.subplots_adjust(left=0.11, right=0.95, top=0.93, bottom=0.10)
                _style_plot(self.fig, self.ax, title="Configure terminals and press RUN")
                self.canvas.draw()
        if hasattr(self, "_dark_btn"):
            if self.app._dark_mode:
                self._dark_btn.config(text="🌙  Dark", bg="#1A2535", fg="#FCD34D")
            else:
                self._dark_btn.config(text="☀  Light", bg="#243F6A", fg="#93C5FD")

    def _build_ui(self):
        hdr = tk.Frame(self, bg=ACCENT, height=50)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        logo = tk.Frame(hdr, bg=ACCENT); logo.pack(side="left", padx=14, pady=6)
        tk.Label(logo, text="FET Characterisation", bg=ACCENT, fg="white",
                 font=FONT_H1).pack(anchor="w")
        tk.Label(logo, text="Transfer  &  Output Curve Analysis",
                 bg=ACCENT, fg="#7EB6E0", font=FONT_SM).pack(anchor="w")
        self._conn_lbl = tk.Label(
            hdr, text="● SMU: Disconnected",
            bg=ACCENT, fg="#EF4444", font=FONT_MONO)
        self._conn_lbl.pack(side="right", padx=(8, 14))

        self.status_lbl = tk.Label(hdr, text="", bg=ACCENT, fg="#4ADE80",
                                   font=FONT_MONO2)

        dm_label = "🌙  Dark" if self.app._dark_mode else "☀  Light"
        dm_bg    = "#1A2535" if self.app._dark_mode else "#243F6A"
        dm_fg    = "#FCD34D" if self.app._dark_mode else "#93C5FD"
        self._dark_btn = tk.Button(
            hdr, text=dm_label, font=FONT_SM,
            bg=dm_bg, fg=dm_fg, relief="flat", bd=0,
            activebackground="#243040", cursor="hand2",
            padx=10, pady=3, command=self._toggle_dark_fet)
        self._dark_btn.pack(side="right", padx=(0, 8))

        mode_bar = tk.Frame(self, bg=_THEME["PANEL2"],
                            highlightbackground=_THEME["BORDER"], highlightthickness=1)
        mode_bar.pack(fill="x", padx=8, pady=(6,0))
        tk.Label(mode_bar, text="Measurement Mode:", bg=_THEME["PANEL2"], fg=_THEME["TEXT2"],
                 font=FONT_H3).pack(side="left", padx=12, pady=7)
        self.mode_var = tk.StringVar(value=self.MODES[0])
        for m in self.MODES:
            tk.Radiobutton(mode_bar, text=m, variable=self.mode_var, value=m,
                           bg=_THEME["PANEL2"], fg=_THEME["TEXT"], selectcolor=_THEME["PANEL2"],
                           activebackground=_THEME["PANEL2"], font=FONT_BODY,
                           command=self._mode_changed).pack(side="left", padx=8, pady=7)

        body = tk.Frame(self, bg=_THEME["BG"])
        body.pack(fill="both", expand=True, padx=8, pady=6)

        lf = tk.Frame(body, bg=_THEME["BG"], width=730)
        lf.pack(side="left", fill="y"); lf.pack_propagate(False)
        cvs = tk.Canvas(lf, bg=_THEME["BG"], highlightthickness=0, width=720)
        vsb = ttk.Scrollbar(lf, orient="vertical", command=cvs.yview)
        self._inner = tk.Frame(cvs, bg=_THEME["BG"])
        self._inner.bind("<Configure>",
                         lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        self._win = cvs.create_window((0,0), window=self._inner, anchor="nw")
        cvs.bind("<Configure>",
                 lambda e: cvs.itemconfig(self._win, width=e.width))
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        cvs.bind_all("<MouseWheel>",
                     lambda e: cvs.yview_scroll(int(-1*(e.delta/120)),"units"))
        self._build_table(self._inner)

        rf = tk.Frame(body, bg=_THEME["BG"])
        rf.pack(side="left", fill="both", expand=True, padx=(8,0))
        self._build_graph(rf)

    def _build_table(self, p):
        LBL_W = 22

        hdr = tk.Frame(p, bg=_THEME["PANEL"],
                       highlightbackground=_THEME["BORDER"], highlightthickness=1)
        hdr.pack(fill="x", pady=(0,2))
        tk.Label(hdr, text="Parameter", bg=_THEME["PANEL2"], fg=_THEME["TEXT"],
                 font=FONT_H3, width=LBL_W, anchor="w").pack(
                 side="left", padx=(6,0), ipady=6)
        for t in self.TERMINALS:
            col = self.TERM_COL[t]
            f = tk.Frame(hdr, bg=col)
            f.pack(side="left", fill="both", expand=True, padx=2, pady=3)
            tk.Label(f, text=t, bg=col, fg="white",
                     font=FONT_H2).pack(expand=True, pady=5)

        def sec(label):
            f = tk.Frame(p, bg=_THEME["ACCENT2"]); f.pack(fill="x", pady=(5,1))
            tk.Label(f, text=f"  {label}", bg=_THEME["ACCENT2"], fg="white",
                     font=FONT_H3).pack(side="left", pady=4)

        def row(label, section=False):
            r = tk.Frame(p, bg=_THEME["PANEL"],
                         highlightbackground=_THEME["BORDER"], highlightthickness=1)
            r.pack(fill="x", pady=1)
            bg = _THEME["PANEL"] if not section else _THEME["PANEL2"]
            tk.Label(r, text=label, bg=bg, fg=_THEME["TEXT"],
                     font=FONT_BODY if not section else FONT_H3,
                     width=LBL_W, anchor="w").pack(
                     side="left", ipady=4, padx=(6,0))
            cells = {}
            for t in self.TERMINALS:
                col = self.TERM_COL[t]
                f = tk.Frame(r, bg=_THEME["PANEL"],
                             highlightbackground=col, highlightthickness=1)
                f.pack(side="left", fill="both", expand=True, padx=2, pady=2)
                cells[t] = f
            return cells

        sec("Instrument Assignment")
        cr = row("Instrument", section=True)
        defaults_smu = {"Gate":"SMU1 (Node 1 — Master)",
                        "Drain":"SMU2 (Node 2 — Slave)",
                        "Source":"GNDU"}
        for t in self.TERMINALS:
            self.v[t]["smu"] = tk.StringVar(value=defaults_smu[t])
            ttk.Combobox(cr[t], textvariable=self.v[t]["smu"],
                         values=self.SMU_OPTS, state="readonly",
                         width=20, font=FONT_SM).pack(fill="x", padx=4, pady=4)
            self.v[t]["smu"].trace_add("write",
                lambda *_, term=t: self._on_smu_change(term))

        sec("Force")
        op_r = row("Operation Mode")
        op_defaults = {"Gate":"Voltage Linear Sweep",
                       "Drain":"Voltage Bias","Source":"Voltage Bias"}
        for t in self.TERMINALS:
            self.v[t]["op"] = tk.StringVar(value=op_defaults[t])
            ttk.Combobox(op_r[t], textvariable=self.v[t]["op"],
                         values=self.OP_MODES, state="readonly",
                         width=20, font=FONT_SM).pack(fill="x", padx=4, pady=4)
            self.v[t]["op"].trace_add("write",
                lambda *_, term=t: self._on_op_change(term))

        def entry_row(label, key, default_d, unit="V"):
            r = row(label)
            for t in self.TERMINALS:
                self.v[t][key] = tk.StringVar(value=default_d.get(t,"0"))
                f = tk.Frame(r[t], bg=_THEME["PANEL"]); f.pack(fill="x", padx=4, pady=3)
                ttk.Entry(f, textvariable=self.v[t][key],
                          width=10, font=FONT_SM).pack(side="left")
                tk.Label(f, text=unit, bg=_THEME["PANEL"], fg=_THEME["TEXT3"],
                         font=FONT_SM).pack(side="left", padx=3)
            return r

        bias_r = entry_row("Bias", "bias", {"Gate":"0","Drain":"10","Source":"0"})
        for t in self.TERMINALS: self._pf[t]["bias_row"] = bias_r[t]

        list_r = row("List Values")
        for t in self.TERMINALS:
            cell = list_r[t]
            lbl = tk.Label(cell, text="(empty)", bg=_THEME["SEP"], fg=_THEME["TEXT3"],
                           font=FONT_SM, anchor="w")
            lbl.pack(side="left", padx=5, fill="x", expand=True)
            btn = ttk.Button(cell, text="Edit…",
                             command=lambda term=t, lb=lbl:
                                 self._open_list_editor(term, lb))
            btn.pack(side="right", padx=4, pady=4)
            self._pf[t]["list_lbl"] = lbl
            self._pf[t]["list_row"] = list_r[t]
            cell.configure(bg=_THEME["SEP"])
            lbl.configure(state="disabled"); btn.configure(state="disabled")

        start_r = entry_row("Start", "start", {"Gate":"0","Drain":"10","Source":"0"})
        stop_r  = entry_row("Stop",  "stop",  {"Gate":"8","Drain":"10","Source":"0"})
        step_r  = entry_row("Step",  "step",  {"Gate":"0.1","Drain":"1","Source":"0"})

        pts_r = row("Points")
        for t in self.TERMINALS:
            self.v[t]["pts"]      = tk.StringVar(value="81" if t=="Gate" else "76")
            self.v[t]["pts_auto"] = tk.BooleanVar(value=True)
            f = tk.Frame(pts_r[t], bg=_THEME["PANEL"]); f.pack(fill="x", padx=4, pady=3)

            ent = ttk.Entry(f, textvariable=self.v[t]["pts"], width=8, font=FONT_SM)
            ent.pack(side="left")
            self._pf[t]["pts_entry"] = ent

            auto_lbl = tk.Label(f, text="auto", bg=_THEME["PANEL2"], fg=_THEME["ACCENT2"],
                                font=FONT_SM, padx=4)
            auto_lbl.pack(side="left", padx=(2,0))
            self._pf[t]["pts_auto_lbl"] = auto_lbl

            def _make_toggle(term, entry_widget, lbl_widget):
                def _toggle(*_):
                    if self.v[term]["pts_auto"].get():
                        entry_widget.configure(state="disabled")
                        lbl_widget.pack(side="left", padx=(2,0))
                        self._recalc_pts_for(term)
                    else:
                        lbl_widget.pack_forget()
                        entry_widget.configure(state="normal")
                return _toggle

            toggle_fn = _make_toggle(t, ent, auto_lbl)
            self.v[t]["pts_auto"].trace_add("write", lambda *_, fn=toggle_fn: fn())

            chk = tk.Checkbutton(f, text="Auto", variable=self.v[t]["pts_auto"],
                                  bg=_THEME["PANEL"], fg=_THEME["TEXT2"], activebackground=_THEME["PANEL"],
                                  selectcolor=_THEME["PANEL"], relief="flat",
                                  highlightthickness=0, font=FONT_SM)
            chk.pack(side="left", padx=(6,0))

            for k in ("start","stop","step"):
                self.v[t][k].trace_add("write",
                    lambda *_, term=t: self._recalc_pts_for(term))

            toggle_fn()

        dual_r = row("Dual Sweep")
        for t in self.TERMINALS:
            self.v[t]["dual"] = tk.BooleanVar(value=False)
            f = tk.Frame(dual_r[t], bg=_THEME["PANEL"]); f.pack(fill="x", padx=4, pady=3)
            tk.Checkbutton(f, variable=self.v[t]["dual"], bg=_THEME["PANEL"],
                           activebackground=_THEME["PANEL"], selectcolor=_THEME["PANEL"],
                           relief="flat", highlightthickness=0).pack(side="left")

        for t in self.TERMINALS:
            self._pf[t]["start_row"] = start_r[t]
            self._pf[t]["stop_row"]  = stop_r[t]
            self._pf[t]["step_row"]  = step_r[t]
            self._pf[t]["pts_row"]   = pts_r[t]

        rng_r = row("Source Range")
        for t in self.TERMINALS:
            self.v[t]["range"] = tk.StringVar(value="Best Fixed")
            ttk.Combobox(rng_r[t], textvariable=self.v[t]["range"],
                         values=self.RANGE_OPTS, state="readonly",
                         width=13, font=FONT_SM).pack(fill="x", padx=4, pady=4)

        comp_r = entry_row("Compliance",      "comp",
                           {"Gate":"0.01","Drain":"0.5","Source":"0.1"}, unit="A")
        delay_r= entry_row("Source Delay",    "delay",
                           {"Gate":"0.05","Drain":"0.05","Source":"0.05"}, unit="s")
        pod_r  = entry_row("Power-On Delay",  "pod",
                           {"Gate":"0","Drain":"0","Source":"0"}, unit="s")
        ovp_r  = row("Overvoltage Protect.")
        for t in self.TERMINALS:
            self.v[t]["ovp"] = tk.StringVar(value="OFF")
            ttk.Combobox(ovp_r[t], textvariable=self.v[t]["ovp"],
                         values=self.OVP_OPTS, state="readonly",
                         width=13, font=FONT_SM).pack(fill="x", padx=4, pady=4)

        sec("Measure")

        def chk_row(label, key, defaults):
            r = row(label)
            for t in self.TERMINALS:
                self.v[t][key] = tk.BooleanVar(value=defaults.get(t, True))
                f = tk.Frame(r[t], bg=_THEME["PANEL"]); f.pack(fill="x", padx=4, pady=3)
                tk.Checkbutton(f, variable=self.v[t][key], bg=_THEME["PANEL"],
                               activebackground=_THEME["PANEL"], selectcolor=_THEME["PANEL"],
                               relief="flat", highlightthickness=0
                               ).pack(side="left")
            return r

        def col_name_row(label, key, defaults):
            r = row(label)
            for t in self.TERMINALS:
                self.v[t][key] = tk.StringVar(value=defaults.get(t,""))
                ttk.Entry(r[t], textvariable=self.v[t][key],
                          width=14, font=FONT_SM).pack(fill="x", padx=4, pady=3)

        chk_row("Current Measure",  "meas_i",
                {"Gate":False,"Drain":True,"Source":True})
        col_name_row("  Column Name",   "icol",
                     {"Gate":"GateI","Drain":"DrainI","Source":"SourceI"})
        meas_irange_r = row("  Current Range")
        for t in self.TERMINALS:
            self.v[t]["irange"] = tk.StringVar(value="Limited Auto")
            ttk.Combobox(meas_irange_r[t], textvariable=self.v[t]["irange"],
                         values=["Limited Auto","Auto","1nA","10nA","100nA",
                                 "1uA","10uA","100uA","1mA","10mA","100mA","1A"],
                         state="readonly", width=13,
                         font=FONT_SM).pack(fill="x", padx=4, pady=4)

        chk_row("Voltage Measure", "meas_v", {t:True for t in self.TERMINALS})
        vrep_r = row("  Report Value")
        for t in self.TERMINALS:
            self.v[t]["vrep"] = tk.StringVar(value="Programmed")
            ttk.Combobox(vrep_r[t], textvariable=self.v[t]["vrep"],
                         values=["Programmed","Actual"],
                         state="readonly", width=13,
                         font=FONT_SM).pack(fill="x", padx=4, pady=4)
        col_name_row("  Column Name", "vcol",
                     {"Gate":"GateV","Drain":"DrainV","Source":"SourceV"})

        sec("Instrument Settings")
        nplc_r = row("NPLC")
        for t in self.TERMINALS:
            self.v[t]["nplc"] = tk.StringVar(value="1")
            ttk.Entry(nplc_r[t], textvariable=self.v[t]["nplc"],
                      width=10, font=FONT_SM).pack(fill="x", padx=4, pady=3)

        ctrl = tk.Frame(p, bg=_THEME["PANEL"],
                        highlightbackground=_THEME["BORDER"], highlightthickness=1)
        ctrl.pack(fill="x", pady=(8,0))

        row1 = tk.Frame(ctrl, bg=_THEME["PANEL"]); row1.pack(fill="x", padx=10, pady=(8,4))
        tk.Label(row1, text="VISA Address", bg=_THEME["PANEL"], fg=_THEME["TEXT"],
                 font=FONT_BODY, width=16, anchor="w").pack(side="left")
        self.addr_var = tk.StringVar(value=self.app.master_addr)
        ttk.Entry(row1, textvariable=self.addr_var,
                  width=38, font=FONT_BODY).pack(side="left", padx=6)

        row2 = tk.Frame(ctrl, bg=_THEME["PANEL"]); row2.pack(fill="x", padx=10, pady=(0,4))
        tk.Label(row2, text="Input Jacks", bg=_THEME["PANEL"], fg=_THEME["TEXT"],
                 font=FONT_BODY, width=16, anchor="w").pack(side="left")
        self.terminals_var = tk.StringVar(value="Rear")
        ttk.Combobox(row2, textvariable=self.terminals_var,
                     values=["Rear", "Front"],
                     state="readonly", width=8,
                     font=FONT_BODY).pack(side="left", padx=(0, 18))
        tk.Label(row2, text="Sense Mode", bg=_THEME["PANEL"], fg=_THEME["TEXT"],
                 font=FONT_BODY).pack(side="left")
        self.sense_var = tk.StringVar(value="2-Wire (Local)")
        ttk.Combobox(row2, textvariable=self.sense_var,
                     values=["2-Wire (Local)","4-Wire (Remote)"],
                     state="readonly", width=16,
                     font=FONT_BODY).pack(side="left", padx=6)
        self.nplc_all = tk.StringVar(value="1")
        tk.Label(row2, text="NPLC (all)", bg=_THEME["PANEL"], fg=_THEME["TEXT"],
                 font=FONT_BODY).pack(side="left", padx=(20,4))
        ttk.Entry(row2, textvariable=self.nplc_all,
                  width=6, font=FONT_BODY).pack(side="left")
        ttk.Button(row2, text="Apply to all",
                   command=self._apply_nplc_all).pack(side="left", padx=6)

        mq_hdr = tk.Frame(ctrl, bg=_THEME["ACCENT2"]); mq_hdr.pack(fill="x")
        tk.Label(mq_hdr, text="  ⚙  Measurement Quality",
                 bg=_THEME["ACCENT2"], fg="white", font=FONT_SM).pack(side="left", pady=3)
        mq_row = tk.Frame(ctrl, bg=_THEME["PANEL"]); mq_row.pack(fill="x", padx=10, pady=(4,6))

        tk.Label(mq_row, text="Avg/pt", bg=_THEME["PANEL"], fg=_THEME["TEXT"], font=FONT_BODY).pack(side="left")
        ttk.Combobox(mq_row, textvariable=self.mq_avg,
                     values=["1","2","4","8","16"], state="readonly",
                     width=4, font=FONT_BODY).pack(side="left", padx=(3,16))

        tk.Checkbutton(mq_row, text="Settling check", variable=self.mq_settle,
                       bg=_THEME["PANEL"], fg=_THEME["TEXT"], activebackground=_THEME["PANEL"],
                       selectcolor=_THEME["PANEL"], relief="flat",
                       highlightthickness=0, font=FONT_BODY).pack(side="left")
        tk.Label(mq_row, text="Tol", bg=_THEME["PANEL"], fg=_THEME["TEXT2"], font=FONT_SM).pack(side="left", padx=(6,2))
        ttk.Entry(mq_row, textvariable=self.mq_tol, width=5, font=FONT_BODY).pack(side="left")
        tk.Label(mq_row, text="%", bg=_THEME["PANEL"], fg=_THEME["TEXT3"], font=FONT_SM).pack(side="left", padx=(1,4))
        tk.Label(mq_row, text="Max att", bg=_THEME["PANEL"], fg=_THEME["TEXT2"], font=FONT_SM).pack(side="left", padx=(4,2))
        ttk.Entry(mq_row, textvariable=self.mq_maxatt, width=4, font=FONT_BODY).pack(side="left", padx=(0,16))

        tk.Checkbutton(mq_row, text="Compliance pre-check", variable=self.mq_precheck,
                       bg=_THEME["PANEL"], fg=_THEME["TEXT"], activebackground=_THEME["PANEL"],
                       selectcolor=_THEME["PANEL"], relief="flat",
                       highlightthickness=0, font=FONT_BODY).pack(side="left")

        row3 = tk.Frame(ctrl, bg=_THEME["PANEL"]); row3.pack(fill="x", padx=10, pady=(0,8))
        ttk.Button(row3, text="▶  RUN", style="FRun.TButton",
                   command=self._run).pack(side="left", fill="x",
                                           expand=True, ipady=5, padx=(0,6))
        ttk.Button(row3, text="■  STOP", style="FStop.TButton",
                   command=self._do_stop).pack(side="left", ipady=5, ipadx=12)
        ttk.Button(row3, text="Export CSV",
                   command=self._export_csv).pack(side="left", padx=(10,3))
        ttk.Button(row3, text="Cycle CSV",
                   command=self._export_cycle_csv).pack(side="left", padx=(0,3))
        ttk.Button(row3, text="Save PNG",
                   command=self._save_png).pack(side="left")

        for t in self.TERMINALS:
            self._on_smu_change(t)
            self._on_op_change(t)

    def _on_smu_change(self, t):
        sel   = self.v[t]["smu"].get()
        no_op = sel in self._NO_OP
        for key in ("bias_row","list_row","start_row","stop_row","step_row","pts_row"):
            f = self._pf[t].get(key)
            if f:
                for w in f.winfo_children():
                    try: w.configure(state="disabled" if no_op else "normal")
                    except: pass
                f.configure(bg=_THEME["SEP"] if no_op else _THEME["PANEL"])
        if not no_op:
            self._on_op_change(t)

    def _on_op_change(self, t):
        if self.v[t]["smu"].get() in self._NO_OP: return
        op = self.v[t]["op"].get()
        is_list  = op in self._LIST
        is_bias  = op in self._BIAS
        is_sweep = op in self._SWEEP
        is_log   = op in self._LOG

        vis = {
            "bias_row":  is_bias,
            "list_row":  is_list,
            "start_row": is_sweep or is_log,
            "stop_row":  is_sweep or is_log,
            "step_row":  is_sweep,
            "pts_row":   is_sweep or is_log,
        }
        for key, show in vis.items():
            f = self._pf[t].get(key)
            if not f: continue
            bg = _THEME["PANEL"] if show else _THEME["SEP"]
            f.configure(bg=bg)
            for w in f.winfo_children():
                try: w.configure(state="normal" if show else "disabled")
                except: pass
            if key == "list_row":
                lbl = self._pf[t].get("list_lbl")
                if lbl: lbl.configure(bg=bg)

        if t == "Gate": self._recalc_pts()

    def _recalc_pts(self, *_):
        for t in self.TERMINALS:
            self._recalc_pts_for(t)

    def _recalc_pts_for(self, term, *_):
        try:
            if not self.v[term]["pts_auto"].get():
                return
            s  = float(self.v[term]["start"].get())
            e  = float(self.v[term]["stop"].get())
            st = float(self.v[term]["step"].get())
            if st == 0: return
            n = int(abs(e - s) / abs(st)) + 1
            self.v[term]["pts"].set(str(n))
        except:
            pass

    def _apply_nplc_all(self):
        val = self.nplc_all.get().strip()
        for t in self.TERMINALS:
            self.v[t]["nplc"].set(val)

    def _open_list_editor(self, terminal, lbl_widget):
        def on_save(vals):
            self._list_vals[terminal] = vals
            n = len(vals)
            lbl_widget.config(text=f"{n} value{'s' if n!=1 else ''}"
                              if n else "(empty)")
        ListEditor(self,
                   title=f"{terminal}  —  List Values",
                   initial=self._list_vals[terminal],
                   accent=self.TERM_COL[terminal],
                   on_save=on_save)

    def _build_graph(self, parent):
        info = tk.Frame(parent, bg=_THEME["PANEL"],
                        highlightbackground=_THEME["BORDER"], highlightthickness=1)
        info.pack(fill="x", pady=(0,4))
        self.info_lbl = tk.Label(info, text="Configure terminals and press RUN",
                                 bg=_THEME["PANEL"], fg=_THEME["TEXT3"], font=FONT_BODY, anchor="w")
        self.info_lbl.pack(side="left", padx=10, pady=5)

        self.curve_count = tk.Label(info, text="", bg=_THEME["PANEL"],
                                    fg=_THEME["ACCENT2"], font=FONT_MONO)
        self.curve_count.pack(side="right", padx=10)

        prog_f = tk.Frame(parent, bg=_THEME["PANEL"],
                          highlightbackground=_THEME["BORDER"], highlightthickness=1)
        prog_f.pack(fill="x", pady=(0,4))
        prow = tk.Frame(prog_f, bg=_THEME["PANEL"]); prow.pack(fill="x", padx=10, pady=5)
        self.progress = ProgressRow(prow)
        self.progress.pack(fill="x")

        tb = tk.Frame(parent, bg=_THEME["PANEL"],
                      highlightbackground=_THEME["BORDER"], highlightthickness=1)
        tb.pack(fill="x", pady=(0,4))
        ttk.Button(tb, text="Clear", command=self._clear_curves).pack(
            side="left", padx=6, pady=4)
        ttk.Button(tb, text="Log Y Scale",
                   command=self._toggle_log).pack(side="left", padx=(0,4))
        ttk.Button(tb, text="Auto Scale",
                   command=self._auto_scale).pack(side="left", padx=(0,4))

        tk.Button(tb, text="📊 Compare CSV",
                  bg=_THEME["PANEL2"], fg=_THEME["TEXT2"],
                  font=FONT_SM, relief="flat", bd=0,
                  activebackground=_THEME["HOVER"], cursor="hand2",
                  padx=8, pady=3,
                  command=lambda: self._compare.open_panel()).pack(
                  side="left", padx=(8,0), pady=3)
        self._log_y = False

        self._build_qpoint_panel(parent)
        self._build_cycling_stress_panel(parent)
        self._build_analysis_panel(parent)

        pf = tk.Frame(parent, bg=_THEME["PANEL"],
                      highlightbackground=_THEME["BORDER"], highlightthickness=1)
        pf.pack(fill="both", expand=True)

        self.fig = Figure(figsize=(7, 5.5), dpi=100, facecolor=_THEME["PANEL"])
        self.fig.subplots_adjust(left=0.11, right=0.95, top=0.93, bottom=0.10)
        self.ax  = self.fig.add_subplot(111)
        _style_plot(self.fig, self.ax, title="Configure terminals and press RUN")

        self.canvas = FigureCanvasTkAgg(self.fig, master=pf)
        self.canvas.draw()
        self.canvas.get_tk_widget().config(bg=_THEME["PANEL"])
        self.canvas.get_tk_widget().pack(fill="both", expand=True, padx=2, pady=2)

    def _build_qpoint_panel(self, parent):
        qf = tk.Frame(parent, bg=_THEME["PANEL"],
                      highlightbackground=_THEME["BORDER2"], highlightthickness=1)
        qf.pack(fill="x", pady=(0, 4))

        hdr = tk.Frame(qf, bg=ACCENT)
        hdr.pack(fill="x")
        tk.Label(hdr, text="  ⊙  Q-Point  Analysis",
                 bg=ACCENT, fg="white", font=FONT_H3).pack(side="left", pady=4)
        self.qp_enabled = tk.BooleanVar(value=False)
        tk.Checkbutton(hdr, text="Enable", variable=self.qp_enabled,
                       bg=ACCENT, fg="white", selectcolor=_THEME["ACCENT"],
                       activebackground=_THEME["ACCENT"], font=FONT_SM,
                       command=self._replot).pack(side="right", padx=10)

        bdy = tk.Frame(qf, bg=_THEME["PANEL"])
        bdy.pack(fill="x", padx=10, pady=6)

        self._qp_xfr_f = tk.Frame(bdy, bg=_THEME["PANEL"])
        tk.Label(self._qp_xfr_f, text="Vgs_Q", bg=_THEME["PANEL"], fg=_THEME["TEXT"],
                 font=FONT_BODY).pack(side="left")
        self.qp_vgs = tk.StringVar(value="4.0")
        ttk.Entry(self._qp_xfr_f, textvariable=self.qp_vgs,
                  width=7, font=FONT_BODY).pack(side="left", padx=(4, 2))
        tk.Label(self._qp_xfr_f, text="V", bg=_THEME["PANEL"], fg=_THEME["TEXT3"],
                 font=FONT_SM).pack(side="left", padx=(0, 16))

        self._qp_out_f = tk.Frame(bdy, bg=_THEME["PANEL"])

        tk.Label(self._qp_out_f, text="Vdd", bg=_THEME["PANEL"], fg=_THEME["TEXT"],
                 font=FONT_BODY).pack(side="left")
        self.qp_vdd = tk.StringVar(value="20.0")
        ttk.Entry(self._qp_out_f, textvariable=self.qp_vdd,
                  width=7, font=FONT_BODY).pack(side="left", padx=(4, 2))
        tk.Label(self._qp_out_f, text="V", bg=_THEME["PANEL"], fg=_THEME["TEXT3"],
                 font=FONT_SM).pack(side="left", padx=(0, 16))

        tk.Label(self._qp_out_f, text="RL", bg=_THEME["PANEL"], fg=_THEME["TEXT"],
                 font=FONT_BODY).pack(side="left")
        self.qp_rl = tk.StringVar(value="100.0")
        ttk.Entry(self._qp_out_f, textvariable=self.qp_rl,
                  width=7, font=FONT_BODY).pack(side="left", padx=(4, 2))
        tk.Label(self._qp_out_f, text="Ω", bg=_THEME["PANEL"], fg=_THEME["TEXT3"],
                 font=FONT_SM).pack(side="left", padx=(0, 16))

        tk.Label(self._qp_out_f, text="Vg curve", bg=_THEME["PANEL"], fg=_THEME["TEXT"],
                 font=FONT_BODY).pack(side="left")
        self.qp_curve_var = tk.StringVar(value="")
        self._qp_combo = ttk.Combobox(self._qp_out_f,
                                       textvariable=self.qp_curve_var,
                                       values=[], state="readonly",
                                       width=14, font=FONT_SM)
        self._qp_combo.pack(side="left", padx=(4, 16))

        ttk.Button(bdy, text="Apply", style="Accent.TButton",
                   command=self._replot).pack(side="left", padx=(0, 12))

        self.qp_result_lbl = tk.Label(bdy, text="—", bg=_THEME["PANEL"], fg=_THEME["ACCENT2"],
                                       font=FONT_MONO)
        self.qp_result_lbl.pack(side="left", fill="x")

        self._qp_sync_mode()

    def _qp_sync_mode(self):
        is_xfr = "Transfer" in self.mode_var.get()
        if is_xfr:
            self._qp_out_f.pack_forget()
            self._qp_xfr_f.pack(side="left")
        else:
            self._qp_xfr_f.pack_forget()
            self._qp_out_f.pack(side="left")

    def _build_cycling_stress_panel(self, parent):
        outer = tk.Frame(parent, bg=_THEME["PANEL"],
                         highlightbackground=_THEME["BORDER2"], highlightthickness=1)
        outer.pack(fill="x", pady=(0, 4))

        hdr = tk.Frame(outer, bg=_THEME["ACCENT"]); hdr.pack(fill="x")
        tk.Label(hdr, text="  ⟳  Cycling & Bias Stress",
                 bg=_THEME["ACCENT"], fg="white", font=FONT_H3).pack(side="left", pady=4)
        self._cyc_open = tk.BooleanVar(value=False)
        self._cyc_body = tk.Frame(outer, bg=_THEME["PANEL"])

        def _toggle():
            if self._cyc_open.get():
                self._cyc_body.pack_forget()
                self._cyc_open.set(False)
                arr_lbl.config(text="▸")
            else:
                self._cyc_body.pack(fill="x")
                self._cyc_open.set(True)
                arr_lbl.config(text="▾")

        arr_lbl = tk.Label(hdr, text="▸", bg=_THEME["ACCENT"], fg="white",
                           font=("Segoe UI", 10), cursor="hand2")
        arr_lbl.pack(side="right", padx=10)
        arr_lbl.bind("<Button-1>", lambda _: _toggle())
        hdr.bind("<Button-1>", lambda _: _toggle())

        bdy = self._cyc_body
        pad = dict(padx=10, pady=3)

        cyc_hdr = tk.Frame(bdy, bg=_THEME["ACCENT3"]); cyc_hdr.pack(fill="x", **pad)
        tk.Label(cyc_hdr, text=" Cycling", bg=_THEME["ACCENT3"], fg=_THEME["ACCENT2"],
                 font=FONT_H3).pack(side="left", pady=3)

        r1 = tk.Frame(bdy, bg=_THEME["PANEL"]); r1.pack(fill="x", padx=10, pady=2)
        tk.Checkbutton(r1, text="Enable Cycling", variable=self.cyc_enabled,
                       bg=_THEME["PANEL"], fg=_THEME["TEXT"], activebackground=_THEME["PANEL"],
                       selectcolor=_THEME["PANEL"], relief="flat",
                       highlightthickness=0, font=FONT_BODY).pack(side="left")
        tk.Label(r1, text="Cycles:", bg=_THEME["PANEL"], fg=_THEME["TEXT"], font=FONT_BODY).pack(side="left", padx=(12,2))
        ttk.Entry(r1, textvariable=self.cyc_n, width=5, font=FONT_BODY).pack(side="left")
        tk.Label(r1, text="Rest (s):", bg=_THEME["PANEL"], fg=_THEME["TEXT"], font=FONT_BODY).pack(side="left", padx=(12,2))
        ttk.Entry(r1, textvariable=self.cyc_rest, width=6, font=FONT_BODY).pack(side="left")

        r2 = tk.Frame(bdy, bg=_THEME["PANEL"]); r2.pack(fill="x", padx=10, pady=(0,4))
        for text, var in [("Show all cycles", self.cyc_show_all),
                          ("Gradient colors", self.cyc_gradient),
                          ("Mean±σ band", self.cyc_stats),
                          ("Ghost overlay", self.cyc_overlay)]:
            tk.Checkbutton(r2, text=text, variable=var, command=self._replot,
                           bg=_THEME["PANEL"], fg=_THEME["TEXT"], activebackground=_THEME["PANEL"],
                           selectcolor=_THEME["PANEL"], relief="flat",
                           highlightthickness=0, font=FONT_SM).pack(side="left", padx=(0,10))

        tk.Frame(bdy, bg=_THEME["SEP"], height=1).pack(fill="x", padx=10, pady=(4,0))
        cn_hdr = tk.Frame(bdy, bg=_THEME["PANEL"]); cn_hdr.pack(fill="x", padx=10, pady=(4,2))
        tk.Label(cn_hdr, text="Cycle Names — one per line (auto-used if filled):",
                 bg=_THEME["PANEL"], fg=_THEME["TEXT2"], font=FONT_SM).pack(side="left")

        self._fet_cyc_names_text = tk.Text(bdy, height=4, width=28,
                                            font=FONT_BODY,
                                            bg=_THEME["PANEL"], fg=_THEME["TEXT"],
                                            insertbackground=_THEME["ACCENT2"],
                                            relief="flat", bd=0,
                                            highlightthickness=1,
                                            highlightbackground=_THEME["BORDER"],
                                            highlightcolor=ACCENT2)
        self._fet_cyc_names_text.pack(fill="x", padx=10, pady=(0,2))
        tk.Label(bdy, text="  e.g.  Fresh,  Stressed,  Recovery…",
                 bg=_THEME["PANEL"], fg=_THEME["TEXT3"], font=FONT_SM, anchor="w").pack(
                 fill="x", padx=10, pady=(0,6))

        tk.Frame(bdy, bg=_THEME["SEP"], height=1).pack(fill="x", padx=10)
        str_hdr_col = _THEME["WARN2"] if _THEME["BG"] == DARK["BG"] else "#FEF3C7"
        str_hdr = tk.Frame(bdy, bg=str_hdr_col); str_hdr.pack(fill="x", **pad)
        tk.Label(str_hdr, text=" Bias Stress  (applied between cycles)",
                 bg=str_hdr_col, fg=_THEME["WARN"], font=FONT_H3).pack(side="left", pady=3)

        r3 = tk.Frame(bdy, bg=_THEME["PANEL"]); r3.pack(fill="x", padx=10, pady=2)
        tk.Checkbutton(r3, text="Enable Stress", variable=self.str_enabled,
                       bg=_THEME["PANEL"], fg=_THEME["TEXT"], activebackground=_THEME["PANEL"],
                       selectcolor=_THEME["PANEL"], relief="flat",
                       highlightthickness=0, font=FONT_BODY).pack(side="left")
        tk.Label(r3, text="Terminal:", bg=_THEME["PANEL"], fg=_THEME["TEXT"], font=FONT_BODY).pack(side="left", padx=(12,2))
        ttk.Combobox(r3, textvariable=self.str_term,
                     values=["Gate","Drain"], state="readonly",
                     width=7, font=FONT_BODY).pack(side="left")
        tk.Label(r3, text="Stress V:", bg=_THEME["PANEL"], fg=_THEME["TEXT"], font=FONT_BODY).pack(side="left", padx=(12,2))
        ttk.Entry(r3, textvariable=self.str_v, width=6, font=FONT_BODY).pack(side="left")
        tk.Label(r3, text="V", bg=_THEME["PANEL"], fg=_THEME["TEXT3"], font=FONT_SM).pack(side="left", padx=(1,12))
        tk.Label(r3, text="Duration:", bg=_THEME["PANEL"], fg=_THEME["TEXT"], font=FONT_BODY).pack(side="left", padx=(0,2))
        ttk.Entry(r3, textvariable=self.str_dur, width=6, font=FONT_BODY).pack(side="left")
        tk.Label(r3, text="s", bg=_THEME["PANEL"], fg=_THEME["TEXT3"], font=FONT_SM).pack(side="left", padx=(1,12))
        tk.Label(r3, text="Measure every:", bg=_THEME["PANEL"], fg=_THEME["TEXT"], font=FONT_BODY).pack(side="left", padx=(0,2))
        ttk.Entry(r3, textvariable=self.str_int, width=5, font=FONT_BODY).pack(side="left")
        tk.Label(r3, text="s", bg=_THEME["PANEL"], fg=_THEME["TEXT3"], font=FONT_SM).pack(side="left", padx=1)

        self.str_status_lbl = tk.Label(bdy, text="", bg=_THEME["PANEL"],
                                        fg=WARN, font=FONT_MONO)
        self.str_status_lbl.pack(anchor="w", padx=10, pady=(0,4))

    def _build_analysis_panel(self, parent):
        outer = tk.Frame(parent, bg=_THEME["PANEL"],
                         highlightbackground=_THEME["BORDER2"], highlightthickness=1)
        outer.pack(fill="x", pady=(0, 4))

        hdr = tk.Frame(outer, bg="#166534"); hdr.pack(fill="x")
        tk.Label(hdr, text="  ⚗  Parameter Extraction",
                 bg="#166534", fg="white", font=FONT_H3).pack(side="left", pady=4)
        self._an_open = tk.BooleanVar(value=False)
        self._an_body = tk.Frame(outer, bg=_THEME["PANEL"])

        def _toggle():
            if self._an_open.get():
                self._an_body.pack_forget()
                self._an_open.set(False)
                arr_lbl.config(text="▸")
            else:
                self._an_body.pack(fill="x")
                self._an_open.set(True)
                arr_lbl.config(text="▾")

        arr_lbl = tk.Label(hdr, text="▸", bg="#166534", fg="white",
                           font=("Segoe UI", 10), cursor="hand2")
        arr_lbl.pack(side="right", padx=10)
        arr_lbl.bind("<Button-1>", lambda _: _toggle())
        hdr.bind("<Button-1>", lambda _: _toggle())

        bdy = self._an_body

        r1 = tk.Frame(bdy, bg=_THEME["PANEL"]); r1.pack(fill="x", padx=10, pady=(6,2))
        for text, var in [("Vth", self.an_vth), ("SS (mV/dec)", self.an_ss),
                          ("Ion/Ioff", self.an_ionoff), ("Mobility", self.an_mob),
                          ("Hysteresis", self.an_hyst)]:
            tk.Checkbutton(r1, text=text, variable=var,
                           bg=_THEME["PANEL"], fg=_THEME["TEXT"], activebackground=_THEME["PANEL"],
                           selectcolor=_THEME["PANEL"], relief="flat",
                           highlightthickness=0, font=FONT_BODY).pack(side="left", padx=(0,10))

        self._mob_frame = tk.Frame(bdy, bg=_THEME["PANEL2"],
                                   highlightbackground=_THEME["BORDER"], highlightthickness=1)
        mob_row = tk.Frame(self._mob_frame, bg=_THEME["PANEL2"])
        mob_row.pack(fill="x", padx=8, pady=4)
        tk.Label(mob_row, text="Vds_lin:", bg=_THEME["PANEL2"], fg=_THEME["TEXT"], font=FONT_SM).pack(side="left")
        ttk.Entry(mob_row, textvariable=self.an_vds_lin, width=6, font=FONT_SM).pack(side="left", padx=(3,2))
        tk.Label(mob_row, text="V", bg=_THEME["PANEL2"], fg=_THEME["TEXT3"], font=FONT_SM).pack(side="left", padx=(0,12))
        tk.Label(mob_row, text="Cox (F/cm²):", bg=_THEME["PANEL2"], fg=_THEME["TEXT"], font=FONT_SM).pack(side="left")
        ttk.Entry(mob_row, textvariable=self.an_cox, width=10, font=FONT_SM).pack(side="left", padx=(3,12))
        tk.Label(mob_row, text="W/L (µm):", bg=_THEME["PANEL2"], fg=_THEME["TEXT"], font=FONT_SM).pack(side="left")
        ttk.Entry(mob_row, textvariable=self.an_W, width=7, font=FONT_SM).pack(side="left", padx=(3,1))
        tk.Label(mob_row, text="/", bg=_THEME["PANEL2"], fg=_THEME["TEXT3"], font=FONT_SM).pack(side="left")
        ttk.Entry(mob_row, textvariable=self.an_L, width=7, font=FONT_SM).pack(side="left", padx=(1,2))
        self.an_mob.trace_add("write", lambda *_: (
            self._mob_frame.pack(fill="x", padx=10, pady=2)
            if self.an_mob.get() else self._mob_frame.pack_forget()))

        act_row = tk.Frame(bdy, bg=_THEME["PANEL"]); act_row.pack(fill="x", padx=10, pady=(4,6))
        ttk.Button(act_row, text="Run Analysis", style="Accent.TButton",
                   command=self._run_analysis).pack(side="left", padx=(0,12))
        self.an_result_lbl = tk.Label(act_row, text="—", bg=_THEME["PANEL"], fg=_THEME["ACCENT2"],
                                       font=FONT_MONO)
        self.an_result_lbl.pack(side="left", fill="x")

        self.an_trend_lbl = tk.Label(bdy, text="", bg=_THEME["PANEL"],
                                      fg=_THEME["TEXT2"], font=FONT_SM)
        self.an_trend_lbl.pack(anchor="w", padx=10, pady=(0,4))

    def _extract_vth(self, x, y):
        try:
            x = np.array(x, dtype=float)
            y = np.abs(np.array(y, dtype=float))
            order = np.argsort(x); x, y = x[order], y[order]
            valid = y > 0
            if np.sum(valid) < 4: return float("nan")
            sq = np.sqrt(y)
            gm_sq = np.gradient(sq, x)
            idx = int(np.argmax(gm_sq))
            if gm_sq[idx] == 0: return float("nan")
            return float(x[idx] - sq[idx] / gm_sq[idx])
        except Exception:
            return float("nan")

    def _calc_ss(self, x, y):
        try:
            x = np.array(x, dtype=float)
            y = np.abs(np.array(y, dtype=float))
            order = np.argsort(x); x, y = x[order], y[order]
            valid = y > 1e-15
            if np.sum(valid) < 4: return float("nan")
            xv, yv = x[valid], y[valid]
            log_y = np.log10(yv)
            slope = np.gradient(log_y, xv)
            idx_max = int(np.argmax(slope))
            sub_slope = slope[:max(idx_max, 1)]
            if len(sub_slope) < 2: return float("nan")
            best = float(np.max(sub_slope))
            if best <= 0: return float("nan")
            return float(1000.0 / best)
        except Exception:
            return float("nan")

    def _calc_ion_ioff(self, y):
        try:
            y = np.abs(np.array(y, dtype=float))
            valid = y > 0
            if np.sum(valid) < 2: return float("nan")
            ion  = float(np.max(y[valid]))
            ioff = float(np.min(y[valid]))
            return float("nan") if ioff == 0 else ion / ioff
        except Exception:
            return float("nan")

    def _calc_mobility(self, x, y):
        try:
            cox = float(self.an_cox.get())
            W   = float(self.an_W.get()) * 1e-4
            L   = float(self.an_L.get()) * 1e-4
            vds = float(self.an_vds_lin.get())
            if cox <= 0 or W <= 0 or L <= 0 or vds == 0:
                return float("nan")
            x = np.array(x, dtype=float)
            y = np.array(y, dtype=float)
            order = np.argsort(x); x, y = x[order], y[order]
            gm = np.gradient(y, x)
            gm_max = float(np.max(np.abs(gm)))
            return gm_max * L / (W * cox * abs(vds))
        except Exception:
            return float("nan")

    def _calc_hysteresis(self, curve):
        try:
            x = np.array(curve["x"], dtype=float)
            y = np.abs(np.array(curve["y"], dtype=float))
            n = len(x) // 2
            if n < 3: return float("nan")
            x_fwd, y_fwd = x[:n], y[:n]
            x_rev_r = x[n:][::-1]; y_rev_r = y[n:][::-1]
            if len(x_rev_r) < 2: return float("nan")
            y_rev_i = np.interp(x_fwd, x_rev_r, y_rev_r)
            return float(np.trapz(np.abs(y_fwd - y_rev_i), x_fwd))
        except Exception:
            return float("nan")

    def _run_analysis(self):
        is_xfr = "Transfer" in self.mode_var.get()
        src = self._cycle_data[-1] if self._cycle_data else self.curves
        if not src:
            self.an_result_lbl.config(text="No data — run a measurement first.", fg=_THEME["DANGER"])
            return

        c = src[0]
        x = c["x"]; y = c["y"]
        eng = mticker.EngFormatter(sep="")
        parts = []

        if self.an_vth.get() and is_xfr:
            v = self._extract_vth(x, y)
            parts.append(f"Vth = {v:.3g} V" if not np.isnan(v) else "Vth = N/A")

        if self.an_ss.get() and is_xfr:
            v = self._calc_ss(x, y)
            parts.append(f"SS = {v:.0f} mV/dec" if not np.isnan(v) else "SS = N/A")

        if self.an_ionoff.get():
            v = self._calc_ion_ioff(y)
            parts.append(f"Ion/Ioff = {eng(v)}" if not np.isnan(v) else "Ion/Ioff = N/A")

        if self.an_mob.get() and is_xfr:
            v = self._calc_mobility(x, y)
            parts.append(f"µFE = {v:.3g} cm²/Vs" if not np.isnan(v) else "µFE = N/A (check Cox)")

        if self.an_hyst.get():
            v = self._calc_hysteresis(c)
            parts.append(f"Hyst = {eng(v)}A·V" if not np.isnan(v) else "Hyst = N/A (need dual sweep)")

        self.an_result_lbl.config(
            text="    ".join(parts) if parts else "No extractors enabled.",
            fg=_THEME["ACCENT2"])

        if len(self._cycle_data) > 1 and is_xfr:
            trend_parts = []
            vth_list, ion_list = [], []
            for cyc in self._cycle_data:
                if not cyc: continue
                vth_list.append(self._extract_vth(cyc[0]["x"], cyc[0]["y"]))
                ion_list.append(float(np.max(np.abs(cyc[0]["y"]))))
            vth_arr = [v for v in vth_list if not np.isnan(v)]
            if len(vth_arr) >= 2 and self.an_vth.get():
                dv = vth_arr[-1] - vth_arr[0]
                trend_parts.append(f"ΔVth = {dv:+.3g} V over {len(vth_arr)} cycles")
            if len(ion_list) >= 2:
                drift = (ion_list[-1] / ion_list[0] - 1) * 100 if ion_list[0] else 0
                trend_parts.append(f"Ion drift = {drift:+.1f}%")
            self.an_trend_lbl.config(
                text="Cycle trends:  " + "    ".join(trend_parts) if trend_parts else "")
        else:
            self.an_trend_lbl.config(text="")

    def _draw_qpoint(self):
        if not self.qp_enabled.get() or not self.curves:
            self.after(0, lambda: self.qp_result_lbl.config(text="—", fg=_THEME["ACCENT2"]))
            return
        is_xfr = "Transfer" in self.mode_var.get()
        try:
            if is_xfr:
                self._draw_qpoint_transfer()
            else:
                self._draw_qpoint_output()
        except Exception as ex:
            msg = str(ex)
            self.after(0, lambda m=msg: self.qp_result_lbl.config(
                text=f"Q-point: {m}", fg=_THEME["DANGER"]))

    def _draw_qpoint_transfer(self):
        vgs_q = float(self.qp_vgs.get())
        c = self.curves[0]
        x = np.array(c["x"],  dtype=float)
        y = np.array([abs(v) for v in c["y"]], dtype=float)
        if len(x) < 2: return
        order = np.argsort(x)
        x, y  = x[order], y[order]
        id_q = float(np.interp(vgs_q, x, y))
        dy   = np.gradient(y, x)
        gm_q = float(np.interp(vgs_q, x, dy))
        self.ax.axvline(vgs_q, color="#F59E0B", lw=1.2, ls="--", alpha=0.75, zorder=5)
        self.ax.axhline(id_q,  color="#F59E0B", lw=1.2, ls="--", alpha=0.75, zorder=5)
        self.ax.plot(vgs_q, id_q, "*", color="#F59E0B", markersize=15,
                     zorder=11, markeredgecolor="#92400E", markeredgewidth=0.8,
                     label=f"Q  Vgs = {vgs_q:.3g} V,  Id = {_eng_str(id_q)}A")
        ann = (f"  Vgs = {vgs_q:.3g} V\n"
               f"  Id  = {_eng_str(id_q)}A\n"
               f"  gm  = {_eng_str(gm_q)}S")
        self.ax.annotate(ann, xy=(vgs_q, id_q), xytext=(20, 18), textcoords="offset points",
            fontsize=8, color=WARN, fontfamily="Consolas",
            bbox=dict(boxstyle="round,pad=0.5", facecolor=WARN2, edgecolor="#F59E0B", alpha=0.95, linewidth=1.2),
            arrowprops=dict(arrowstyle="->", color="#F59E0B", connectionstyle="arc3,rad=0.1", lw=1.2))
        result = (f"Q:  Vgs = {vgs_q:.3g} V    Id = {_eng_str(id_q)}A    gm = {_eng_str(gm_q)}S")
        self.after(0, lambda r=result: self.qp_result_lbl.config(text=r, fg=_THEME["ACCENT2"]))

    def _draw_qpoint_output(self):
        vdd = float(self.qp_vdd.get())
        rl  = float(self.qp_rl.get())
        if rl <= 0: raise ValueError("RL must be > 0 Ω")
        if vdd <= 0: raise ValueError("Vdd must be > 0 V")
        eng = mticker.EngFormatter(sep="")
        id_max   = vdd / rl
        vds_line = np.array([0.0, vdd])
        id_line  = np.array([id_max, 0.0])
        self.ax.plot(vds_line, id_line, color="#F59E0B", lw=2.0, ls="-.",
                     zorder=8, label=f"Load Line  Vdd = {vdd:.3g} V,  RL = {eng(rl)}Ω")
        self.ax.annotate(f"{eng(vdd)}V", xy=(vdd, 0), xytext=(0, 8), textcoords="offset points",
                         fontsize=7.5, color="#92400E", ha="center")
        self.ax.annotate(f"{eng(id_max)}A", xy=(0, id_max), xytext=(8, 0), textcoords="offset points",
                         fontsize=7.5, color="#92400E", va="center")
        sel   = self.qp_curve_var.get()
        curve = next((c for c in self.curves if c["label"] == sel), self.curves[0])
        x = np.array(curve["x"],  dtype=float)
        y = np.array([abs(v) for v in curve["y"]], dtype=float)
        if len(x) < 2: return
        order = np.argsort(x); x, y  = x[order], y[order]
        id_ll = (vdd - x) / rl
        diff  = y - id_ll
        signs = np.sign(diff)
        xings = np.where(np.diff(signs))[0]
        if len(xings) == 0:
            result = "No Q-point intersection on selected curve"
            self.after(0, lambda r=result: self.qp_result_lbl.config(text=r, fg=_THEME["DANGER"]))
            return
        ci = xings[-1]
        x0, x1 = x[ci], x[ci + 1]
        d0, d1  = diff[ci], diff[ci + 1]
        vds_q = float(x0 - d0 * (x1 - x0) / (d1 - d0))
        id_q  = float(np.interp(vds_q, x, y))
        p_q   = vds_q * id_q
        vds_hyp = np.linspace(max(0.05 * vdd, 0.05), vdd * 0.98, 300)
        id_hyp  = p_q / vds_hyp
        self.ax.plot(vds_hyp, id_hyp, color="#DC2626", lw=1.2, ls=":", alpha=0.65, zorder=4,
                     label=f"Pd = {eng(p_q)}W  hyperbola")
        self.ax.axvline(vds_q, color="#F59E0B", lw=1.2, ls="--", alpha=0.75, zorder=5)
        self.ax.axhline(id_q,  color="#F59E0B", lw=1.2, ls="--", alpha=0.75, zorder=5)
        self.ax.plot(vds_q, id_q, "*", color="#F59E0B", markersize=15, zorder=11,
                     markeredgecolor="#92400E", markeredgewidth=0.8,
                     label=f"Q  Vds = {vds_q:.3g} V,  Id = {eng(id_q)}A")
        ann = (f"  {curve['label']}\n  Vds = {vds_q:.3g} V\n  Id  = {eng(id_q)}A\n  Pd  = {eng(p_q)}W")
        self.ax.annotate(ann, xy=(vds_q, id_q), xytext=(20, 18), textcoords="offset points",
            fontsize=8, color=WARN, fontfamily="Consolas",
            bbox=dict(boxstyle="round,pad=0.5", facecolor=WARN2, edgecolor="#F59E0B", alpha=0.95, linewidth=1.2),
            arrowprops=dict(arrowstyle="->", color="#F59E0B", connectionstyle="arc3,rad=-0.1", lw=1.2))
        result = (f"Q:  Vds = {vds_q:.3g} V    Id = {eng(id_q)}A    Pd = {eng(p_q)}W")
        self.after(0, lambda r=result: self.qp_result_lbl.config(text=r, fg=_THEME["ACCENT2"]))

    def _replot(self):
        self.fig.clear()
        self.fig.patch.set_facecolor(_THEME["PANEL"])
        self.ax = self.fig.add_subplot(111)
        self.fig.subplots_adjust(left=0.11, right=0.95, top=0.93, bottom=0.10)
        mode   = self.mode_var.get()
        is_xfr = "Transfer" in mode
        xlabel = "Gate Voltage  Vg (V)" if is_xfr else "Drain Voltage  Vds (V)"
        ylabel = "|Drain Current  Id| (A)"
        title  = ("Transfer Curve  —  Id vs Vg" if is_xfr else "Output Curve  —  Id vs Vds")
        _style_plot(self.fig, self.ax, title=title)
        self.ax.set_xlabel(xlabel, color=_THEME["TEXT"], fontsize=9)
        self.ax.set_ylabel(ylabel, color=_THEME["TEXT"], fontsize=9)
        _eng_formatter(self.ax, "x")
        if not self._log_y:
            _eng_formatter(self.ax, "y")

        if hasattr(self, "_qp_combo") and self.curves:
            labels = [c["label"] for c in self.curves]
            self._qp_combo["values"] = labels
            if self.qp_curve_var.get() not in labels:
                self.qp_curve_var.set(labels[0])

        if self.cyc_overlay.get() and self._overlay_data:
            ghost_col = "#888888" if not self.app._dark_mode else "#4A5568"
            for c in self._overlay_data:
                x = c.get("x", []); y_raw = [abs(v) for v in c.get("y", [])]
                if not x: continue
                n = min(len(x), len(y_raw))
                if n == 0: continue
                yp = ([v if v > 1e-15 else float("nan") for v in y_raw[:n]]
                      if self._log_y else y_raw[:n])
                self.ax.plot(x[:n], yp, "-", lw=1.4, color=ghost_col, alpha=0.50, zorder=1)

        all_cycle_list = list(self._cycle_data)
        if self.curves:
            all_cycle_list.append(list(self.curves))

        use_cycles = self.cyc_enabled.get() and len(all_cycle_list) >= 1

        if use_cycles:
            n_cyc = len(all_cycle_list)
            cyc_colors = [FET_PAL[i % len(FET_PAL)] for i in range(n_cyc)]

            for ci, (cyc_curves, col) in enumerate(zip(all_cycle_list, cyc_colors)):
                is_current = (ci == len(all_cycle_list) - 1 and self.is_running)
                alpha = 1.0 if (is_current or n_cyc == 1) else max(0.45, 1.0 - 0.07 * (n_cyc - ci - 1))
                lw    = 1.8 if (is_current or n_cyc == 1) else (1.0 if n_cyc > 4 else 1.6)
                ms    = 2.5 if (is_current or n_cyc == 1) else (1.5 if n_cyc > 4 else 2.5)
                sfx_live = " [live]" if is_current else ""

                for c in cyc_curves:
                    x = c.get("x", []); y_raw = [abs(v) for v in c.get("y", [])]
                    y = ([v if v > 1e-15 else float("nan") for v in y_raw]
                         if self._log_y else y_raw)
                    n = min(len(x), len(y))
                    if n == 0: continue
                    lbl = (f"{self._cycle_names_fet[ci] if ci < len(self._cycle_names_fet) else f'Cyc {ci+1}'}  "
                           f"{c['label']}{sfx_live}") if n_cyc > 1 else f"{c['label']}{sfx_live}"
                    self.ax.plot(x[:n], y[:n], "o-", markersize=ms,
                                 linewidth=lw, color=col, label=lbl,
                                 alpha=alpha, zorder=3 + ci)

            completed_cycs = all_cycle_list[:-1] if self.is_running else all_cycle_list
            if (self.cyc_show_all.get() and self.cyc_stats.get()
                    and len(completed_cycs) >= 2):
                try:
                    ref_x = np.array(completed_cycs[0][0]["x"])
                    if len(ref_x) >= 3:
                        all_y = []
                        for cyc_curves in completed_cycs:
                            if not cyc_curves: continue
                            c  = cyc_curves[0]
                            xc = np.array(c["x"]); yc = np.abs(np.array(c["y"]))
                            order = np.argsort(xc)
                            yi = np.interp(ref_x, xc[order], yc[order])
                            all_y.append(yi)
                        if len(all_y) >= 2:
                            ay    = np.array(all_y)
                            mu    = np.nanmean(ay, axis=0)
                            sg    = np.nanstd(ay, axis=0)
                            valid = ~np.isnan(mu)
                            band_col = "#22D3EE" if self.app._dark_mode else "#3B82F6"
                            self.ax.fill_between(
                                ref_x[valid], (mu - sg)[valid], (mu + sg)[valid],
                                alpha=0.20 if self.app._dark_mode else 0.15,
                                color=band_col, zorder=2,
                                label=f"μ±σ ({len(completed_cycs)} cycles)")
                            self.ax.plot(ref_x[valid], mu[valid], "--",
                                         lw=1.8, color=band_col, alpha=0.90, zorder=4, label="Mean")
                except Exception:
                    pass
        else:
            src = (self.curves if self.curves
                   else (self._cycle_data[-1] if self._cycle_data else []))
            for i, c in enumerate(src):
                col = FET_PAL[i % len(FET_PAL)]
                x = c.get("x", []); y_raw = [abs(v) for v in c.get("y", [])]
                y = ([v if v > 1e-15 else float("nan") for v in y_raw]
                     if self._log_y else y_raw)
                n = min(len(x), len(y))
                if n == 0: continue
                self.ax.plot(x[:n], y[:n], "o-", markersize=2.5,
                             linewidth=1.8, color=col, label=c["label"])

        if self._log_y:
            try:
                self.ax.set_yscale("log")
                import matplotlib.ticker as _mticker
                self.ax.yaxis.set_major_formatter(
                    _mticker.LogFormatterSciNotation(labelOnlyBase=False))
                self.ax.yaxis.set_minor_formatter(
                    _mticker.LogFormatterSciNotation(labelOnlyBase=False,
                                                     minor_thresholds=(2, 0.4)))
            except Exception:
                pass

        self._draw_qpoint()

        if self._compare.refs:
            self._compare.draw(self.ax)

        handles, auto_labels = self.ax.get_legend_handles_labels()
        if handles:
            self.ax.legend(handles, auto_labels, fontsize=7.5, loc="best",
                           framealpha=0.85, edgecolor=_THEME["BORDER"],
                           facecolor=_THEME["PANEL"], labelcolor=_THEME["TEXT"],
                           labelspacing=0.3)
        self.canvas.draw_idle()
        n_cyc = len(self._cycle_data)
        n_cur = len(self.curves)
        if n_cyc > 0 and self.cyc_enabled.get():
            self.curve_count.config(
                text=f"Cycle {self._current_cycle + 1}/{n_cyc}  ·  {n_cur} curve{'s' if n_cur!=1 else ''}")
        else:
            self.curve_count.config(
                text=f"{n_cur} curve{'s' if n_cur!=1 else ''}" if n_cur else "")

    def _mode_changed(self):
        mode = self.mode_var.get()
        is_xfr = "Transfer" in mode
        if is_xfr:
            self.v["Gate"]["op"].set("Voltage Linear Sweep")
            self.v["Drain"]["op"].set("Voltage Bias")
        else:
            self.v["Gate"]["op"].set("Voltage Bias")
            self.v["Drain"]["op"].set("Voltage Linear Sweep")
        for t in self.TERMINALS:
            self._on_op_change(t)
        if hasattr(self, "_qp_xfr_f"):
            self._qp_sync_mode()
        self._replot()

    def _clear_curves(self):
        if self.is_running: return
        self.curves           = []
        self._cycle_data      = []
        self._cycle_names_fet = []
        self._overlay_data    = None
        self._current_cycle   = 0
        self.fig.clear()
        self.fig.patch.set_facecolor(_THEME["PANEL"])
        self.ax = self.fig.add_subplot(111)
        self.fig.subplots_adjust(left=0.11, right=0.95, top=0.93, bottom=0.10)
        _style_plot(self.fig, self.ax, title="Configure terminals and press RUN")
        self.canvas.draw()
        self.curve_count.config(text="")
        self.progress.reset()
        if hasattr(self, "an_result_lbl"):
            self.an_result_lbl.config(text="—", fg=_THEME["ACCENT2"])
        if hasattr(self, "an_trend_lbl"):
            self.an_trend_lbl.config(text="")

    def _auto_scale(self):
        try:
            self.ax.relim(); self.ax.autoscale_view()
            self.canvas.draw_idle()
        except: pass

    def _toggle_log(self):
        self._log_y = not self._log_y
        self._replot()

    def _fv(self, t, k, fb=0.0):
        try: return float(self.v[t][k].get())
        except: return fb

    def _get_fet_cycle_name(self, cycle_idx: int) -> str:
        try:
            raw   = self._fet_cyc_names_text.get("1.0", tk.END)
            lines = [l.strip() for l in raw.splitlines() if l.strip()]
            if cycle_idx < len(lines):
                return lines[cycle_idx]
        except Exception:
            pass
        return f"Cyc {cycle_idx + 1}"

    def _step_list(self, start, stop, step):
        if step == 0: return [start]
        if (stop-start)*step < 0: step = -step
        vals, v = [], start
        while ((step>0 and v<=stop+1e-9) or (step<0 and v>=stop-1e-9)):
            vals.append(round(v,9)); v += step
        return vals

    def _resolve(self, t):
        sel = self.v[t]["smu"].get()
        return self._TSP.get(sel, None), sel not in self._NO_OP

    def _set_status(self, text, col="#4ADE80"):
        self.after(0, lambda: self.status_lbl.config(text=f"●  {text}", fg=col))

    def _run(self):
        if self.is_running: return
        is_xfr = "Transfer" in self.mode_var.get()
        pterm  = "Gate" if is_xfr else "Drain"
        _, real = self._resolve(pterm)
        if not real:
            messagebox.showerror("Configuration Error",
                f"'{pterm}' must be assigned to SMU1 or SMU2.")
            return
        self.is_running     = True
        self._stop          = False
        if self.cyc_overlay.get():
            src = []
            if self._cycle_data:
                for cyc in self._cycle_data:
                    src.extend(cyc)
            elif self.curves:
                src = self.curves
            if src:
                self._overlay_data = copy.deepcopy(src)
        self.curves           = []
        self._cycle_data      = []
        self._cycle_names_fet = []
        self._current_cycle   = 0
        self.fig.clear()
        self.fig.patch.set_facecolor(_THEME["PANEL"])
        self.ax = self.fig.add_subplot(111)
        self.fig.subplots_adjust(left=0.11, right=0.95, top=0.93, bottom=0.10)
        _style_plot(self.fig, self.ax, title="Acquiring…")
        self.canvas.draw()
        self._set_status("Acquiring…", "#F59E0B")
        self.progress.start()
        threading.Thread(target=self._measure, daemon=True).start()

    def _do_stop(self):
        self._stop = True
        self._set_status("Stopping…", "#EF4444")
        self.after(0, lambda: self.progress.stop(False))
        try:
            if self.inst:
                self.inst.write(TSP_SAFE_ABORT)
                for t in self.TERMINALS:
                    tsp, real = self._resolve(t)
                    if real: self.inst.write(f"{tsp}.source.output = {tsp}.OFF")
        except: pass

    def _safe_abort(self):
        try: self.inst.write(TSP_SAFE_ABORT); time.sleep(0.15)
        except: pass

    def _measure(self):
        sweep_tsp = bias_tsp = None
        try:
            self.inst = self.app.rm.open_resource(self.addr_var.get().strip())
            self.inst.timeout = 60000
            self.after(0, lambda: self._conn_lbl.config(
                text=f"● SMU: {self.addr_var.get().strip().split('::')[-2] if '::' in self.addr_var.get() else self.addr_var.get()[:12]}",
                fg="#4ADE80") if hasattr(self, "_conn_lbl") else None)

            mode   = self.mode_var.get()
            is_xfr = "Transfer" in mode

            gate_tsp  = self._TSP.get(self.v["Gate"]["smu"].get())
            drain_tsp = self._TSP.get(self.v["Drain"]["smu"].get())
            if not gate_tsp:
                raise ValueError("Gate must be assigned to SMU1 or SMU2.")
            if not drain_tsp:
                raise ValueError("Drain must be assigned to SMU1 or SMU2.")

            if is_xfr:
                sweep_tsp, sweep_term = gate_tsp,  "Gate"
                bias_tsp,  bias_term  = drain_tsp, "Drain"
            else:
                sweep_tsp, sweep_term = drain_tsp, "Drain"
                bias_tsp,  bias_term  = gate_tsp,  "Gate"

            read_tsp = drain_tsp

            sw_start = self._fv(sweep_term, "start",   0.0)
            sw_stop  = self._fv(sweep_term, "stop",    8.0)
            sw_step  = self._fv(sweep_term, "step",    0.1)
            sw_comp  = self._fv(sweep_term, "comp",    0.01)
            sw_nplc  = self._fv(sweep_term, "nplc",    1.0)
            sw_delay = self._fv(sweep_term, "delay",   0.05)
            sw_dual  = self.v[sweep_term]["dual"].get()

            bi_op    = self.v[bias_term]["op"].get()
            bi_comp  = self._fv(bias_term, "comp",  0.5)
            bi_bias  = self._fv(bias_term, "bias",  10.0)
            bi_start = self._fv(bias_term, "start", 10.0)
            bi_stop  = self._fv(bias_term, "stop",   0.0)
            bi_step  = self._fv(bias_term, "step",  -1.0)

            if bi_op in self._BIAS:
                bias_levels = [bi_bias]
            elif bi_op in self._SWEEP or bi_op in self._LOG:
                bias_levels = self._step_list(bi_start, bi_stop, bi_step)
            elif bi_op in self._LIST:
                lv = self._list_vals.get(bias_term, [])
                bias_levels = lv if lv else [bi_bias]
            else:
                bias_levels = [bi_bias]

            sw_op   = self.v[sweep_term]["op"].get()
            sw_auto = self.v[sweep_term]["pts_auto"].get()
            if sw_op in self._LIST:
                lv = self._list_vals.get(sweep_term, [])
                if not lv:
                    raise ValueError(f"No list values for {sweep_term}.")
                sweep_base = list(lv)
            elif sw_auto:
                if abs(sw_step) < 1e-12:
                    raise ValueError("Sweep step cannot be zero.")
                sweep_base = self._step_list(sw_start, sw_stop, sw_step)
            else:
                try:
                    n_manual = max(2, int(self.v[sweep_term]["pts"].get()))
                except ValueError:
                    n_manual = 81
                sweep_base = list(np.linspace(sw_start, sw_stop, n_manual))

            sweep_vals = (list(sweep_base) + list(reversed(sweep_base))
                          if sw_dual else list(sweep_base))
            n_pts = len(sweep_vals)
            if n_pts == 0:
                raise ValueError("Sweep produced 0 points.")

            sense_c  = ("SENSE_4WIRE" if "4-Wire" in self.sense_var.get() else "SENSE_2WIRE")
            term_str = self.terminals_var.get().upper()

            try: mq_avg = max(1, int(self.mq_avg.get()))
            except ValueError: mq_avg = 1
            mq_settle  = self.mq_settle.get()
            try: mq_tol = max(0.001, float(self.mq_tol.get())) / 100.0
            except ValueError: mq_tol = 0.01
            try: mq_maxatt = max(1, int(self.mq_maxatt.get()))
            except ValueError: mq_maxatt = 5
            mq_precheck = self.mq_precheck.get()

            cyc_enabled = self.cyc_enabled.get()
            try: n_cycles = max(1, int(self.cyc_n.get())) if cyc_enabled else 1
            except ValueError: n_cycles = 1
            try: cyc_rest = max(0.0, float(self.cyc_rest.get()))
            except ValueError: cyc_rest = 0.0

            str_enabled = self.str_enabled.get() and cyc_enabled and n_cycles > 1
            try:
                str_v   = float(self.str_v.get())
                str_dur = max(1.0, float(self.str_dur.get()))
                str_int = max(0.5, float(self.str_int.get()))
            except ValueError:
                str_v = 20.0; str_dur = 60.0; str_int = 10.0
            str_tsp = (sweep_tsp if self.str_term.get() == sweep_term else bias_tsp)

            def W(cmd): self.inst.write(cmd)
            def Qf(cmd, default=float("nan")):
                try:
                    r = self.inst.query(cmd).strip()
                    if r in ("nil", "", "nil\n"): return default
                    return float(r)
                except Exception: return default
            def Q(cmd, default="nan"):
                try:
                    r = self.inst.query(cmd).strip()
                    return default if r in ("nil", "", "nil\n") else r
                except Exception: return default

            def read_point(tsp):
                if mq_avg == 1 and not mq_settle:
                    return Qf(f"print({tsp}.measure.read())")
                readings = []
                for _ in range(mq_avg):
                    readings.append(Qf(f"print({tsp}.measure.read())"))
                val = float(np.nanmean(readings))
                if mq_settle:
                    for _ in range(mq_maxatt - 1):
                        prev = val
                        new_readings = [Qf(f"print({tsp}.measure.read())") for _ in range(mq_avg)]
                        val = float(np.nanmean(new_readings))
                        if abs(prev) > 1e-18:
                            if abs(val - prev) / abs(prev) < mq_tol: break
                        else:
                            if abs(val - prev) < 1e-15: break
                return val

            needs_node2 = "node[2]" in (sweep_tsp + bias_tsp)
            self._set_status("Resetting…", "#F59E0B")
            W("*CLS")

            if needs_node2:
                W("tsplink.reset(1)")
                time.sleep(2.0)
                state = Q("print(tsplink.state)", "offline")
                if "online" not in state.lower():
                    raise RuntimeError(
                        f"TSP-Link not online (state={state!r}).\n\n"
                        "• Connect TSP-Link cable between both 2450s\n"
                        "• Set Node ID=1 on Master, Node ID=2 on Slave\n"
                        "  MENU → System → TSP-Link Node")
                W("node[1].reset()"); time.sleep(0.5)
                W("node[2].reset()"); time.sleep(0.5)
            else:
                W("reset()"); time.sleep(0.6)

            self._set_status("Configuring…", "#F59E0B")

            def cfg_smu(tsp, comp, nplc, sense, term):
                W(f"{tsp}.terminals = {tsp}.TERMINALS_{term}")
                W(f"{tsp}.source.func = {tsp}.FUNC_DC_VOLTAGE")
                W(f"{tsp}.source.ilimit.level = {comp}")
                W(f"{tsp}.source.autorange = {tsp}.ON")
                W(f"{tsp}.source.level = 0")
                W(f"{tsp}.measure.func = {tsp}.FUNC_DC_CURRENT")
                W(f"{tsp}.measure.sense = {tsp}.{sense}")
                W(f"{tsp}.measure.nplc = {nplc}")
                W(f"{tsp}.measure.autorange = {tsp}.ON")
                W(f"{tsp}.measure.autozero.enable = {tsp}.OFF")
                W(f"{tsp}.measure.autozero.once()")
                time.sleep(0.2)
                W(f"{tsp}.source.output = {tsp}.ON")
                time.sleep(0.4)

            cfg_smu(sweep_tsp, sw_comp, sw_nplc, sense_c, term_str)
            cfg_smu(bias_tsp,  bi_comp, sw_nplc, sense_c, term_str)
            time.sleep(0.3)

            if mq_precheck:
                self._set_status("Pre-check…", "#F59E0B")
                W(f"{sweep_tsp}.source.level = {sweep_vals[0]:.6g}")
                W(f"{bias_tsp}.source.level = {bias_levels[0]:.6g}")
                time.sleep(max(sw_delay, 0.1))
                i_test = Qf(f"print({read_tsp}.measure.read())")
                if abs(i_test) >= sw_comp * 0.95:
                    W(f"{sweep_tsp}.source.level = 0")
                    W(f"{bias_tsp}.source.level = 0")
                    raise RuntimeError(
                        f"Compliance pre-check failed:\n"
                        f"Id = {i_test:.4g} A at start point — "
                        f"device may be shorted or compliance too low.\n\n"
                        f"Raise Compliance or reduce start voltage and retry.")
                W(f"{sweep_tsp}.source.level = 0")
                W(f"{bias_tsp}.source.level = 0")
                time.sleep(0.1)

            total_pts_all = n_cycles * len(bias_levels) * n_pts

            for cycle_idx in range(n_cycles):
                if self._stop: break

                self._current_cycle = cycle_idx
                if n_cycles > 1:
                    self._set_status(f"Cycle {cycle_idx+1}/{n_cycles}  — Measuring…", "#F59E0B")

                if str_enabled and cycle_idx > 0 and not self._stop:
                    self._set_status(f"Cycle {cycle_idx}/{n_cycles}  → Stress {str_dur:.0f}s…", WARN)
                    W(f"{str_tsp}.source.level = {str_v:.6g}")
                    t_str_start = time.time()
                    next_meas   = t_str_start + str_int
                    while time.time() - t_str_start < str_dur and not self._stop:
                        now = time.time()
                        if now >= next_meas:
                            i_s = Qf(f"print({read_tsp}.measure.read())")
                            elapsed = now - t_str_start
                            msg = (f"Stress  t={elapsed:.0f}s  Id={mticker.EngFormatter(sep='')(i_s)}A")
                            self.after(0, lambda m=msg: self.str_status_lbl.config(text=m))
                            next_meas += str_int
                        time.sleep(0.2)
                    W(f"{str_tsp}.source.level = 0")
                    time.sleep(0.15)
                    self.after(0, lambda: self.str_status_lbl.config(text=""))

                if cyc_rest > 0 and cycle_idx > 0 and not self._stop:
                    self._set_status(f"Cycle {cycle_idx}/{n_cycles}  → Rest {cyc_rest:.1f}s…", "#60A5FA")
                    time.sleep(cyc_rest)

                cycle_curves = []
                total_curves = len(bias_levels)

                for ci, bv in enumerate(bias_levels):
                    if self._stop: break

                    bias_lbl = (f"Vds = {bv:.4g} V" if is_xfr else f"Vg = {bv:.4g} V")
                    self._set_status(
                        f"Cyc {cycle_idx+1}/{n_cycles}  Curve {ci+1}/{total_curves}  —  {bias_lbl}", "#F59E0B")

                    W(f"{bias_tsp}.source.level = {bv:.6g}")
                    time.sleep(sw_delay + 0.15)

                    curve_dict = {
                        "label": bias_lbl + ("  [dual]" if sw_dual else ""),
                        "cycle": cycle_idx, "step": bv,
                        "x": [], "y": []
                    }
                    cycle_curves.append(curve_dict)
                    self.curves = cycle_curves
                    all_x, all_y = [], []
                    interval = max(1, n_pts // 20)

                    for si, sv in enumerate(sweep_vals):
                        if self._stop: break
                        W(f"{sweep_tsp}.source.level = {sv:.6g}")
                        time.sleep(max(sw_delay, 0.05))
                        i_val = read_point(read_tsp)
                        all_x.append(sv)
                        all_y.append(i_val)

                        if si % interval == 0 or si == n_pts - 1:
                            curve_dict["x"] = list(all_x)
                            curve_dict["y"] = list(all_y)
                            self.after(0, self._replot)

                        done = (cycle_idx * len(bias_levels) * n_pts + ci * n_pts + si + 1)
                        pct = min(int(100 * done / total_pts_all), 99)
                        self.after(0, lambda pv=pct: self.progress.update(pv))

                    curve_dict["x"] = all_x
                    curve_dict["y"] = all_y
                    self.after(0, self._replot)

                self._cycle_data.append(cycle_curves)
                cyc_name_fet = self._get_fet_cycle_name(cycle_idx)
                self._cycle_names_fet.append(cyc_name_fet)
                self.curves = []
                self.after(0, self._replot)

            try:
                W(f"{sweep_tsp}.source.level = 0")
                W(f"{bias_tsp}.source.level = 0")
                time.sleep(0.15)
                W(f"{sweep_tsp}.source.output = {sweep_tsp}.OFF")
                W(f"{bias_tsp}.source.output = {bias_tsp}.OFF")
                time.sleep(0.2)
                self.inst.close()
            except Exception:
                pass

            if self._stop:
                self._set_status("Stopped", "#EF4444")
                self.after(0, lambda: self.progress.stop(False))
            else:
                n_c = len(self._cycle_data)
                n_cur = sum(len(c) for c in self._cycle_data)
                self._set_status(
                    f"✓  Done  —  {n_c} cycle{'s' if n_c!=1 else ''}  ·  {n_cur} curve{'s' if n_cur!=1 else ''}  acquired",
                    "#22C55E")
                self.after(0, lambda: self.progress.stop(True))
                self.after(0, self._auto_save)
                if any([self.an_vth.get(), self.an_ss.get(),
                        self.an_ionoff.get(), self.an_mob.get()]):
                    self.after(200, self._run_analysis)

        except Exception as e:
            err = str(e)
            self.after(0, lambda em=err: messagebox.showerror("FET Error", em))
            self._set_status("Error", "#EF4444")
            self.after(0, lambda: self.progress.stop(False))
            try:
                if sweep_tsp:
                    self.inst.write(f"{sweep_tsp}.source.output = {sweep_tsp}.OFF")
                if bias_tsp:
                    self.inst.write(f"{bias_tsp}.source.output = {bias_tsp}.OFF")
            except Exception:
                pass
        finally:
            self.is_running = False
            self._stop      = False
            self.after(0, lambda: self._conn_lbl.config(
                text="● SMU: Disconnected", fg="#EF4444")
                if hasattr(self, "_conn_lbl") else None)

    def _make_fn(self):
        tag = "Transfer" if "Transfer" in self.mode_var.get() else "Output"
        return f"FET_{tag}_{time.strftime('%Y%m%d_%H%M%S')}"

    def _write_fet_csv(self, path):
        mode   = self.mode_var.get()
        is_xfr = "Transfer" in mode
        # FIX: self.curves is cleared after each cycle — fall back to last cycle data
        src = self.curves if self.curves else (
              self._cycle_data[-1] if self._cycle_data else [])
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["# Keithley 2450  Lab View  —  FET Characterisation"])
            w.writerow(["# Mode", mode])
            for t in self.TERMINALS:
                w.writerow([f"# {t}", self.v[t]["smu"].get(), self.v[t]["op"].get()])
            w.writerow(["# Date", time.strftime("%d/%m/%Y  %H:%M:%S")])
            w.writerow([])
            hdr = ["Vg (V)" if is_xfr else "Vds (V)"]
            for c in src: hdr.append(f"Id  {c['label']} (A)")
            w.writerow(hdr)
            if not src: return
            max_pts = max(len(c["x"]) for c in src)
            for i in range(max_pts):
                row = [f"{src[0]['x'][i]:.6g}" if i < len(src[0]["x"]) else ""]
                for c in src:
                    row.append(f"{c['y'][i]:.9g}" if i < len(c["y"]) else "")
                w.writerow(row)

    def _export_csv(self):
        # FIX: also check _cycle_data
        if not self.curves and not self._cycle_data:
            messagebox.showwarning("No Data","Run a measurement first."); return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv", initialdir=self.app.save_dir,
            initialfile=f"{self._make_fn()}.csv",
            filetypes=[("CSV","*.csv"),("All","*.*")])
        if not path: return
        try:
            self._write_fet_csv(path)
            messagebox.showinfo("Saved", f"CSV saved:\n{path}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def _export_cycle_csv(self):
        if not self._cycle_data:
            messagebox.showwarning("No Cycle Data",
                "No cycle data recorded.\nEnable cycling and run a measurement first.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv", initialdir=self.app.save_dir,
            initialfile=f"{self._make_fn()}_cycles.csv",
            filetypes=[("CSV","*.csv"),("All","*.*")])
        if not path: return
        try:
            mode   = self.mode_var.get()
            is_xfr = "Transfer" in mode
            x_col  = "Vg (V)" if is_xfr else "Vds (V)"
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["# Keithley 2450  Lab View  —  Cycle Export"])
                w.writerow(["# Mode", mode])
                w.writerow(["# Cycles", len(self._cycle_data)])
                w.writerow(["# Date", time.strftime("%d/%m/%Y  %H:%M:%S")])
                if hasattr(self, "an_result_lbl"):
                    w.writerow(["# Extracted", self.an_result_lbl.cget("text")])
                w.writerow([])
                header = [x_col]
                for ci, cyc in enumerate(self._cycle_data):
                    cyc_name = (self._cycle_names_fet[ci]
                                if ci < len(self._cycle_names_fet) else f"Cyc {ci+1}")
                    for c in cyc:
                        header.append(f"{cyc_name}  {c['label']} (A)")
                w.writerow(header)
                ref_x = self._cycle_data[0][0]["x"] if self._cycle_data else []
                max_pts = max(
                    (max(len(c["x"]) for c in cyc) for cyc in self._cycle_data if cyc), default=0)
                for i in range(max_pts):
                    row = [f"{ref_x[i]:.6g}" if i < len(ref_x) else ""]
                    for cyc in self._cycle_data:
                        for c in cyc:
                            row.append(f"{c['y'][i]:.9g}" if i < len(c["y"]) else "")
                    w.writerow(row)
            messagebox.showinfo("Saved", f"Cycle CSV saved:\n{path}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def _save_png(self):
        # FIX: also check _cycle_data
        if not self.curves and not self._cycle_data:
            messagebox.showwarning("No Data","Run a measurement first."); return
        path = filedialog.asksaveasfilename(
            defaultextension=".png", initialdir=self.app.save_dir,
            initialfile=f"{self._make_fn()}.png",
            filetypes=[("PNG","*.png"),("All","*.*")])
        if path:
            try:
                self.app._save_fig_light(self.fig, path)
                self._set_status("PNG saved")
            except Exception as e:
                messagebox.showerror("Save Error", str(e))

    def _auto_save(self):
        try:
            base = self._make_fn()
            self._write_fet_csv(os.path.join(self.app.save_dir, f"{base}.csv"))
            if self._cycle_data and len(self._cycle_data) > 1:
                with open(os.path.join(self.app.save_dir, f"{base}_cycles.csv"),
                          "w", newline="", encoding="utf-8-sig") as f:
                    w = csv.writer(f)
                    mode   = self.mode_var.get()
                    is_xfr = "Transfer" in mode
                    x_col  = "Vg (V)" if is_xfr else "Vds (V)"
                    w.writerow(["# Keithley 2450  Lab View  — Cycle Export"])
                    w.writerow(["# Mode", mode])
                    w.writerow(["# Cycles", len(self._cycle_data)])
                    w.writerow(["# Date", time.strftime("%d/%m/%Y  %H:%M:%S")])
                    w.writerow([])
                    header = [x_col]
                    for ci, cyc in enumerate(self._cycle_data):
                        cyc_name = (self._cycle_names_fet[ci]
                                    if ci < len(self._cycle_names_fet) else f"Cyc {ci+1}")
                        for c in cyc:
                            header.append(f"{cyc_name}  {c['label']} (A)")
                    w.writerow(header)
                    ref_x  = self._cycle_data[0][0]["x"] if self._cycle_data else []
                    max_pts = max(
                        (max(len(c["x"]) for c in cyc) for cyc in self._cycle_data if cyc), default=0)
                    for i in range(max_pts):
                        row = [f"{ref_x[i]:.6g}" if i < len(ref_x) else ""]
                        for cyc in self._cycle_data:
                            for c in cyc:
                                row.append(f"{c['y'][i]:.9g}" if i < len(c["y"]) else "")
                        w.writerow(row)
            self.app._save_fig_light(self.fig, os.path.join(self.app.save_dir, f"{base}.png"))
            self._set_status(f"Auto-saved  →  {self.app.save_dir}")
        except Exception:
            pass


# ═══════════════════════════════════════════════════════════════════════════════
#  LOADING SCREEN
# ═══════════════════════════════════════════════════════════════════════════════

class SplashScreen(tk.Toplevel):
    _STEPS = [
        ("Initialising VISA driver…",      0.14),
        ("Loading instrument profiles…",   0.28),
        ("Building measurement engine…",   0.44),
        ("Configuring plot framework…",    0.60),
        ("Setting up FET analysis…",       0.74),
        ("Preparing UI components…",       0.88),
        ("Ready.",                          1.00),
    ]

    def __init__(self, root):
        super().__init__(root)
        self.overrideredirect(True)
        w, h = 520, 300
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self.configure(bg="#0D1117")
        self.attributes("-topmost", True)
        self.lift()

        tk.Frame(self, bg="#1F6FEB", height=4).pack(fill="x")

        body = tk.Frame(self, bg="#0D1117")
        body.pack(fill="both", expand=True, padx=40, pady=(28, 0))

        tk.Label(body, text="KEITHLEY  2450",
                 bg="#0D1117", fg="#E6EDF3",
                 font=("Segoe UI", 20, "bold")).pack(anchor="w")
        tk.Label(body, text="Source Measure Unit  ·  Lab View",
                 bg="#0D1117", fg="#58A6FF",
                 font=("Segoe UI", 9)).pack(anchor="w", pady=(2, 0))
        tk.Label(body, text="by Rangaraajan Muralidaran",
                 bg="#0D1117", fg="#484F58",
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(2, 0))

        tk.Frame(body, bg="#21262D", height=1).pack(fill="x", pady=(20, 16))

        self._step_lbl = tk.Label(body, text="Starting up…",
                                   bg="#0D1117", fg="#8B949E",
                                   font=("Segoe UI", 9))
        self._step_lbl.pack(anchor="w")

        track = tk.Frame(body, bg="#21262D", height=6)
        track.pack(fill="x", pady=(10, 0))
        track.pack_propagate(False)
        self._bar = tk.Frame(track, bg="#1F6FEB", height=6)
        self._bar.place(x=0, y=0, relheight=1.0, relwidth=0.0)

        foot = tk.Frame(self, bg="#161B22")
        foot.pack(fill="x", side="bottom")
        tk.Label(foot, text="Developed with  ♥  for precision measurement",
                 bg="#161B22", fg="#30363D",
                 font=("Segoe UI", 7)).pack(pady=6)

        self._step_idx = 0
        self._animate()

    def _animate(self):
        if self._step_idx >= len(self._STEPS):
            return
        label, frac = self._STEPS[self._step_idx]
        self._step_lbl.config(text=label)
        self._bar.place_configure(relwidth=frac)
        self._step_idx += 1
        delay = 180 if self._step_idx < len(self._STEPS) else 350
        self.after(delay, self._animate)

    def finish(self):
        self.destroy()


# ═══════════════════════════════════════════════════════════════════════════════
#  MODE SELECTOR
# ═══════════════════════════════════════════════════════════════════════════════

class ModeSelector(tk.Toplevel):
    MODES = [
        ("IV",  "⚡",  "I-V Characterisation",
         "Voltage/current sweeps, cycling,\ncurve comparison & parameter extraction"),
        ("FET", "🔬",  "FET Characterisation",
         "Transfer & output curves,\nbias stress, Q-point analysis"),
        ("DMM", "📊",  "Digital Multimeter",
         "Live voltage & current readout\nfor probe landing & contact checking"),
    ]

    def __init__(self, root, callback, on_quit):
        super().__init__(root)
        self._callback = callback
        self._on_quit  = on_quit
        self.title("Keithley 2450 - Lab View")
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        self.geometry(f"{sw}x{sh}+0+0")
        self.configure(bg="#0D1117")
        self.state("zoomed")
        try: self.attributes("-zoomed", True)
        except Exception: pass
        self.lift()
        self._build()
        self.protocol("WM_DELETE_WINDOW", self._quit)
        self.bind("<Escape>", lambda _: self._quit())
        self.focus_force()

    def _build(self):
        top = tk.Frame(self, bg="#161B22"); top.pack(fill="x")
        tk.Frame(top, bg="#1F6FEB", height=4).pack(fill="x")
        quit_row = tk.Frame(top, bg="#161B22"); quit_row.pack(fill="x", padx=16, pady=6)
        tk.Label(quit_row, text="KEITHLEY 2450  —  Lab View",
                 bg="#161B22", fg="#484F58",
                 font=("Segoe UI", 9)).pack(side="left")
        tk.Button(quit_row, text="✕  Quit",
                  bg="#161B22", fg="#F87171",
                  font=("Segoe UI", 9), relief="flat", bd=0,
                  activebackground="#21262D", cursor="hand2",
                  padx=8, pady=2,
                  command=self._quit).pack(side="right")

        hdr = tk.Frame(self, bg="#0D1117")
        hdr.pack(fill="x", pady=(30, 0))
        tk.Label(hdr, text="KEITHLEY  2450  —  Lab View",
                 bg="#0D1117", fg="#E6EDF3",
                 font=("Segoe UI", 22, "bold")).pack()
        tk.Label(hdr, text="Select measurement mode to continue",
                 bg="#0D1117", fg="#58A6FF",
                 font=("Segoe UI", 11)).pack(pady=(6, 0))

        tk.Frame(self, bg="#21262D", height=1).pack(fill="x", pady=30)

        cards_f = tk.Frame(self, bg="#0D1117")
        cards_f.pack(expand=True)

        for mode_key, icon, title, desc in self.MODES:
            card = tk.Frame(cards_f, bg="#161B22",
                            highlightbackground="#30363D", highlightthickness=2,
                            cursor="hand2", padx=40, pady=30)
            card.pack(side="left", padx=24, ipadx=10, ipady=10)

            tk.Label(card, text=icon, bg="#161B22", fg="#E6EDF3",
                     font=("Segoe UI", 40)).pack()
            tk.Label(card, text=title, bg="#161B22", fg="#E6EDF3",
                     font=("Segoe UI", 15, "bold")).pack(pady=(8, 4))
            tk.Label(card, text=desc, bg="#161B22", fg="#8B949E",
                     font=("Segoe UI", 9), justify="center").pack()

            btn = tk.Button(card, text=f"Open  {mode_key}",
                            bg="#1F6FEB", fg="white",
                            font=("Segoe UI", 10, "bold"),
                            relief="flat", bd=0, padx=24, pady=8,
                            activebackground="#388BFD", cursor="hand2",
                            command=lambda k=mode_key: self._pick(k))
            btn.pack(pady=(18, 0))

            def _enter(e, f=card):
                f.config(highlightbackground="#58A6FF")
            def _leave(e, f=card):
                f.config(highlightbackground="#30363D")
            card.bind("<Enter>", _enter)
            card.bind("<Leave>", _leave)
            for child in card.winfo_children():
                child.bind("<Enter>", _enter)
                child.bind("<Leave>", _leave)

        tk.Frame(self, bg="#21262D", height=1).pack(fill="x", pady=30)
        tk.Label(self, text="by Rangaraajan Muralidaran  ·  Developed with ♥  ·  Press Esc to quit",
                 bg="#0D1117", fg="#30363D",
                 font=("Segoe UI", 8)).pack(pady=(0, 16))

    def _pick(self, mode_key):
        self.destroy()
        self._callback(mode_key)

    def _quit(self):
        self.destroy()
        self._on_quit()


# ═══════════════════════════════════════════════════════════════════════════════
#  DMM WINDOW
# ═══════════════════════════════════════════════════════════════════════════════

class DMMWindow(tk.Toplevel):
    _POLL_MS=400; _CONT_THRESH=10.0
    MEAS_MODES=[
        ("DCV","DC Voltage",   "smu.FUNC_DC_VOLTAGE","smu.FUNC_DC_VOLTAGE","Source V","V",True),
        ("DCI","DC Current",   "smu.FUNC_DC_CURRENT","smu.FUNC_DC_CURRENT","Source I","A",True),
        ("R2W","Resistance 2W","smu.FUNC_DC_VOLTAGE","smu.FUNC_DC_CURRENT","Test V",  "V",True),
        ("R4W","Resistance 4W","smu.FUNC_DC_VOLTAGE","smu.FUNC_DC_CURRENT","Test V",  "V",False),
        ("DIODE","Diode Test", "smu.FUNC_DC_CURRENT","smu.FUNC_DC_VOLTAGE","Test I",  "A",True),
        ("CONT","Continuity",  "smu.FUNC_DC_VOLTAGE","smu.FUNC_DC_CURRENT","Test V",  "V",True),
        ("PWR","Power (VxI)",  "smu.FUNC_DC_VOLTAGE","smu.FUNC_DC_CURRENT","Source V","V",True),
    ]
    _SRC={"DCV":"0","DCI":"0","R2W":"1","R4W":"1","DIODE":"0.001","CONT":"0.5","PWR":"1"}
    _COL={"DCV":("#58A6FF","#1F6FEB"),"DCI":("#3FB950","#2EA043"),
          "R2W":("#E3B341","#B08800"),"R4W":("#E3B341","#B08800"),
          "DIODE":("#F778BA","#C050A0"),"CONT":("#22D3EE","#0E9DBB"),
          "PWR":("#F97316","#C05000")}

    def __init__(self,root):
        super().__init__(root)
        self.title("Keithley 2450  —  Digital Multimeter")
        self.configure(bg=_THEME["BG"])
        self.state("zoomed")
        try: self.attributes("-zoomed",True)
        except: pass
        self.rm=pyvisa.ResourceManager(); self.inst=None
        self._running=False; self._after_id=None
        self._hold=False; self._min=None; self._max=None
        self.addr_var  =tk.StringVar(value=_cfg.get("visa_address"))
        self.mode_var  =tk.StringVar(value="DCV")
        self.src_var   =tk.StringVar(value="0")
        self.sense_var =tk.StringVar(value="2-Wire")
        self.panel_var =tk.StringVar(value="Rear")
        self.rate_var  =tk.StringVar(value="Medium (1 PLC)")
        self.status_var=tk.StringVar(value="● SMU: Disconnected")
        self.v_val =tk.StringVar(value="—"); self.i_val=tk.StringVar(value="—")
        self.r_val =tk.StringVar(value="—"); self.p_val=tk.StringVar(value="—")
        self.main_val=tk.StringVar(value="—"); self.sub_val=tk.StringVar(value="")
        self.minmax_var=tk.StringVar(value="")
        setup_styles()
        self._build()

    def _build(self):
        for w in self.winfo_children(): w.destroy()
        self.configure(bg=_THEME["BG"])
        BG =_THEME["BG"]; PNL=_THEME["PANEL"]; P2 =_THEME["PANEL2"]
        BD =_THEME["BORDER"]; TXT=_THEME["TEXT"]; T2 =_THEME["TEXT2"]
        T3 =_THEME["TEXT3"]; A2 =_THEME["ACCENT2"]; A3=_THEME["ACCENT3"]
        SUC=_THEME["SUCCESS"]; DNG=_THEME["DANGER"]; HOV=_THEME["HOVER"]
        tk.Frame(self,bg=A2,height=4).pack(fill="x")
        hdr=tk.Frame(self,bg=P2); hdr.pack(fill="x")
        tk.Label(hdr,text="  📊  Digital Multimeter",bg=P2,fg=TXT,
                 font=("Segoe UI",13,"bold")).pack(side="left",padx=16,pady=8)
        is_dark=_THEME["BG"]==DARK["BG"]
        dm_t="🌙  Dark" if is_dark else "☀  Light"
        dm_b="#1A2535" if is_dark else "#243F6A"; dm_f="#FCD34D" if is_dark else "#93C5FD"
        self._dark_btn=tk.Button(hdr,text=dm_t,font=("Segoe UI",8),
                                  bg=dm_b,fg=dm_f,relief="flat",bd=0,
                                  activebackground=P2,cursor="hand2",
                                  padx=8,pady=2,command=self._toggle_dark_dmm)
        self._dark_btn.pack(side="right",padx=(0,8),pady=8)
        self._conn_lbl=tk.Label(hdr,textvariable=self.status_var,
                                 bg=P2,fg=DNG,font=FONT_MONO)
        self._conn_lbl.pack(side="right",padx=(8,14),pady=8)
        mbar=tk.Frame(self,bg=BG); mbar.pack(fill="x",padx=8,pady=(6,0))
        self._mode_btns={}
        for key,name,*_ in self.MEAS_MODES:
            b=tk.Button(mbar,text=name,bg=PNL,fg=T2,font=("Segoe UI",9),
                         relief="flat",bd=0,padx=12,pady=6,cursor="hand2",
                         activebackground=P2,command=lambda k=key:self._set_mode(k))
            b.pack(side="left",padx=2); self._mode_btns[key]=b
        sb=tk.Frame(self,bg=P2,highlightbackground=BD,highlightthickness=1)
        sb.pack(fill="x",padx=8,pady=6)
        def L(t): tk.Label(sb,text=t,bg=P2,fg=T2,font=("Segoe UI",9)).pack(side="left",padx=(6,2))
        def E(var,w=8):
            e=tk.Entry(sb,textvariable=var,width=w,bg=BG,fg=TXT,insertbackground=A2,
                       relief="flat",font=("Consolas",8),highlightthickness=1,highlightbackground=BD)
            e.pack(side="left",padx=(0,8),ipady=3); return e
        def K(var,vals,w=8):
            c=ttk.Combobox(sb,textvariable=var,values=vals,state="readonly",width=w,font=("Segoe UI",8))
            c.pack(side="left",padx=(0,8)); return c
        L("VISA:"); E(self.addr_var,34)
        self._src_lbl=tk.Label(sb,text="Source V (V):",bg=P2,fg=T2,font=("Segoe UI",9))
        self._src_lbl.pack(side="left",padx=(0,2)); E(self.src_var,8)
        L("Sense:"); self._sense_cb=K(self.sense_var,["2-Wire","4-Wire"],7)
        L("Rate:");  K(self.rate_var,["Fast (0.1 PLC)","Medium (1 PLC)","Slow (10 PLC)"],16)
        L("Terminals:"); K(self.panel_var,["Rear","Front"],6)
        tk.Button(sb,text="ℹ  Wiring Guide",bg=A3,fg=A2,font=("Segoe UI",9),
                  relief="flat",bd=0,padx=10,pady=4,cursor="hand2",
                  activebackground=HOV,command=self._show_wiring_guide).pack(side="left",padx=(0,8))
        self._connect_btn=tk.Button(sb,text="►  Connect",bg=SUC,fg="white",
                                     font=("Segoe UI",9,"bold"),relief="flat",bd=0,
                                     padx=12,pady=4,cursor="hand2",activebackground=SUC,
                                     command=self._toggle_connect)
        self._connect_btn.pack(side="left",pady=4)
        if self._running: self._connect_btn.config(text="■  Disconnect",bg=DNG)
        self._disp_frame=tk.Frame(self,bg=BG)
        self._disp_frame.pack(fill="both",expand=True,padx=8,pady=(4,4))
        self._build_display()
        ctrl=tk.Frame(self,bg=P2,highlightbackground=BD,highlightthickness=1)
        ctrl.pack(fill="x",padx=8,pady=(0,4))
        for lt,cmd in [("HOLD",self._toggle_hold),("MIN/MAX",self._toggle_minmax),("RESET",self._reset_minmax)]:
            tk.Button(ctrl,text=lt,bg=PNL,fg=T2,font=("Segoe UI",8),
                      relief="flat",bd=0,padx=10,pady=4,cursor="hand2",
                      activebackground=P2,command=cmd).pack(side="left",padx=4,pady=6)
        tk.Label(ctrl,textvariable=self.minmax_var,bg=P2,
                 fg=self._COL["PWR"][0],font=("Consolas",9)).pack(side="left",padx=12)
        foot=tk.Frame(self,bg=BG); foot.pack(fill="x",padx=8,pady=(0,4))
        tk.Label(foot,text="NPLC: Fast=0.1 / Med=1 / Slow=10  |  R=V/I  |  P=VxI",
                 bg=BG,fg=T3,font=("Segoe UI",7)).pack(side="left")
        tk.Button(foot,text="✕  Close",bg=BG,fg=DNG,font=("Segoe UI",9),
                  relief="flat",bd=0,cursor="hand2",command=self._close).pack(side="right")
        self.protocol("WM_DELETE_WINDOW",self._close)
        self._set_mode(self.mode_var.get())

    def _build_display(self):
        for w in self._disp_frame.winfo_children(): w.destroy()
        BG=_THEME["BG"]; PNL=_THEME["PANEL"]; BD=_THEME["BORDER"]
        T2=_THEME["TEXT2"]; T3=_THEME["TEXT3"]
        key=self.mode_var.get()
        col1=self._COL.get(key,("#58A6FF","#1F6FEB"))[0]
        mn=next((n for k,n,*_ in self.MEAS_MODES if k==key),key)
        left=tk.Frame(self._disp_frame,bg=PNL,highlightbackground=BD,highlightthickness=1)
        left.pack(side="left",fill="both",expand=True,padx=(0,6))
        tk.Label(left,text=mn.upper(),bg=PNL,fg=T2,font=("Segoe UI",12)).pack(pady=(20,0))
        tk.Label(left,textvariable=self.main_val,bg=PNL,fg=col1,font=("Consolas",72,"bold")).pack(expand=True)
        tk.Label(left,textvariable=self.sub_val, bg=PNL,fg=T3, font=("Consolas",16)).pack(pady=(0,20))
        right=tk.Frame(self._disp_frame,bg=BG); right.pack(side="left",fill="y")
        for sym,var,col in [("V",self.v_val,self._COL["DCV"][0]),
                             ("I",self.i_val,self._COL["DCI"][0]),
                             ("R",self.r_val,self._COL["R2W"][0]),
                             ("P",self.p_val,self._COL["PWR"][0])]:
            tile=tk.Frame(right,bg=PNL,highlightbackground=BD,highlightthickness=1)
            tile.pack(fill="x",padx=6,pady=3,ipadx=8,ipady=4)
            tk.Label(tile,text=sym,bg=PNL,fg=T3,font=("Segoe UI",9)).pack(side="left",padx=(6,4))
            tk.Label(tile,textvariable=var,bg=PNL,fg=col,font=("Consolas",16,"bold")).pack(side="left")

    def _set_mode(self,key):
        self.mode_var.set(key)
        row=next((r for r in self.MEAS_MODES if r[0]==key),None)
        if row:
            self._src_lbl.config(text=f"{row[4]} ({row[5]}):")
            self.src_var.set(self._SRC.get(key,"1"))
            if not row[6]: self.sense_var.set("4-Wire"); self._sense_cb.config(state="disabled")
            else:
                if self.sense_var.get()=="4-Wire" and key!="R4W": self.sense_var.set("2-Wire")
                self._sense_cb.config(state="readonly")
        for k,b in self._mode_btns.items():
            col=self._COL[k][0]
            (b.config(bg=col,fg="#0D1117") if k==key
             else b.config(bg=_THEME["PANEL"],fg=_THEME["TEXT2"]))
        self._reset_minmax(); self.main_val.set("—"); self.sub_val.set("")
        self._build_display()
        if self._running and self.inst: self._configure_inst()

    def _nplc(self):
        r=self.rate_var.get()
        return "0.1" if "Fast" in r else ("10" if "Slow" in r else "1")

    def _configure_inst(self):
        key=self.mode_var.get(); row=next(r for r in self.MEAS_MODES if r[0]==key)
        sense="smu.SENSE_4WIRE" if self.sense_var.get()=="4-Wire" else "smu.SENSE_2WIRE"
        panel=self.panel_var.get().upper()
        try: src=float(self.src_var.get())
        except: src=0.0
        try:
            self.inst.write(f"smu.terminals = smu.TERMINALS_{panel}")
            self.inst.write(f"smu.source.func = {row[2]}")
            self.inst.write(f"smu.source.level = {src}")
            self.inst.write(f"smu.measure.func = {row[3]}")
            self.inst.write(f"smu.measure.sense = {sense}")
            self.inst.write(f"smu.measure.nplc = {self._nplc()}")
            self.inst.write("smu.measure.autorange = smu.ON")
            if key=="CONT":  self.inst.write("smu.source.ilimit.level = 0.1")
            if key=="DIODE": self.inst.write("smu.source.vlimit.level = 5")
            self.inst.write("smu.source.output = smu.ON")
        except Exception as e: self.status_var.set(f"Config err: {e}")

    def _toggle_connect(self):
        if self._running: self._disconnect()
        else:             self._connect()

    def _connect(self):
        addr=self.addr_var.get().strip()
        if not addr: self.status_var.set("● SMU: No address"); return
        try:
            self.inst=self.rm.open_resource(addr); self.inst.timeout=8000
            self.inst.write("*RST"); time.sleep(0.3)
            self._running=True
            self._connect_btn.config(text="■  Disconnect",bg=_THEME["DANGER"])
            self.status_var.set("● SMU: Connected")
            if hasattr(self,"_conn_lbl"): self._conn_lbl.config(fg=_THEME["SUCCESS"])
            _cfg.set("visa_address",addr)
            self._configure_inst(); self._poll()
        except Exception as e:
            self.status_var.set(f"● SMU: Error")
            if hasattr(self,"_conn_lbl"): self._conn_lbl.config(fg=_THEME["DANGER"])
            if self.inst:
                try: self.inst.close()
                except: pass
            self.inst=None

    def _disconnect(self):
        self._running=False
        if self._after_id:
            try: self.after_cancel(self._after_id)
            except: pass
        if self.inst:
            try: self.inst.write("smu.source.output = smu.OFF"); self.inst.close()
            except: pass
        self.inst=None
        self._connect_btn.config(text="►  Connect",bg=_THEME["SUCCESS"])
        self.status_var.set("● SMU: Disconnected")
        if hasattr(self,"_conn_lbl"): self._conn_lbl.config(fg=_THEME["DANGER"])
        for v in (self.v_val,self.i_val,self.r_val,self.p_val,self.main_val,self.sub_val): v.set("—")

    def _poll(self):
        if not self._running: return
        key=self.mode_var.get(); fmt=mticker.EngFormatter(sep="")
        try:
            rm=self.inst.query("print(smu.measure.read(defbuffer1))").strip()
            rs=self.inst.query("print(defbuffer1.sourcevalues[defbuffer1.n])").strip()
            meas=float(rm); src=float(rs)
            if   key=="DCV": v,i=meas,src
            elif key=="DCI": i,v=meas,src
            elif key=="DIODE": i,v=src,meas
            else: v,i=src,meas
            r=abs(v/i) if abs(i)>1e-15 else float("inf")
            p=abs(v*i)
            self.v_val.set(fmt(v)); self.i_val.set(fmt(i))
            self.r_val.set("∞" if r==float("inf") else fmt(r)); self.p_val.set(fmt(p))
            if   key=="DCV":   main=fmt(v); sub=f"{fmt(i)}A  |  {fmt(r)}Ω"
            elif key=="DCI":   main=fmt(i); sub=f"{fmt(v)}V  |  {fmt(r)}Ω"
            elif key in("R2W","R4W"): main="∞" if r==float("inf") else fmt(r); sub=f"{fmt(v)}V  |  {fmt(i)}A"
            elif key=="DIODE": main=fmt(v); sub=f"Vf @ {fmt(i)}A"
            elif key=="CONT":
                ok=r<self._CONT_THRESH; main="CLOSED" if ok else "OPEN"; sub=f"{fmt(r)}Ω"
                try:
                    lf=self._disp_frame.winfo_children()[0]
                    lf.config(highlightbackground=_THEME["SUCCESS"] if ok else _THEME["DANGER"])
                except: pass
            elif key=="PWR":   main=fmt(p); sub=f"{fmt(v)}V x {fmt(i)}A"
            else:              main=fmt(meas); sub=""
            if not self._hold:
                self.main_val.set(main); self.sub_val.set(sub)
                try:
                    num=float(rm)
                    if self._min is None or num<self._min: self._min=num
                    if self._max is None or num>self._max: self._max=num
                    self.minmax_var.set(f"MIN {fmt(self._min)}   MAX {fmt(self._max)}")
                except: pass
            self.status_var.set("● SMU: Connected")
            if hasattr(self,"_conn_lbl"): self._conn_lbl.config(fg=_THEME["SUCCESS"])
        except Exception as e: self.status_var.set("● SMU: Read error")
        if self._running: self._after_id=self.after(self._POLL_MS,self._poll)

    def _toggle_hold(self):
        self._hold=not self._hold; self.sub_val.set("▌ HOLD" if self._hold else "")

    def _toggle_minmax(self):
        if self._min is not None:
            fmt=mticker.EngFormatter(sep="")
            self.minmax_var.set(f"MIN {fmt(self._min)}   MAX {fmt(self._max)}")

    def _reset_minmax(self): self._min=None; self._max=None; self.minmax_var.set("")

    def _close(self): self._disconnect(); self.destroy()

    def _toggle_dark_dmm(self):
        is_dark=_THEME["BG"]==DARK["BG"]
        _THEME.update(LIGHT if is_dark else DARK)
        _sync_globals(); _cfg.set("dark_mode",not is_dark); setup_styles()
        self._build()

    def _show_wiring_guide(self):
        GUIDES=[
            ("DC Voltage","#58A6FF","Measure voltage across a component.",
             ["HI -> Force High (FHIGH)","LO -> Force Low (FLOW)","Leave SENSE unconnected"],
             " FHIGH --[DUT]-- FLOW","Rear or Front. 2-Wire only."),
            ("DC Current","#3FB950","Measure current. Insert SMU IN SERIES.",
             ["FHIGH -> Force High (current exits)","FLOW -> Force Low (current returns)","Set compliance V to protect circuit"],
             " FHIGH -> [DUT] -> FLOW  (series)","Compliance voltage limits V across DUT."),
            ("Resistance 2W","#E3B341","Source V, measure I, R=V/I. Includes lead resistance.",
             ["FHIGH -> one end of resistor","FLOW -> other end"],
             " FHIGH --[R_dut]-- FLOW","Lead resistance (~0.1-1 ohm) included. Use 4W for R < 100 ohm."),
            ("Resistance 4W","#E3B341","Kelvin. Eliminates lead resistance. Required for low-R.",
             ["FHIGH -> Force+ (current source)","FLOW -> Force- (current return)",
              "SHIGH -> Sense+ near DUT (no current)","SLOW -> Sense- near DUT (no current)"],
             " FHIGH --+--[R_dut]--+-- FLOW\n        |              |\n      SHIGH         SLOW",
             "4 separate probe wires. Required for R < 10 ohm or precision."),
            ("Diode Test","#F778BA","Force constant current, measure forward voltage Vf.",
             ["FHIGH -> Anode (+) of diode","FLOW -> Cathode (-) of diode","Typical test current: 1 mA"],
             " FHIGH --[>|]-- FLOW\n   (+) Anode->Cathode (-)","Si: 0.6-0.7V  Schottky: 0.2-0.4V  LED: 1.8-3.5V"),
            ("Continuity","#22D3EE","Check for low-resistance path. CLOSED/OPEN.",
             ["FHIGH -> probe tip 1","FLOW -> probe tip 2","Touch both to conductor"],
             " FHIGH -- probe -- track -- probe -- FLOW",
             "< 10 ohm = CLOSED (green)  |  >= 10 ohm = OPEN (red)"),
            ("Power (VxI)","#F97316","Source V, measure I, P=VxI.",
             ["FHIGH -> + terminal of load","FLOW -> - terminal"],
             " FHIGH --[Load]-- FLOW\n  V*I = P","Positive P = power delivered to DUT."),
        ]
        win=tk.Toplevel(self); win.title("DMM Wiring Guide")
        win.configure(bg=_THEME["BG"]); win.geometry("660x560"); win.resizable(True,True)
        tk.Frame(win,bg=_THEME["ACCENT2"],height=3).pack(fill="x")
        tk.Label(win,text="  DMM Wiring & Connection Guide",
                 bg=_THEME["BG"],fg=_THEME["TEXT"],font=("Segoe UI",12,"bold")).pack(anchor="w",padx=14,pady=(10,2))
        tk.Label(win,text="FHIGH=Force High | FLOW=Force Low | SHIGH=Sense High | SLOW=Sense Low",
                 bg=_THEME["BG"],fg=_THEME["TEXT2"],font=("Segoe UI",8)).pack(anchor="w",padx=14,pady=(0,6))
        tk.Frame(win,bg=_THEME["SEP"],height=1).pack(fill="x",padx=14,pady=(0,6))
        cont=tk.Frame(win,bg=_THEME["BG"]); cont.pack(fill="both",expand=True)
        cvs=tk.Canvas(cont,bg=_THEME["BG"],highlightthickness=0)
        vsb=ttk.Scrollbar(cont,orient="vertical",command=cvs.yview)
        inner=tk.Frame(cvs,bg=_THEME["BG"])
        inner.bind("<Configure>",lambda e:cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.create_window((0,0),window=inner,anchor="nw")
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left",fill="both",expand=True); vsb.pack(side="right",fill="y")
        cvs.bind("<MouseWheel>",lambda e:cvs.yview_scroll(int(-e.delta/120),"units"))
        for title,col,desc,conns,diag,notes in GUIDES:
            card=tk.Frame(inner,bg=_THEME["PANEL"],highlightbackground=_THEME["BORDER"],highlightthickness=1)
            card.pack(fill="x",padx=12,pady=4)
            tbar=tk.Frame(card,bg=col); tbar.pack(fill="x")
            tk.Label(tbar,text=f"  {title}",bg=col,fg="#0D1117",font=("Segoe UI",10,"bold")).pack(side="left",pady=4)
            body=tk.Frame(card,bg=_THEME["PANEL"]); body.pack(fill="x",padx=10,pady=6)
            tk.Label(body,text=desc,bg=_THEME["PANEL"],fg=_THEME["TEXT"],font=("Segoe UI",9),anchor="w").pack(anchor="w")
            tk.Label(body,text="Connections:",bg=_THEME["PANEL"],fg=_THEME["TEXT2"],font=("Segoe UI",8,"bold")).pack(anchor="w",pady=(4,1))
            for ln in conns: tk.Label(body,text=f"  - {ln}",bg=_THEME["PANEL"],fg=_THEME["TEXT"],font=("Consolas",8),anchor="w").pack(anchor="w")
            tk.Label(body,text=diag,bg=_THEME["BG"],fg=col,font=("Consolas",9),anchor="w",justify="left",padx=8,pady=4).pack(fill="x",pady=(6,2))
            tk.Label(body,text=notes,bg=_THEME["PANEL"],fg=_THEME["TEXT2"],font=("Segoe UI",8),anchor="w",justify="left").pack(anchor="w",pady=(2,0))
        tk.Button(win,text="Close",bg=_THEME["PANEL2"],fg=_THEME["TEXT"],font=("Segoe UI",9),
                  relief="flat",bd=0,padx=14,pady=6,cursor="hand2",
                  activebackground=_THEME["HOVER"],command=win.destroy).pack(pady=8)


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT  —  BUG FIX: _back() destroys ALL root children (not just Toplevels)
#  so IV windows (which pack frames directly on root) are fully cleaned up
#  before the mode selector is shown again.
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()

    splash = SplashScreen(root)
    root.update()

    def _on_mode(mode_key):
        """Called after user picks a mode from the selector."""
        root.deiconify()
        root.state("zoomed")
        try: root.attributes("-zoomed", True)
        except Exception: pass
        root.lift(); root.focus_force()

        def _back():
            """Destroy ALL root children (IV packs frames on root directly) then
            return to mode selector.  Previously only Toplevels were destroyed,
            which caused a second IV window to stack on top of the first."""
            for w in list(root.winfo_children()):
                try: w.destroy()
                except Exception: pass
            _show_selector()

        if mode_key == "IV":
            app = KeithleyApp(root)
            root.protocol("WM_DELETE_WINDOW", lambda: (
                app._force_stop() if app.is_running else None,
                _back()
            ))

        elif mode_key == "FET":
            root.withdraw()
            _app = KeithleyApp.__new__(KeithleyApp)
            _app.root        = root
            _app._dark_mode  = bool(_cfg.get("dark_mode"))
            _app.master_addr = _cfg.get("visa_address")
            _app.save_dir    = _cfg.get("save_dir")
            _app.rm          = pyvisa.ResourceManager()
            _THEME.update(DARK if _app._dark_mode else LIGHT)
            _sync_globals()
            fet = FETWindow(_app)
            fet.protocol("WM_DELETE_WINDOW", lambda: (
                fet._do_stop() if fet.is_running else None,
                fet.destroy(),
                _back()
            ))

        elif mode_key == "DMM":
            root.withdraw()
            dmm = DMMWindow(root)
            dmm.protocol("WM_DELETE_WINDOW", lambda: (
                dmm._close(),
                _back()
            ))

    def _show_selector():
        """Show the mode selector. Closing/Escaping it quits the app."""
        root.withdraw()
        sel = ModeSelector(root, _on_mode, on_quit=lambda: root.quit())

    def _init():
        root.after(500, lambda: (splash.finish(), _show_selector()))

    root.after(200, _init)
    root.mainloop()
